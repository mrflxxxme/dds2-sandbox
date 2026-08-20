# ruff: noqa: RUF001, RUF002, RUF003
"""Волна D v2: XLSX-выгрузка аналитики (CONTRACT-V2 §4).

Файл читается обратно openpyxl прямо в тесте: проверяем не «ручка ответила 200»,
а что в книге лежат те же цифры, что отдают ручки статистики.
"""

import io

from openpyxl import load_workbook

from backend.utils.time import utcnow
from tests.test_api_design_tasks import BASE, _mk_task, _msg, env  # noqa: F401

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def _download(client, headers, **params):
    resp = await client.get(f"{BASE}/stats/export.xlsx", params=params, headers=headers)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == XLSX_MIME
    return load_workbook(io.BytesIO(resp.content)), resp


def _cells(ws):
    """Все непустые значения листа строкой — для поиска подстрок."""
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


async def test_export_headers_and_sheets(client, env):
    await _mk_task(client, env.author.h)
    wb, resp = await _download(client, env.viewer.h)

    assert wb.sheetnames == ["Сводка", "Задачи"]
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith("attachment; filename*=UTF-8''")
    assert "design-tasks_" in disposition
    assert resp.headers["x-content-type-options"] == "nosniff"


async def test_export_filename_uses_resolved_window(client, env):
    """Без параметров в имя подставляются дефолтные границы, а не пустые строки."""
    _wb, resp = await _download(client, env.viewer.h)
    disposition = resp.headers["content-disposition"]
    today = utcnow().date().isoformat()
    assert today in disposition, disposition


async def test_task_sheet_matches_list(client, env):
    """Строки листа «Задачи» — те же задачи, что отдаёт GET /design-tasks."""
    made = [await _mk_task(client, env.author.h) for _ in range(3)]
    wb, _resp = await _download(client, env.viewer.h)

    ws = wb["Задачи"]
    numbers = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    for task in made:
        assert task["number"] in numbers


async def test_summary_matches_stats_endpoints(client, env):
    """Цифры «Сводки» совпадают с ответами ручек за то же окно."""
    task = await _mk_task(client, env.author.h)
    label = (
        await client.post(
            f"{BASE}/refs/labels", json={"name": "Срочно", "color": "red"}, headers=env.lead.h
        )
    ).json()
    await client.put(f"{BASE}/{task['id']}/labels", json={"label_ids": [label["id"]]}, headers=env.lead.h)

    by_attr = (await client.get(f"{BASE}/stats/by-attribute", headers=env.viewer.h)).json()
    expected_label_count = next(r["count"] for r in by_attr["labels"] if r["name"] == "Срочно")

    wb, _resp = await _download(client, env.viewer.h)
    cells = _cells(wb["Сводка"])
    assert "Метки" in cells
    idx = cells.index("Срочно")
    assert cells[idx + 1] == str(expected_label_count)


async def test_attribute_columns_and_multi_value_format(client, env):
    """Мультизначные ячейки — через запятую с пробелом; пустые остаются пустыми."""
    attr = (
        await client.post(
            f"{BASE}/refs/attributes", json={"name": "Площадки", "is_multi": True}, headers=env.lead.h
        )
    ).json()
    values = []
    for v in ("WB", "Ozon"):
        values.append(
            (
                await client.post(
                    f"{BASE}/refs/attributes/{attr['id']}/values",
                    json={"value": v},
                    headers=env.lead.h,
                )
            ).json()
        )

    task = await _mk_task(client, env.author.h)
    await client.put(
        f"{BASE}/{task['id']}/attributes",
        json={"value_ids": [values[0]["id"], values[1]["id"]]},
        headers=env.lead.h,
    )
    empty = await _mk_task(client, env.author.h)

    wb, _resp = await _download(client, env.viewer.h)
    ws = wb["Задачи"]
    head = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col = head.index("Площадки") + 1

    by_number = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
    filled = ws.cell(row=by_number[task["number"]], column=col).value
    assert filled in ("WB, Ozon", "Ozon, WB"), filled
    assert ws.cell(row=by_number[empty["number"]], column=col).value in (None, "")


async def test_truncation_is_announced(client, env, monkeypatch):
    """Cap строк проговаривается в «Сводке» — тихое усечение запрещено."""
    from backend.services.design import export_xlsx

    monkeypatch.setattr(export_xlsx, "MAX_TASK_ROWS", 2)
    for _ in range(4):
        await _mk_task(client, env.author.h)

    wb, _resp = await _download(client, env.viewer.h)
    assert wb["Задачи"].max_row == 3  # шапка + два ряда
    assert any("Показано 2 из" in c for c in _cells(wb["Сводка"]))


async def test_field_columns_cap_is_announced(client, env, monkeypatch):
    """Cap колонок-реквизитов тоже проговаривается."""
    from backend.services.design import export_xlsx

    monkeypatch.setattr(export_xlsx, "MAX_ATTRIBUTE_COLUMNS", 1)
    for name in ("Бренд", "Кабинет"):
        await client.post(f"{BASE}/refs/attributes", json={"name": name}, headers=env.lead.h)
    await _mk_task(client, env.author.h)

    wb, _resp = await _download(client, env.viewer.h)
    assert any("Показаны первые 1 полей-реквизитов из 2" in c for c in _cells(wb["Сводка"]))


async def test_export_window_guard(client, env):
    resp = await client.get(
        f"{BASE}/stats/export.xlsx", params={"date_to": "2026-08-01"}, headers=env.viewer.h
    )
    assert resp.status_code == 400, resp.text
    assert _msg(resp) == "Укажите обе границы диапазона"


async def test_export_is_project_scoped(client, env, db_session, auth_headers):
    """В выгрузке нет задач чужого проекта (Iron rule 1)."""
    from tests.design_helpers import make_task
    from tests.test_api_design_tasks import _create_project

    other = await _create_project(client, auth_headers, "Design D выгрузка чужая")
    alien = await make_task(db_session, other["id"], env.lead.id, number="ALIEN-1")
    mine = await _mk_task(client, env.author.h)

    wb, _resp = await _download(client, env.viewer.h)
    numbers = _cells(wb["Задачи"])
    assert mine["number"] in numbers
    assert alien.number not in numbers


async def test_viewer_can_download(client, env):
    """Р32: выгрузка доступна всем с page-ключом, включая viewer."""
    await _mk_task(client, env.author.h)
    wb, _resp = await _download(client, env.viewer.h)
    assert wb.sheetnames == ["Сводка", "Задачи"]
