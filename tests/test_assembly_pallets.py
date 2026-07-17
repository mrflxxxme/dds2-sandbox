"""Tests: раскладка по паллетам (pallet_manifest) + расчётный вес товаров.

Covers:
  compute_goods_weight — Σ(qty×weight), недостающие ШК, изоляция project_id.
  update_pallet_manifest — сохранение, сброс (null), строгий инвариант (409),
    ШК не из состава (400).
  build_pallet_layout_xlsx — валидные .xlsx для internal и wb.
"""

from decimal import Decimal
from io import BytesIO

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy import text

from backend.models.cost import CostOrder, CostOrderItem
from backend.schemas.assembly import (
    AssemblyItemCreate,
    AssemblyRequestCreate,
    BoxContent,
    PalletBox,
    PalletManifest,
)
from backend.services.assembly_service import (
    PalletManifestConflict,
    build_pallet_layout_xlsx,
    compute_goods_weight,
    create_assembly_request,
    get_assembly_request,
    update_pallet_manifest,
)

PROJECT_ID = 0
OTHER_PROJECT_ID = 0
BC1 = "TEST_BC_PAL_001"
BC2 = "TEST_BC_PAL_002"


@pytest_asyncio.fixture(autouse=True)
async def setup(db_session, project, other_project):
    """Свежий проект + FULFILLMENT-склад + номенклатура с весом/кратностью."""
    global PROJECT_ID, OTHER_PROJECT_ID
    PROJECT_ID = project.id
    OTHER_PROJECT_ID = other_project.id

    await db_session.execute(
        text(
            "INSERT INTO warehouses (project_id, name, warehouse_type, sort_order, is_active, is_deleted, created_at, updated_at) "
            "VALUES (:pid, 'PalTest FF', 'FULFILLMENT', 1, true, false, NOW(), NOW())"
        ),
        {"pid": PROJECT_ID},
    )
    # bc1: вес 0.5кг, кратность 10; bc2: без веса (недостающий), кратность 10.
    await db_session.execute(
        text(
            "INSERT INTO nomenclature (project_id, barcode, subject, article_seller, weight_kg, box_qty_override, use_box_multiplicity, updated_at) "
            "VALUES (:pid, :bc, 'Товар 1', 'ART-1', 0.5, 10, true, NOW())"
        ),
        {"pid": PROJECT_ID, "bc": BC1},
    )
    await db_session.execute(
        text(
            "INSERT INTO nomenclature (project_id, barcode, subject, article_seller, weight_kg, box_qty_override, use_box_multiplicity, updated_at) "
            "VALUES (:pid, :bc, 'Товар 2', 'ART-2', NULL, 10, true, NOW())"
        ),
        {"pid": PROJECT_ID, "bc": BC2},
    )
    await db_session.commit()

    # Сток на складе-источнике (create_assembly_request проверяет доступное кол-во).
    wh_id = await _wh_id(db_session)
    for bc in (BC1, BC2):
        nom_id = (
            await db_session.execute(
                text("SELECT id FROM nomenclature WHERE project_id = :pid AND barcode = :bc"),
                {"pid": PROJECT_ID, "bc": bc},
            )
        ).scalar()
        await db_session.execute(
            text(
                "INSERT INTO warehouse_stock (project_id, warehouse_id, nomenclature_id, barcode, quantity, in_transit, updated_at) "
                "VALUES (:pid, :wid, :nid, :bc, 500, 0, NOW())"
            ),
            {"pid": PROJECT_ID, "wid": wh_id, "nid": nom_id, "bc": bc},
        )
    await db_session.commit()
    yield


async def _wh_id(db_session) -> int:
    return (
        await db_session.execute(
            text("SELECT id FROM warehouses WHERE project_id = :pid AND name = 'PalTest FF' LIMIT 1"),
            {"pid": PROJECT_ID},
        )
    ).scalar()


async def _seed_machine_weight(db_session, order_no: str, barcode: str, weight: Decimal):
    """Строка машины (CostOrderItem) с весом для barcode — источник fallback-веса.

    Через ORM (а не raw SQL): применяет python-дефолты обязательных полей заказа
    (delivery_cost_cny, rate_*, country и т.п.), которых нет в raw INSERT."""
    db_session.add(CostOrder(project_id=PROJECT_ID, order_no=order_no))
    db_session.add(
        CostOrderItem(project_id=PROJECT_ID, order_no=order_no, barcode=barcode, qty=1, weight_kg=weight)
    )
    await db_session.commit()


async def _make_request(db_session, items: list[tuple[str, int]]):
    payload = AssemblyRequestCreate(
        warehouse_id=await _wh_id(db_session),
        wb_warehouse_name_manual="Электросталь",
        pallets_count=2,
        pallet_weight_kg=Decimal("100.00"),
        items=[AssemblyItemCreate(barcode=bc, quantity=q) for bc, q in items],
    )
    return await create_assembly_request(db_session, PROJECT_ID, payload)


# ─── compute_goods_weight ───────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_goods_weight_sum_and_missing(db_session):
    req = await _make_request(db_session, [(BC1, 20), (BC2, 3)])
    req = await get_assembly_request(db_session, PROJECT_ID, req.id)
    weight, missing = await compute_goods_weight(db_session, PROJECT_ID, req)
    assert weight == Decimal("10.000")  # 20 × 0.5, у bc2 веса нет
    assert missing == [BC2]


@pytest.mark.asyncio
async def test_goods_weight_none_when_no_weights(db_session):
    req = await _make_request(db_session, [(BC2, 5)])
    req = await get_assembly_request(db_session, PROJECT_ID, req.id)
    weight, missing = await compute_goods_weight(db_session, PROJECT_ID, req)
    assert weight is None
    assert missing == [BC2]


@pytest.mark.asyncio
async def test_goods_weight_machine_fallback(db_session):
    """BC2 без веса в справочнике, но есть ненулевой вес в машине (CostOrderItem)
    → берётся из машины, BC2 НЕ в missing."""
    await _seed_machine_weight(db_session, f"V-PAL-{PROJECT_ID}", BC2, Decimal("0.8"))

    req = await _make_request(db_session, [(BC2, 4)])
    req = await get_assembly_request(db_session, PROJECT_ID, req.id)
    weight, missing = await compute_goods_weight(db_session, PROJECT_ID, req)
    assert weight == Decimal("3.200")  # 4 × 0.8 из машины
    assert missing == []


@pytest.mark.asyncio
async def test_goods_weight_nomenclature_wins_over_machine(db_session):
    """Справочник приоритетнее машины: BC1 имеет вес 0.5 в справочнике; даже если
    в машине другой вес — берётся справочный."""
    await _seed_machine_weight(db_session, f"V-PALB-{PROJECT_ID}", BC1, Decimal("9.0"))

    req = await _make_request(db_session, [(BC1, 2)])
    req = await get_assembly_request(db_session, PROJECT_ID, req.id)
    weight, _ = await compute_goods_weight(db_session, PROJECT_ID, req)
    assert weight == Decimal("1.000")  # 2 × 0.5 (справочник), не 9.0 из машины


@pytest.mark.asyncio
async def test_goods_weight_project_isolation(db_session):
    """Вес bc1 задан только в НАШЕМ проекте — из другого не подтягивается и наоборот."""
    # Дать bc1 вес в другом проекте (без веса он там), затем убрать у нас →
    # compute для другого проекта не должен видеть наш вес.
    await db_session.execute(
        text(
            "INSERT INTO nomenclature (project_id, barcode, weight_kg, updated_at) VALUES (:pid, :bc, 9.0, NOW())"
        ),
        {"pid": OTHER_PROJECT_ID, "bc": BC1},
    )
    await db_session.commit()
    req = await _make_request(db_session, [(BC1, 2)])
    req = await get_assembly_request(db_session, PROJECT_ID, req.id)
    weight, _ = await compute_goods_weight(db_session, PROJECT_ID, req)
    assert weight == Decimal("1.000")  # 2 × 0.5 (наш проект), а не 9.0 из чужого


# ─── update_pallet_manifest ─────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_manifest_save_and_reset(db_session):
    req = await _make_request(db_session, [(BC1, 20)])
    manifest = PalletManifest(
        pallets=[PalletBox(pallet_no=1, boxes=[BoxContent(barcode=BC1, box_count=2, loose_units=0)])]
    )
    saved = await update_pallet_manifest(db_session, PROJECT_ID, req.id, manifest)
    assert saved.pallet_manifest is not None
    assert saved.pallet_manifest[0]["pallet_no"] == 1

    reset = await update_pallet_manifest(db_session, PROJECT_ID, req.id, None)
    assert reset.pallet_manifest is None


@pytest.mark.asyncio
async def test_manifest_loose_tail(db_session):
    """box_count×box_qty + loose == quantity (хвост-россыпь)."""
    req = await _make_request(db_session, [(BC1, 23)])  # 2 короба по 10 + 3 россыпью
    manifest = PalletManifest(
        pallets=[PalletBox(pallet_no=1, boxes=[BoxContent(barcode=BC1, box_count=2, loose_units=3)])]
    )
    saved = await update_pallet_manifest(db_session, PROJECT_ID, req.id, manifest)
    assert saved.pallet_manifest[0]["boxes"][0]["loose_units"] == 3


@pytest.mark.asyncio
async def test_manifest_invariant_violation_409(db_session):
    req = await _make_request(db_session, [(BC1, 20)])
    bad = PalletManifest(
        pallets=[PalletBox(pallet_no=1, boxes=[BoxContent(barcode=BC1, box_count=1, loose_units=0)])]
    )  # 1×10 = 10 ≠ 20
    with pytest.raises(PalletManifestConflict):
        await update_pallet_manifest(db_session, PROJECT_ID, req.id, bad)


@pytest.mark.asyncio
async def test_manifest_unknown_barcode_400(db_session):
    req = await _make_request(db_session, [(BC1, 20)])
    bad = PalletManifest(
        pallets=[PalletBox(pallet_no=1, boxes=[BoxContent(barcode="NOPE", box_count=2, loose_units=0)])]
    )
    with pytest.raises(ValueError):
        await update_pallet_manifest(db_session, PROJECT_ID, req.id, bad)


@pytest.mark.asyncio
async def test_manifest_project_isolation(db_session):
    """Чужой проект не может тронуть манифест нашей заявки (get вернёт None → ValueError)."""
    req = await _make_request(db_session, [(BC1, 20)])
    manifest = PalletManifest(
        pallets=[PalletBox(pallet_no=1, boxes=[BoxContent(barcode=BC1, box_count=2, loose_units=0)])]
    )
    with pytest.raises(ValueError):
        await update_pallet_manifest(db_session, OTHER_PROJECT_ID, req.id, manifest)


# ─── build_pallet_layout_xlsx ───────────────────────────────────────────────


@pytest.mark.asyncio
async def test_xlsx_internal_and_wb(db_session):
    req = await _make_request(db_session, [(BC1, 23)])  # 2 короба + 3 россыпь
    req = await get_assembly_request(db_session, PROJECT_ID, req.id)

    internal = await build_pallet_layout_xlsx(db_session, PROJECT_ID, req, fmt="internal")
    ws = load_workbook(BytesIO(internal)).active
    assert ws["A1"].value == "Паллета"  # заголовок
    assert any(row and row[0] == "ИТОГО" for row in ws.iter_rows(values_only=True))

    wb_bytes = await build_pallet_layout_xlsx(db_session, PROJECT_ID, req, fmt="wb")
    ws2 = load_workbook(BytesIO(wb_bytes)).active
    header = [c.value for c in ws2[1]]
    assert header == ["Баркод товара", "Кол-во товаров", "ШК короба", "Срок годности"]
    rows = [r for r in ws2.iter_rows(min_row=2, values_only=True) if r[0]]
    # 2 полных короба (по 10) + 1 россыпь (3) = 3 строки; ШК короба/срок — пусто.
    assert len(rows) == 3
    assert all(r[2] in (None, "") and r[3] in (None, "") for r in rows)
    assert sorted(r[1] for r in rows) == [3, 10, 10]


# ─── boxes_count + расчётный вес отгрузки + apply ───────────────────────────


@pytest.mark.asyncio
async def test_boxes_count_ceil(db_session):
    """Число коробов = Σ ⌈qty/кратность⌉ (неполный хвост = отдельный короб)."""
    from backend.services.assembly.weight import compute_boxes_count, resolve_box_qty_by_warehouse

    req = await _make_request(db_session, [(BC1, 20), (BC2, 3)])
    req = await get_assembly_request(db_session, PROJECT_ID, req.id)
    box_qty = await resolve_box_qty_by_warehouse(db_session, PROJECT_ID, {req.warehouse_id: [BC1, BC2]})
    # BC1 20/10 = 2 короба; BC2 ⌈3/10⌉ = 1 (неполный) → 3.
    assert compute_boxes_count(req, box_qty) == 3


def test_suggested_total_weight_adds_box_tara():
    from backend.services.assembly.weight import compute_suggested_total_weight

    # нетто 10.0 + вес_коробки 0.4 × 3 короба = 11.2.
    assert compute_suggested_total_weight(Decimal("10.0"), 3, Decimal("0.4")) == Decimal("11.200")
    # без веса коробки → чистое нетто.
    assert compute_suggested_total_weight(Decimal("10.0"), 3, None) == Decimal("10.000")
    # нет нетто → нечего предлагать.
    assert compute_suggested_total_weight(None, 3, Decimal("0.4")) is None


@pytest.mark.asyncio
async def test_apply_goods_weight_includes_box_tara(db_session):
    """apply_goods_weight: вес паллеты = (нетто + вес_коробки×коробов) / паллеты."""
    from backend.services.assembly_service import apply_goods_weight
    from backend.services.settings_service import set_box_weight_kg

    await set_box_weight_kg(db_session, PROJECT_ID, Decimal("0.4"))
    req = await _make_request(db_session, [(BC1, 20)])  # нетто 10, 2 короба, pallets=2
    req = await apply_goods_weight(db_session, PROJECT_ID, req.id)
    # total = 10 + 0.4×2 = 10.8; габаритов нет → pallets остаётся 2; 10.8/2 = 5.40.
    assert req.pallets_count == 2
    assert req.pallet_weight_kg == Decimal("5.40")


@pytest.mark.asyncio
async def test_apply_goods_weight_raises_without_weight(db_session):
    from backend.services.assembly_service import apply_goods_weight

    req = await _make_request(db_session, [(BC2, 5)])  # BC2 без веса
    with pytest.raises(ValueError):
        await apply_goods_weight(db_session, PROJECT_ID, req.id)


@pytest.mark.asyncio
async def test_apply_goods_weight_preserve_set_pallets(db_session, monkeypatch):
    """preserve_set_pallets=True (путь mark_ready): уже заданное pallets_count (>0)
    НЕ перетирается автооценкой геометрии — заявки из commit черновика несут паллеты,
    посчитанные фронтом ПО КЛАССАМ совместимости категорий (бэкендовый footprint
    классов не знает), а ручная правка оператора — тем более авторитет. Вес при этом
    делится на сохранённое число паллет. Дефолт (кнопка «Рассчитать вес») — прежнее
    поведение: геометрия перетирает."""
    from backend.services.assembly import pallets as pallets_mod

    async def fake_suggest(db, project_id, request, box_qty_by_wh_bc):
        return 3  # бэкендовая оценка расходится с фронтовой per-class раскладкой

    monkeypatch.setattr(pallets_mod, "_suggest_pallets_count", fake_suggest)

    req = await _make_request(db_session, [(BC1, 20)])  # pallets_count=2 задан фронтом
    kept = await pallets_mod.apply_goods_weight(
        db_session, PROJECT_ID, req.id, preserve_set_pallets=True
    )
    assert kept.pallets_count == 2  # значение фронта сохранено
    assert kept.pallet_weight_kg == Decimal("5.00")  # нетто 10 / 2 паллеты

    # Дефолтный путь — прежнее поведение: автооценка перетирает.
    recalced = await pallets_mod.apply_goods_weight(db_session, PROJECT_ID, req.id)
    assert recalced.pallets_count == 3


@pytest.mark.asyncio
async def test_apply_goods_weight_preserve_still_suggests_when_unset(db_session, monkeypatch):
    """preserve_set_pallets=True при pallets_count=0: автооценка ПО-ПРЕЖНЕМУ
    подставляется (нечего сохранять) — pre-dist заявки без паллет от фронта."""
    from backend.services.assembly import pallets as pallets_mod
    from backend.services.assembly_service import create_assembly_request as create_req

    async def fake_suggest(db, project_id, request, box_qty_by_wh_bc):
        return 4

    monkeypatch.setattr(pallets_mod, "_suggest_pallets_count", fake_suggest)

    payload = AssemblyRequestCreate(
        warehouse_id=await _wh_id(db_session),
        wb_warehouse_name_manual="Электросталь",
        pallets_count=0,
        pallet_weight_kg=Decimal("0"),
        items=[AssemblyItemCreate(barcode=BC1, quantity=20)],
    )
    req = await create_req(db_session, PROJECT_ID, payload)
    applied = await pallets_mod.apply_goods_weight(
        db_session, PROJECT_ID, req.id, preserve_set_pallets=True
    )
    assert applied.pallets_count == 4
    assert applied.pallet_weight_kg == Decimal("2.50")  # нетто 10 / 4 паллеты


@pytest.mark.asyncio
async def test_apply_goods_weight_bulk_partial(db_session):
    """Массовый авто-вес: заявка с весом применяется, без веса — в skipped."""
    from backend.services.assembly_service import apply_goods_weight_bulk

    r1 = await _make_request(db_session, [(BC1, 20)])  # есть вес
    r2 = await _make_request(db_session, [(BC2, 5)])  # нет веса
    applied, skipped = await apply_goods_weight_bulk(db_session, PROJECT_ID, [r1.id, r2.id])
    assert [r.id for r in applied] == [r1.id]
    assert len(skipped) == 1
    assert skipped[0]["id"] == r2.id and skipped[0]["number"] == r2.number


@pytest.mark.asyncio
async def test_resolve_box_qty_map_multi_matches_single(db_session):
    """Батч-резолв на неск. складов == per-складу resolve_box_qty_map."""
    from backend.services.box_multiplicity_service import resolve_box_qty_map, resolve_box_qty_map_multi

    wh_id = await _wh_id(db_session)
    single = await resolve_box_qty_map(db_session, PROJECT_ID, wh_id, [BC1, BC2])
    multi = await resolve_box_qty_map_multi(db_session, PROJECT_ID, {wh_id: [BC1, BC2]})
    assert multi.get((wh_id, BC1)) == single.get(BC1) == 10  # box_qty_override
    assert multi.get((wh_id, BC2)) == single.get(BC2) == 10
