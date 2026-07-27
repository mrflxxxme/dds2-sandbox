"""
Tests for the extended loan service: interest engine, extend (продление),
analytics (dashboard/by-lender/forecast), Excel import, lender portal scoping.
Uses real PG database for service-level tests; engine tests are pure.
"""

import io
import uuid
from datetime import date
from decimal import Decimal

import openpyxl
import pytest
from fastapi import HTTPException

from backend.schemas.counterparty import CounterpartyCreate
from backend.schemas.loan import LenderAccessCreate, LoanCreate, LoanExtend, LoanUpdate
from backend.services import lender_access_service, loan_analytics, loan_interest
from backend.services import lender_portal_service as lps
from backend.services.counterparty_service import CounterpartyService
from backend.services.loan_import import import_loans_from_xlsx
from backend.services.loan_service import LoanService

# ─── Helpers ──────────────────────────────────────────────────────────────────


def _inn() -> str:
    return str(int(uuid.uuid4().hex[:9], 16) % 9000000000 + 1000000000)


async def _create_project(client, auth_headers) -> int:
    resp = await client.post(
        "/api/v1/projects",
        json={"name": f"loanx_{uuid.uuid4().hex[:6]}"},
        headers=auth_headers,
    )
    return resp.json()["id"]


async def _create_cp(db_session, project_id: int, name: str = "Кредитор") -> int:
    svc = CounterpartyService(db_session)
    cp = await svc.create(
        CounterpartyCreate(inn=_inn(), name=f"{name} {uuid.uuid4().hex[:4]}", primary_type="OTHER"),
        project_id=project_id,
    )
    return cp.id


def _incoming(cp_id: int, contract: str, principal="1000000", rate="0.28") -> LoanCreate:
    return LoanCreate(
        counterparty_id=cp_id,
        direction="INCOMING",
        principal=Decimal(principal),
        currency="RUB",
        rate=Decimal(rate),
        contract_number=contract,
        contract_date=date(2025, 1, 1),
        start_date=date(2025, 1, 1),
        maturity_date=date(2025, 7, 1),
        entity_type="IP",
    )


# ─── Interest engine (pure) ───────────────────────────────────────────────────


def test_engine_current_period_accrual():
    calc = loan_interest.compute_loan(
        principal=Decimal("300000"),
        rate=Decimal("0.28"),
        start_date=date(2024, 3, 20),
        maturity_date=date(2024, 9, 19),
        status="ACTIVE",
        payments=[],
        as_of=date(2024, 4, 15),
    )
    # Период начисления — календарный 25→25, а НЕ годовщина выдачи займа:
    # текущий период 25.03 → 25.04, начислено с 25.03 по 15.04 = 21 день.
    assert calc.accrual_period_start == date(2024, 3, 25)
    assert calc.accrual_period_end == date(2024, 4, 25)
    assert calc.accrued_interest == Decimal("4832.88")
    # весь период 25.03 → 25.04 = 31 день — столько заплатят 25 апреля
    assert calc.interest_due_period == Decimal("7134.25")
    assert calc.monthly_interest == Decimal("7000.00")
    assert calc.daily_interest == Decimal("230.14")
    assert calc.next_interest_date == date(2024, 4, 25)
    # full-term projection: 300000 * 0.28 * 183/365
    assert calc.total_interest_projected == Decimal("42115.07")
    assert calc.is_overdue is False


def test_engine_accrual_period_boundaries():
    """Границы сетки 25→25 и первый период только что выданного займа."""
    assert loan_interest.accrual_period(date(2026, 7, 24)) == (date(2026, 6, 25), date(2026, 7, 25))
    assert loan_interest.accrual_period(date(2026, 7, 25)) == (date(2026, 7, 25), date(2026, 8, 25))
    # займ выдан 23.07 — в периоде 25.07→25.08 он живёт с 25-го, а не с выдачи
    calc = loan_interest.compute_loan(
        principal=Decimal("2000000"), rate=Decimal("0.255"),
        start_date=date(2026, 7, 23), maturity_date=date(2027, 1, 21),
        status="ACTIVE", payments=[], as_of=date(2026, 7, 27),
    )
    assert calc.accrued_interest == Decimal("2794.52")  # 2 дня, не 4


def test_engine_closed_mid_period_accrues_days_alive():
    """Погашенный в середине периода займ начисляет за прожитые дни (боевой Семериков)."""
    calc = loan_interest.compute_loan(
        principal=Decimal("3000000"), rate=Decimal("0.28"),
        start_date=date(2025, 10, 8), maturity_date=date(2026, 5, 27),
        status="CLOSED",
        payments=[loan_interest.PaymentLite("PRINCIPAL_REPAY", Decimal("3000000"), date(2026, 5, 27))],
        as_of=date(2026, 6, 24),  # внутри периода 25.05→25.06, платят его 25.06
    )
    assert calc.remaining_principal == Decimal("0.00")  # долга нет
    assert calc.accrued_interest == Decimal("4602.74")  # но 2 дня в периоде отработал
    assert calc.interest_due_period == Decimal("4602.74")  # столько и заплатят 25.06
    # займ, погашенный в прошлых периодах, в текущем не начисляет ничего
    old = loan_interest.compute_loan(
        principal=Decimal("2000000"), rate=Decimal("0.35"),
        start_date=date(2025, 1, 30), maturity_date=date(2025, 7, 31),
        status="CLOSED",
        payments=[loan_interest.PaymentLite("PRINCIPAL_REPAY", Decimal("2000000"), date(2025, 7, 31))],
        as_of=date(2026, 6, 25),
    )
    assert old.accrued_interest == Decimal("0.00")


def test_engine_overdue_and_closed():
    overdue = loan_interest.compute_loan(
        principal=Decimal("100000"), rate=Decimal("0.3"),
        start_date=date(2024, 1, 1), maturity_date=date(2024, 6, 1),
        status="ACTIVE", payments=[], as_of=date(2024, 8, 1),
    )
    assert overdue.is_overdue is True
    assert overdue.days_to_maturity < 0

    closed = loan_interest.compute_loan(
        principal=Decimal("100000"), rate=Decimal("0.3"),
        start_date=date(2024, 1, 1), maturity_date=date(2024, 6, 1),
        status="CLOSED", payments=[], as_of=date(2024, 5, 1),
    )
    assert closed.remaining_principal == Decimal("0.00")
    assert closed.accrued_interest == Decimal("0.00")
    assert closed.monthly_interest == Decimal("0.00")


def test_engine_partial_repayment_reduces_remaining():
    calc = loan_interest.compute_loan(
        principal=Decimal("1000000"), rate=Decimal("0.2"),
        start_date=date(2025, 1, 1), maturity_date=date(2026, 1, 1),
        status="ACTIVE",
        payments=[loan_interest.PaymentLite("PRINCIPAL_REPAY", Decimal("400000"), date(2025, 6, 1))],
        as_of=date(2025, 7, 1),
    )
    assert calc.remaining_principal == Decimal("600000.00")
    assert calc.principal_repaid == Decimal("400000.00")


# ─── extend (продление) ───────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_extend_creates_linked_successor(db_session, client, auth_headers):
    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id)
    svc = LoanService(db_session)
    loan = await svc.create(data=_incoming(cp_id, "06"), project_id=project_id)

    successor = await svc.extend(
        loan_id=loan.id,
        data=LoanExtend(new_rate=Decimal("0.31"), record_repayment=True),
        project_id=project_id,
    )
    assert successor.parent_loan_id == loan.id
    assert successor.status == "ACTIVE"
    assert successor.rate == Decimal("0.31")
    assert successor.contract_number == "06-01"
    assert successor.principal == Decimal("1000000.00")

    # old loan closed + has a repayment recorded
    detail = await svc.get_detail(loan_id=loan.id, project_id=project_id)
    assert detail.status == "CLOSED"
    assert detail.has_extension is True
    assert any(p.payment_type == "PRINCIPAL_REPAY" for p in detail.payments)


@pytest.mark.asyncio
async def test_extend_closed_loan_rejected(db_session, client, auth_headers):
    from fastapi import HTTPException

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id)
    svc = LoanService(db_session)
    loan = await svc.create(data=_incoming(cp_id, "07"), project_id=project_id)
    await svc.update(loan_id=loan.id, data=LoanUpdate(status="CLOSED"), project_id=project_id)
    with pytest.raises(HTTPException):
        await svc.extend(loan_id=loan.id, data=LoanExtend(), project_id=project_id)


# ─── analytics ────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_dashboard_and_by_lender(db_session, client, auth_headers):
    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id)
    svc = LoanService(db_session)
    await svc.create(data=_incoming(cp_id, "A1", principal="2000000", rate="0.30"), project_id=project_id)
    await svc.create(data=_incoming(cp_id, "A2", principal="1000000", rate="0.20"), project_id=project_id)

    dash = await loan_analytics.dashboard(db_session, project_id, date(2025, 3, 1))
    assert dash["kpis"]["active_count"] == 2
    assert Decimal(str(dash["kpis"]["total_outstanding"])) == Decimal("3000000")
    # weighted rate = (2M*0.3 + 1M*0.2)/3M = 0.2667
    assert abs(Decimal(str(dash["kpis"]["weighted_avg_rate"])) - Decimal("0.2667")) < Decimal("0.001")

    by = await loan_analytics.by_lender(db_session, project_id, date(2025, 3, 1))
    assert len(by["items"]) == 1
    assert by["items"][0]["active_count"] == 2
    assert Decimal(str(by["items"][0]["outstanding"])) == Decimal("3000000")

    fc = await loan_analytics.forecast(db_session, project_id, date(2025, 3, 1), 12)
    assert len(fc["months"]) == 12
    # principal due in 2025-07 (maturity)
    jul = next(m for m in fc["months"] if m["month"] == "2025-07")
    assert Decimal(str(jul["principal_due"])) == Decimal("3000000")


# ─── Excel import idempotency ─────────────────────────────────────────────────


def _build_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расчет процентов"
    ws.append(["Дата", "Тип", "Сумма", "Контрагент", "Номер договора", "Ставка", "От", "До", "Статус"])
    ws.append([date(2025, 1, 1), "Приход", 500000, "Иванов", "100", 0.28, date(2025, 1, 1), date(2025, 7, 1), "Активен"])
    ws.append([date(2025, 7, 1), "Возврат", 500000, "Иванов", "100", None, date(2025, 7, 1), None, "Не активен"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_import_idempotent(db_session, client, auth_headers):
    project_id = await _create_project(client, auth_headers)
    content = _build_xlsx()

    r1 = await import_loans_from_xlsx(db_session, project_id=project_id, content=content)
    assert r1.created_loans == 1
    assert r1.created_counterparties == 1
    assert r1.created_payments == 1

    r2 = await import_loans_from_xlsx(db_session, project_id=project_id, content=content)
    assert r2.created_loans == 0
    assert r2.updated_loans == 1
    assert r2.created_counterparties == 0
    assert r2.created_payments == 0  # repayment already present (idempotent)


def _build_rollover_xlsx() -> bytes:
    """Продление БЕЗ смены номера договора: старый транш гасится в день старта нового."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расчет процентов"
    ws.append(["Дата", "Тип", "Сумма", "Контрагент", "Номер договора", "Ставка", "От", "До", "Статус"])
    # старый транш
    ws.append([date(2025, 8, 18), "Приход", 5000000, "Еремин", "64", 0.315, date(2025, 8, 18), date(2026, 2, 16), "Не активен"])
    # возврат датирован ровно днём погашения старого = днём старта нового
    ws.append([date(2026, 2, 16), "Возврат", 5000000, "Еремин", "64", 0.315, date(2026, 2, 16), date(2026, 2, 16), "Не активен"])
    # преемник под ТЕМ ЖЕ номером
    ws.append([date(2026, 2, 16), "Приход", 5000000, "Еремин", "64", 0.265, date(2026, 2, 16), date(2026, 8, 17), "Активен"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_import_rollover_same_contract_closes_old_tranche(db_session, client, auth_headers):
    """Возврат закрывает транш, который в этот день ПОГАШАЕТСЯ, а не тот, что стартует."""
    from sqlalchemy import select

    from backend.models.loan import Loan, LoanPayment

    project_id = await _create_project(client, auth_headers)
    res = await import_loans_from_xlsx(db_session, project_id=project_id, content=_build_rollover_xlsx())
    assert res.created_loans == 2
    assert res.created_payments == 1

    loans = (
        (await db_session.execute(select(Loan).where(Loan.project_id == project_id).order_by(Loan.start_date)))
        .scalars()
        .all()
    )
    old, new = loans
    assert old.start_date == date(2025, 8, 18)
    assert new.start_date == date(2026, 2, 16)
    # старый — закрыт возвратом, новый — живой
    assert old.status == "CLOSED"
    assert new.status == "ACTIVE"

    payment = (await db_session.execute(select(LoanPayment).where(LoanPayment.loan_id.in_([old.id, new.id])))).scalars().one()
    assert payment.loan_id == old.id, "возврат повешен на новый транш вместо погашаемого старого"


def _build_overdue_xlsx() -> bytes:
    """Срок вышел, строки «Возврат» нет. Колонка «Статус» — формула по датам, ей верить нельзя."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расчет процентов"
    ws.append(["Дата", "Тип", "Сумма", "Контрагент", "Номер договора", "Ставка", "От", "До", "Статус"])
    # формула =IF(AND(TODAY()>=От; TODAY()<=До)…) на просроченном займе даёт «Не активен»
    ws.append([date(2025, 1, 1), "Приход", 3000000, "Петров", "200", 0.28, date(2025, 1, 1), date(2025, 7, 1), "Не активен"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_import_overdue_without_return_stays_active(db_session, client, auth_headers):
    """Просроченный займ без возврата остаётся ACTIVE — иначе долг исчезает из остатка."""
    from sqlalchemy import select

    from backend.models.loan import Loan

    project_id = await _create_project(client, auth_headers)
    res = await import_loans_from_xlsx(db_session, project_id=project_id, content=_build_overdue_xlsx())
    assert res.created_loans == 1
    assert res.created_payments == 0

    loan = (await db_session.execute(select(Loan).where(Loan.project_id == project_id))).scalars().one()
    assert loan.status == "ACTIVE", "займ без строки «Возврат» не должен закрываться по формуле «Статус»"
    assert loan.maturity_date == date(2025, 7, 1)


def _build_entity_sheet_xlsx() -> bytes:
    """Второй лист боевого реестра: «Контрагент» встречается ДВАЖДЫ.

    Первое вхождение — служебный столбец (один заёмщик), а физ/ИП и банк лежат
    рядом со вторым, вплотную к колонкам «куда».
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "База"
    ws.append(["Дата", "Тип", "Сумма", "Контрагент", "Номер договора", "Ставка", "От", "До", "Статус"])
    ws.append([date(2025, 1, 1), "Приход", 500000, "Прохоров", "10", 0.28, date(2025, 1, 1), date(2025, 7, 1), "Активен"])
    ws.append([date(2025, 1, 1), "Приход", 300000, "Карелин", "11", 0.28, date(2025, 1, 1), date(2025, 7, 1), "Активен"])

    ws2 = wb.create_sheet("база 2")
    # столбец A — «Контрагент»-обманка, реальные данные в C рядом с «куда»
    ws2.append(["Контрагент", "Сумма процентов", "Контрагент", "куда", "куда"])
    ws2.append(["Степан", 0, "Прохоров", "ип", None])
    ws2.append(["Степан", 0, "Карелин", "физ", "альфа"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_import_entity_map_picks_column_next_to_kuda(db_session, client, auth_headers):
    """Физ/ИП и банк берутся из «Контрагент», ближайшего к «куда», а не из первого."""
    from sqlalchemy import select

    from backend.models.loan import Loan

    project_id = await _create_project(client, auth_headers)
    res = await import_loans_from_xlsx(db_session, project_id=project_id, content=_build_entity_sheet_xlsx())
    assert res.created_loans == 2

    loans = (
        (await db_session.execute(select(Loan).where(Loan.project_id == project_id)))
        .scalars()
        .all()
    )
    by_contract = {loan.contract_number: loan for loan in loans}
    assert by_contract["10"].entity_type == "IP", "Прохоров должен приехать как ИП"
    assert by_contract["11"].entity_type == "PHYSICAL", "Карелин должен приехать как физлицо"
    assert by_contract["11"].lender_bank == "альфа"


# ─── Lender portal scoping ────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_lender_access_and_scoping(db_session, client, auth_headers):
    from fastapi import HTTPException

    from backend.lender_context import LenderContext
    from backend.models.auth import Project
    from backend.models.counterparty import Counterparty
    from sqlalchemy import select

    project_id = await _create_project(client, auth_headers)
    cp1 = await _create_cp(db_session, project_id, "Лендер1")
    cp2 = await _create_cp(db_session, project_id, "Лендер2")
    svc = LoanService(db_session)
    loan1 = await svc.create(data=_incoming(cp1, "L1"), project_id=project_id)
    loan2 = await svc.create(data=_incoming(cp2, "L2"), project_id=project_id)

    created = await lender_access_service.create_access(
        db_session, project_id, LenderAccessCreate(counterparty_id=cp1)
    )
    assert created.password
    assert created.user_id is not None

    # duplicate → 409
    with pytest.raises(HTTPException):
        await lender_access_service.create_access(db_session, project_id, LenderAccessCreate(counterparty_id=cp1))

    # build a lender context scoped to cp1 and verify isolation
    from backend.models.auth import User

    user = (await db_session.execute(select(User).where(User.id == created.user_id))).scalar_one()
    project = (await db_session.execute(select(Project).where(Project.id == project_id))).scalar_one()
    cp1_obj = (await db_session.execute(select(Counterparty).where(Counterparty.id == cp1))).scalar_one()
    ctx = LenderContext(
        user=user,
        projects={project_id: project},
        counterparties={project_id: [cp1]},
        counterparty_by_id={cp1: cp1_obj},
    )

    listing = await lps.list_loans(db_session, ctx)
    ids = {row.id for row in listing.items}
    assert loan1.id in ids
    assert loan2.id not in ids  # isolation

    # scoped fetch of other lender's loan → 404
    with pytest.raises(HTTPException):
        await lps.get_scoped_loan(db_session, ctx, loan2.id)

    # access list reflects the grant
    access = await lender_access_service.list_access(db_session, project_id)
    assert any(a.counterparty_id == cp1 for a in access)


@pytest.mark.asyncio
async def test_lender_context_drops_cross_project_counterparty(db_session, client, auth_headers):
    """get_lender_context must drop a counterparty that belongs to a DIFFERENT project."""
    import json as _json

    from backend.auth import hash_password
    from backend.lender_context import get_lender_context, lender_counterparties_key
    from backend.models.auth import ProjectMember, User
    from backend.models.refs import ProjectSetting

    proj_a = await _create_project(client, auth_headers)
    proj_b = await _create_project(client, auth_headers)
    cp_b = await _create_cp(db_session, proj_b, "ЧужойПроект")

    user = User(username=f"lx_{uuid.uuid4().hex[:8]}", password_hash=hash_password("x123456"), is_active=True, is_external=True)
    db_session.add(user)
    await db_session.flush()
    # lender membership in project A, but scope list points at a cp from project B
    db_session.add(ProjectMember(project_id=proj_a, user_id=user.id, role="lender"))
    db_session.add(ProjectSetting(project_id=proj_a, key=lender_counterparties_key(user.id), value=_json.dumps([cp_b])))
    await db_session.commit()

    ctx = await get_lender_context(user=user, db=db_session)
    # cp_b belongs to project B, not A → must be excluded from the resolved scope
    assert cp_b not in ctx.counterparty_ids
    assert ctx.counterparties.get(proj_a) == []


# ─── Карточка заёмщика, архив, разделение ИП/физ ─────────────────────────────


@pytest.mark.asyncio
async def test_lender_archived_when_no_active_loans(db_session, client, auth_headers):
    """Заёмщик уходит в архив, когда закрыт последний займ — без флага в БД."""
    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id, "Архивный")
    svc = LoanService(db_session)
    loan = await svc.create(data=_incoming(cp_id, "AR1"), project_id=project_id)

    by = await loan_analytics.by_lender(db_session, project_id, date(2025, 3, 1))
    row = next(i for i in by["items"] if i["counterparty_id"] == cp_id)
    assert row["is_archived"] is False
    assert by["archived_count"] == 0

    await svc.update(loan_id=loan.id, data=LoanUpdate(status="CLOSED"), project_id=project_id)
    by = await loan_analytics.by_lender(db_session, project_id, date(2025, 3, 1))
    row = next(i for i in by["items"] if i["counterparty_id"] == cp_id)
    assert row["is_archived"] is True, "нет активных займов → архив"
    assert by["archived_count"] == 1
    # сущность остаётся видна и в архиве (берётся с любого займа, не только активного)
    assert row["entity_type"] == "IP"


@pytest.mark.asyncio
async def test_by_lender_splits_interest_by_entity(db_session, client, auth_headers):
    """Проценты за период разделены по ИП и физлицам."""
    project_id = await _create_project(client, auth_headers)
    ip_cp = await _create_cp(db_session, project_id, "ИПшник")
    ph_cp = await _create_cp(db_session, project_id, "Физик")
    svc = LoanService(db_session)

    ip_loan = _incoming(ip_cp, "E1", principal="2000000", rate="0.24")
    await svc.create(data=ip_loan, project_id=project_id)
    ph_loan = _incoming(ph_cp, "E2", principal="1000000", rate="0.24")
    ph_loan.entity_type = "PHYSICAL"
    await svc.create(data=ph_loan, project_id=project_id)

    by = await loan_analytics.by_lender(db_session, project_id, date(2025, 3, 1))
    splits = {s["entity_type"]: s for s in by["by_entity"]}
    assert set(splits) == {"IP", "PHYSICAL"}
    assert Decimal(str(splits["IP"]["outstanding"])) == Decimal("2000000")
    assert Decimal(str(splits["PHYSICAL"]["outstanding"])) == Decimal("1000000")
    # тело ИП вдвое больше при той же ставке → и проценты вдвое
    ip_due = Decimal(str(splits["IP"]["interest_due_period"]))
    ph_due = Decimal(str(splits["PHYSICAL"]["interest_due_period"]))
    assert ip_due == ph_due * 2 > Decimal("0")
    assert by["accrual_period_end"] == "2025-03-25"


@pytest.mark.asyncio
async def test_lender_detail_history_and_periods(db_session, client, auth_headers):
    """Карточка заёмщика: история займов + проценты по периодам 25→25."""
    from backend.services.loan_service import get_lender_detail

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id, "Карточка")
    svc = LoanService(db_session)
    await svc.create(data=_incoming(cp_id, "H1", principal="1200000", rate="0.24"), project_id=project_id)
    await svc.create(data=_incoming(cp_id, "H2", principal="600000", rate="0.24"), project_id=project_id)

    # окно шире срока займов (они в 2025-м) — иначе периодов в выборке не будет
    detail = await get_lender_detail(
        db_session, counterparty_id=cp_id, project_id=project_id, months_back=36
    )
    assert detail.name.startswith("Карточка")
    assert detail.total_count == 2
    assert detail.active_count == 2
    assert detail.outstanding == Decimal("1800000.00")
    assert detail.principal_total == Decimal("1800000")
    assert detail.weighted_avg_rate == Decimal("0.2400")
    assert detail.is_archived is False
    assert len(detail.loans) == 2
    assert detail.periods, "должна быть разбивка по периодам"
    # метки периодов — даты выплаты (25-е число), по возрастанию
    assert all(p.period_end.day == 25 for p in detail.periods)
    assert [p.period_end for p in detail.periods] == sorted(p.period_end for p in detail.periods)


@pytest.mark.asyncio
async def test_lender_detail_404_for_foreign_project(db_session, client, auth_headers):
    """Чужой контрагент не отдаётся — изоляция по project_id."""
    from fastapi import HTTPException

    from backend.services.loan_service import get_lender_detail

    p1 = await _create_project(client, auth_headers)
    p2 = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, p1, "Чужой")
    with pytest.raises(HTTPException) as exc:
        await get_lender_detail(db_session, counterparty_id=cp_id, project_id=p2)
    assert exc.value.status_code == 404


@pytest.mark.asyncio
async def test_extend_term_months_preset(db_session, client, auth_headers):
    """Пресет продления: 3 месяца от даты продления, с клампом дня месяца."""
    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id)
    svc = LoanService(db_session)
    loan = await svc.create(data=_incoming(cp_id, "T1"), project_id=project_id)

    successor = await svc.extend(
        loan_id=loan.id,
        data=LoanExtend(term_months=3, new_start_date=date(2025, 11, 30), new_rate=Decimal("0.25")),
        project_id=project_id,
    )
    assert successor.start_date == date(2025, 11, 30)
    assert successor.maturity_date == date(2026, 2, 28), "30 ноября + 3 мес → 28 февраля"
    assert successor.rate == Decimal("0.25")


# ─── Возврат займа ───────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_repay_full_closes_loan(db_session, client, auth_headers):
    """Полный возврат закрывает займ и обнуляет остаток."""
    from backend.schemas.loan import LoanRepay

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id)
    svc = LoanService(db_session)
    loan = await svc.create(data=_incoming(cp_id, "R1"), project_id=project_id)

    await svc.repay(loan_id=loan.id, data=LoanRepay(paid_at=date(2025, 7, 1)), project_id=project_id)

    detail = await svc.get_detail(loan_id=loan.id, project_id=project_id)
    assert detail.status == "CLOSED"
    assert detail.remaining_principal == Decimal("0.00")
    repays = [p for p in detail.payments if p.payment_type == "PRINCIPAL_REPAY"]
    assert len(repays) == 1
    assert repays[0].amount == Decimal("1000000.00")
    assert repays[0].paid_at == date(2025, 7, 1)


@pytest.mark.asyncio
async def test_repay_partial_keeps_loan_active(db_session, client, auth_headers):
    """Частичный возврат оставляет займ живым и уменьшает тело под проценты."""
    from backend.schemas.loan import LoanRepay

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id)
    svc = LoanService(db_session)
    loan = await svc.create(data=_incoming(cp_id, "R2"), project_id=project_id)

    await svc.repay(
        loan_id=loan.id,
        data=LoanRepay(amount=Decimal("400000"), paid_at=date(2025, 3, 1)),
        project_id=project_id,
    )
    detail = await svc.get_detail(loan_id=loan.id, project_id=project_id)
    assert detail.status == "ACTIVE"
    assert detail.remaining_principal == Decimal("600000.00")


@pytest.mark.asyncio
async def test_repay_rejects_overpay_and_double(db_session, client, auth_headers):
    """Возврат больше остатка и повторный возврат по закрытому — 400."""
    from fastapi import HTTPException

    from backend.schemas.loan import LoanRepay

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id)
    svc = LoanService(db_session)
    loan = await svc.create(data=_incoming(cp_id, "R3"), project_id=project_id)

    with pytest.raises(HTTPException) as exc:
        await svc.repay(
            loan_id=loan.id, data=LoanRepay(amount=Decimal("2000000")), project_id=project_id
        )
    assert exc.value.status_code == 400

    await svc.repay(loan_id=loan.id, data=LoanRepay(), project_id=project_id)
    with pytest.raises(HTTPException) as exc2:
        await svc.repay(loan_id=loan.id, data=LoanRepay(), project_id=project_id)
    assert exc2.value.status_code == 400


# ─── «Занёс за всё время»: продления не двоят деньги ─────────────────────────


@pytest.mark.asyncio
async def test_new_money_excludes_extensions(db_session, client, auth_headers):
    """Цепочка продлений — одни и те же деньги, а не сумма всех тел."""
    from backend.services.loan_service import get_lender_detail

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id, "Продлятор")
    svc = LoanService(db_session)
    loan = await svc.create(data=_incoming(cp_id, "X1", principal="2000000"), project_id=project_id)

    # два продления подряд: тело всё то же
    s1 = await svc.extend(loan_id=loan.id, data=LoanExtend(), project_id=project_id)
    await svc.extend(loan_id=s1.id, data=LoanExtend(), project_id=project_id)

    d = await get_lender_detail(db_session, counterparty_id=cp_id, project_id=project_id, months_back=36)
    assert d.total_count == 3, "три записи в истории"
    assert d.outstanding == Decimal("2000000.00")
    assert d.principal_total == Decimal("2000000"), "занёс 2 млн, а не 6"
    assert d.principal_repaid == Decimal("0"), "деньги не возвращались — только переоформлялись"


@pytest.mark.asyncio
async def test_new_money_counts_topup_and_real_return(db_session, client, auth_headers):
    """Доплата при продлении — новые деньги; закрытая цепочка — реальный возврат."""
    from backend.schemas.loan import LoanRepay
    from backend.services.loan_service import get_lender_detail

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id, "Доплата")
    svc = LoanService(db_session)
    loan = await svc.create(data=_incoming(cp_id, "Y1", principal="1000000"), project_id=project_id)

    # продление с доплатой: было 1 млн, стало 1.5 млн → новых денег 0.5 млн
    successor = await svc.extend(
        loan_id=loan.id,
        data=LoanExtend(principal=Decimal("1500000")),
        project_id=project_id,
    )
    d = await get_lender_detail(db_session, counterparty_id=cp_id, project_id=project_id, months_back=36)
    assert d.principal_total == Decimal("1500000")
    assert d.principal_repaid == Decimal("0")

    # теперь реально вернули всё
    await svc.repay(loan_id=successor.id, data=LoanRepay(), project_id=project_id)
    d2 = await get_lender_detail(db_session, counterparty_id=cp_id, project_id=project_id, months_back=36)
    assert d2.outstanding == Decimal("0")
    assert d2.principal_total == Decimal("1500000")
    assert d2.principal_repaid == Decimal("1500000")
    assert d2.is_archived is True


# ─── Зависшие займы ──────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_stuck_loans_only_unresolved(db_session, client, auth_headers):
    """Зависшие = срок вышел и ничего не сделали; возврат и продление убирают из списка."""
    from backend.schemas.loan import LoanRepay
    from backend.services.loan_service import list_stuck_loans

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id, "Зависший")
    svc = LoanService(db_session)
    stuck = await svc.create(data=_incoming(cp_id, "S1"), project_id=project_id)  # до 2025-07-01
    repaid = await svc.create(data=_incoming(cp_id, "S2"), project_id=project_id)
    extended = await svc.create(data=_incoming(cp_id, "S3"), project_id=project_id)
    future = await svc.create(data=_incoming(cp_id, "S4"), project_id=project_id)
    await svc.update(
        loan_id=future.id, data=LoanUpdate(maturity_date=date(2099, 1, 1)), project_id=project_id
    )

    res = await list_stuck_loans(db_session, project_id=project_id)
    assert {i.loan.id for i in res.items} == {stuck.id, repaid.id, extended.id}
    assert res.count == 3

    # вернули один, продлили другой — остаётся только необработанный
    await svc.repay(loan_id=repaid.id, data=LoanRepay(), project_id=project_id)
    # продлеваем в будущее: иначе преемник сам окажется просроченным (и это верно)
    await svc.extend(
        loan_id=extended.id,
        data=LoanExtend(new_start_date=date.today(), term_months=6),
        project_id=project_id,
    )

    res2 = await list_stuck_loans(db_session, project_id=project_id)
    assert [i.loan.id for i in res2.items] == [stuck.id]
    assert res2.total_amount == Decimal("1000000.00")
    item = res2.items[0]
    assert item.days_overdue > 0
    assert item.accrued_since_maturity > 0, "после срока проценты продолжают капать"


@pytest.mark.asyncio
async def test_stuck_loans_isolated_by_project(db_session, client, auth_headers):
    """Чужие зависшие займы не видны."""
    from backend.services.loan_service import list_stuck_loans

    p1 = await _create_project(client, auth_headers)
    p2 = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, p1, "Изоляция")
    svc = LoanService(db_session)
    await svc.create(data=_incoming(cp_id, "Z1"), project_id=p1)

    assert (await list_stuck_loans(db_session, project_id=p1)).count == 1
    assert (await list_stuck_loans(db_session, project_id=p2)).count == 0


@pytest.mark.asyncio
async def test_by_lender_historical_period(db_session, client, auth_headers):
    """Исторический срез: «сколько было к выплате 25.07» считается за тот период."""
    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id, "Историк")
    svc = LoanService(db_session)
    # займ 25.02 → 25.08, ставка 24% на 1 200 000 ₽
    data = _incoming(cp_id, "P1", principal="1200000", rate="0.24")
    data.start_date = date(2025, 2, 25)
    data.maturity_date = date(2025, 8, 25)
    await svc.create(data=data, project_id=project_id)

    # период 25.06 → 25.07 = 30 дней: 1 200 000 × 24% × 30/365
    by = await loan_analytics.by_lender(
        db_session, project_id, date.today(), period_end=date(2025, 7, 25)
    )
    assert by["accrual_period_start"] == "2025-06-25"
    assert by["accrual_period_end"] == "2025-07-25"
    row = next(i for i in by["items"] if i["counterparty_id"] == cp_id)
    assert Decimal(str(row["interest_due_period"])) == Decimal("23671.23")
    # период давно закрыт → «начислено» равно «к выплате», без сдвига на день
    assert Decimal(str(row["accrued_interest"])) == Decimal("23671.23")


# ─── Кредитная линия: выборки + плавающая ставка ─────────────────────────────


def test_credit_line_matches_bank_statement():
    """ВКЛ ВБ Банка: 6 траншей, смена ключевой внутри месяца — сверка с выпиской."""
    from backend.services.loan_interest import RatePeriod, accrued_in_window

    # Выборки по договору РЛ-21/26: лимит 101 млн выбран за месяц
    tranches = [
        (date(2026, 5, 25), 70_000_000), (date(2026, 6, 9), 12_000_000),
        (date(2026, 6, 16), 6_000_000), (date(2026, 6, 19), 4_000_000),
        (date(2026, 6, 23), 5_000_000), (date(2026, 6, 25), 4_000_000),
    ]
    pays = [loan_interest.PaymentLite("DISBURSEMENT", Decimal(a), d) for d, a in tranches]
    # Ставка = ключевая ЦБ + 5 %: 14.5+5 до 21.06, с 22.06 ключевая упала до 14.25
    rates = [
        RatePeriod(date(2026, 5, 25), Decimal("0.195")),
        RatePeriod(date(2026, 6, 22), Decimal("0.1925")),
    ]

    def accrue(a: date, b: date) -> Decimal:
        return accrued_in_window(
            principal=Decimal("0"), rate=None, start_date=date(2026, 5, 25),
            payments=pays, win_start=a, win_end=b, rate_periods=rates,
        )

    # Факт из выписки ВБ Банка — до копейки
    assert accrue(date(2026, 5, 25), date(2026, 5, 31)) == Decimal("224383.56")
    assert accrue(date(2026, 5, 31), date(2026, 6, 30)) == Decimal("1348267.12")


def test_credit_line_body_grows_on_drawdown():
    """Тело линии растёт на выборках и падает на погашениях."""
    from backend.services.loan_interest import accrued_in_window

    pays = [
        loan_interest.PaymentLite("DISBURSEMENT", Decimal("1000000"), date(2026, 1, 1)),
        loan_interest.PaymentLite("DISBURSEMENT", Decimal("1000000"), date(2026, 1, 11)),
        loan_interest.PaymentLite("PRINCIPAL_REPAY", Decimal("2000000"), date(2026, 1, 21)),
    ]
    kw = dict(principal=Decimal("0"), rate=Decimal("0.365"), start_date=date(2026, 1, 1), payments=pays)
    # 10 дней по 1 млн + 10 дней по 2 млн = 30 млн·дней под 36.5 % → 1000 ₽/день на млн
    assert accrued_in_window(**kw, win_start=date(2026, 1, 1), win_end=date(2026, 1, 21)) == Decimal("30000.00")
    # после полного погашения начисление прекращается
    assert accrued_in_window(**kw, win_start=date(2026, 1, 21), win_end=date(2026, 1, 31)) == Decimal("0.00")


def test_rate_change_applies_from_its_own_day():
    """День смены ставки считается уже по новой — иначе расхождение с банком."""
    from backend.services.loan_interest import RatePeriod, accrued_in_window

    kw = dict(
        principal=Decimal("36500000"), rate=Decimal("0.10"),
        start_date=date(2026, 3, 1), payments=[],
        rate_periods=[RatePeriod(date(2026, 3, 3), Decimal("0.20"))],
    )
    # тело 36.5 млн: 10 % = 10 000 ₽/день, 20 % = 20 000 ₽/день
    # дни 2 и 3 марта: 2-е по старой (10 000), 3-е уже по новой (20 000)
    assert accrued_in_window(**kw, win_start=date(2026, 3, 1), win_end=date(2026, 3, 3)) == Decimal("30000.00")


@pytest.mark.asyncio
async def test_credit_line_service_end_to_end(db_session, client, auth_headers):
    """Линия целиком: выборки, лимит, плавающая ставка, сальдо движений."""
    from backend.models.loan import Loan
    from backend.schemas.loan import CreditLineDraw, LoanRatePeriodIn
    from backend.services.loan_service import draw_credit_line, get_credit_line, set_rate_period

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id, "ВБ Банк")
    line = Loan(
        project_id=project_id, counterparty_id=cp_id, direction="INCOMING",
        principal=Decimal("0.01"), currency="RUB", contract_number="РЛ-21/26",
        contract_date=date(2026, 5, 25), start_date=date(2026, 5, 25),
        maturity_date=date(2027, 5, 25), status="ACTIVE",
        loan_kind="CREDIT_LINE", accrual_kind="CALENDAR_MONTH",
        credit_limit=Decimal("101000000"),
    )
    db_session.add(line)
    await db_session.commit()
    await db_session.refresh(line)

    # Ставка = ключевая ЦБ + 5 %, ключевая упала с 22.06
    for vf, rate, base in ((date(2026, 5, 25), "0.195", "0.145"), (date(2026, 6, 22), "0.1925", "0.1425")):
        await set_rate_period(
            db_session, loan_id=line.id, project_id=project_id,
            data=LoanRatePeriodIn(valid_from=vf, rate=Decimal(rate), base_rate=Decimal(base), spread=Decimal("0.05")),
        )
    for when, amount in (
        (date(2026, 5, 25), 70_000_000), (date(2026, 6, 9), 12_000_000),
        (date(2026, 6, 16), 6_000_000), (date(2026, 6, 19), 4_000_000),
        (date(2026, 6, 23), 5_000_000), (date(2026, 6, 25), 4_000_000),
    ):
        await draw_credit_line(
            db_session, loan_id=line.id, project_id=project_id,
            data=CreditLineDraw(amount=Decimal(amount), drawn_at=when),
        )

    cl = await get_credit_line(db_session, loan_id=line.id, project_id=project_id)
    assert cl.drawn == Decimal("101000000")
    assert cl.available == Decimal("0")
    assert cl.utilization == Decimal("1.0000")
    assert cl.current_rate == Decimal("0.1925")
    assert len(cl.movements) == 6
    assert cl.movements[-1].balance_after == Decimal("101000000")

    # Лимит выбран целиком — следующая выборка отбивается
    with pytest.raises(HTTPException) as exc:
        await draw_credit_line(
            db_session, loan_id=line.id, project_id=project_id,
            data=CreditLineDraw(amount=Decimal("1"), drawn_at=date(2026, 7, 1)),
        )
    assert exc.value.status_code == 400


@pytest.mark.asyncio
async def test_credit_line_rejects_draw_on_term_loan(db_session, client, auth_headers):
    """По обычному займу выборки запрещены — только по линии."""
    from backend.schemas.loan import CreditLineDraw
    from backend.services.loan_service import draw_credit_line, get_credit_line

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id)
    svc = LoanService(db_session)
    loan = await svc.create(data=_incoming(cp_id, "T-1"), project_id=project_id)

    with pytest.raises(HTTPException) as exc:
        await draw_credit_line(
            db_session, loan_id=loan.id, project_id=project_id,
            data=CreditLineDraw(amount=Decimal("1000"), drawn_at=date(2025, 2, 1)),
        )
    assert exc.value.status_code == 400
    with pytest.raises(HTTPException):
        await get_credit_line(db_session, loan_id=loan.id, project_id=project_id)


# ─── График платежей (аннуитет Симпл Финанса) ────────────────────────────────

# Приложение № 1 к договору E/2026/0087/01: 30 млн под 21 %, аннуитет
# 3 021 911,33 ₽. Строка 1 — комиссия за выдачу 4,25 %, процентов за первые
# 7 дней договор не берёт.
SIMPLE_SCHEDULE = [
    # seq, начало, конец, дата платежа, дней, тело, проценты, платёж, остаток
    (1, date(2026, 3, 18), date(2026, 3, 25), date(2026, 3, 25), 7, "0", "1275000.00", "1275000.00", "30000000.00"),
    (2, date(2026, 3, 26), date(2026, 4, 27), date(2026, 4, 27), 33, "2452322.29", "569589.04", "3021911.33", "27547677.71"),
    (3, date(2026, 4, 28), date(2026, 5, 25), date(2026, 5, 25), 28, "2578129.56", "443781.77", "3021911.33", "24969548.15"),
    (4, date(2026, 5, 26), date(2026, 6, 25), date(2026, 6, 25), 31, "2576564.05", "445347.28", "3021911.33", "22392984.10"),
    (5, date(2026, 6, 26), date(2026, 7, 27), date(2026, 7, 27), 32, "2609635.02", "412276.31", "3021911.33", "19783349.08"),
]


def _simple_rows(limit: int | None = None):
    from backend.schemas.loan import LoanScheduleRowIn

    rows = []
    for seq, ps, pe, due, days, principal, interest, total, balance in SIMPLE_SCHEDULE[:limit]:
        rows.append(
            LoanScheduleRowIn(
                seq=seq, period_start=ps, period_end=pe, due_date=due, days=days,
                principal_due=Decimal(principal), interest_due=Decimal(interest),
                payment_total=Decimal(total), balance_after=Decimal(balance),
                is_fee=(seq == 1),
            )
        )
    return rows


def test_simple_finance_schedule_matches_contract_formula():
    """Проценты каждой строки = остаток × 21 % × дни / 365 — сверка с приложением № 1."""
    balance = Decimal("30000000.00")
    for seq, _ps, _pe, _due, days, principal, interest, total, after in SIMPLE_SCHEDULE:
        if seq == 1:
            # Первые 7 дней процентов нет — за них взята комиссия за выдачу 4,25 %
            assert Decimal(interest) == (Decimal("30000000") * Decimal("0.0425")).quantize(Decimal("0.01"))
        else:
            calc = (balance * Decimal("0.21") * Decimal(days) / Decimal("365")).quantize(Decimal("0.01"))
            assert abs(calc - Decimal(interest)) <= Decimal("0.02"), f"строка {seq}"
            assert Decimal(principal) + Decimal(interest) == Decimal(total)
        balance = Decimal(after)


def test_schedule_interest_splits_into_calendar_months():
    """Строка графика раскладывается по календарным месяцам пропорционально дням."""
    from backend.services.loan_interest import ScheduleRowLite, schedule_interest_in_window

    rows = [
        ScheduleRowLite(date(2026, 3, 18), date(2026, 3, 25), date(2026, 3, 25), 7,
                        Decimal("1275000.00"), is_fee=True),
        ScheduleRowLite(date(2026, 3, 26), date(2026, 4, 27), date(2026, 4, 27), 33,
                        Decimal("569589.04")),
    ]
    # Март: строка-комиссия не в счёт, от второй строки — дни 26…31 = 6 из 33
    march = schedule_interest_in_window(rows, win_start=date(2026, 2, 28), win_end=date(2026, 3, 31))
    assert march == (Decimal("569589.04") * 6 / 33).quantize(Decimal("0.01"))
    # Апрель — оставшиеся 27 дней; вместе дают ровно сумму строки
    april = schedule_interest_in_window(rows, win_start=date(2026, 3, 31), win_end=date(2026, 4, 30))
    assert march + april == Decimal("569589.04")


def test_one_off_fee_amortizes_without_losing_kopecks():
    """Разовая комиссия размазывается по сроку и в сумме даёт ровно себя."""
    from backend.services.loan_interest import FeeLite, fee_in_window

    fee = FeeLite(
        amount=Decimal("252500.00"), charged_at=date(2026, 5, 25), amortize=True,
        amortize_from=date(2026, 5, 25), amortize_to=date(2027, 5, 25),
    )
    total = Decimal("0")
    cur = date(2026, 5, 1)
    for _ in range(14):
        nxt = loan_interest.add_months(cur, 1)
        total += fee_in_window(
            fee,
            win_start=date.fromordinal(cur.toordinal() - 1),
            win_end=date.fromordinal(nxt.toordinal() - 1),
        )
        cur = nxt
    assert total == Decimal("252500.00")
    # 366 дней срока → в мае живут 7 дней (25…31)
    may = fee_in_window(fee, win_start=date(2026, 4, 30), win_end=date(2026, 5, 31))
    assert may == (Decimal("252500.00") * 7 / 366).quantize(Decimal("0.01"))
    # Без амортизации — вся сумма падает в день уплаты
    lump = FeeLite(amount=Decimal("252500.00"), charged_at=date(2026, 5, 25), amortize=False)
    assert fee_in_window(lump, win_start=date(2026, 4, 30), win_end=date(2026, 5, 31)) == Decimal("252500.00")
    assert fee_in_window(lump, win_start=date(2026, 5, 31), win_end=date(2026, 6, 30)) == Decimal("0")


@pytest.mark.asyncio
async def test_schedule_reconciles_plan_with_fact(db_session, client, auth_headers):
    """График + фактические платежи: PAID / OVERDUE, ближайший платёж, отклонения."""
    from backend.models.loan import LoanPayment
    from backend.schemas.loan import LoanScheduleReplace
    from backend.services import loan_schedule

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id, "СимплФинанс")
    svc = LoanService(db_session)
    loan = await svc.create(
        data=LoanCreate(
            counterparty_id=cp_id, direction="INCOMING", principal=Decimal("30000000"),
            currency="RUB", rate=Decimal("0.21"), contract_number="E/2026/0087/01",
            contract_date=date(2026, 3, 18), start_date=date(2026, 3, 18),
            maturity_date=date(2027, 2, 25),
        ),
        project_id=project_id,
    )
    await loan_schedule.replace_schedule(
        db_session, loan_id=loan.id, project_id=project_id,
        data=LoanScheduleReplace(rows=_simple_rows()),
    )

    # Факт из выписок: комиссия, два аннуитета (второй с опозданием на день + пени)
    for kind, amount, when in (
        ("COMMISSION", "1275000.00", date(2026, 3, 25)),
        ("PRINCIPAL_REPAY", "2452322.29", date(2026, 4, 24)),
        ("INTEREST_PAY", "569589.04", date(2026, 4, 24)),
        ("PRINCIPAL_REPAY", "2578129.56", date(2026, 5, 26)),
        ("INTEREST_PAY", "443781.77", date(2026, 5, 26)),
        ("PENALTY", "1483.31", date(2026, 5, 26)),
    ):
        db_session.add(LoanPayment(
            loan_id=loan.id, payment_type=kind, amount=Decimal(amount),
            currency="RUB", paid_at=when,
        ))
    await db_session.commit()

    res = await loan_schedule.get_schedule(
        db_session, loan_id=loan.id, project_id=project_id, as_of=date(2026, 7, 1)
    )
    by_seq = {r.seq: r for r in res.rows}
    assert by_seq[1].state == "PAID"
    assert by_seq[1].fee_paid == Decimal("1275000.00")
    # Заплатили раньше срока — отрицательное опоздание, план закрыт
    assert by_seq[2].state == "PAID" and by_seq[2].days_late == -3
    # Пени сверх плана видны как переплата строки
    assert by_seq[3].state == "PAID" and by_seq[3].delta == Decimal("1483.31")
    assert by_seq[3].days_late == 1
    # Платёж 25.06 не сделан — строка просрочена, а 27.07 ждёт оплаты
    assert by_seq[4].state == "OVERDUE"
    assert by_seq[5].state == "DUE"
    assert res.next_due_date == date(2026, 7, 27)
    assert res.overdue_count == 1
    assert res.overdue_amount == Decimal("3021911.33")
    assert res.total_payment == Decimal("13362645.32")  # 1 275 000 + 4 × 3 021 911,33
    # «Итого процентов» повторяет договор: там комиссия за выдачу тоже напечатана
    # в колонке процентов (1 870 994,40 + 1 275 000).
    assert res.total_interest == Decimal("3145994.40")
    # А стоимость денег строку-комиссию не берёт — она приходит через LoanFee,
    # иначе расход задвоится (комиссии по этому займу ещё не заведены → 0).
    assert res.total_cost == Decimal("1870994.40")


@pytest.mark.asyncio
async def test_month_cost_uses_schedule_and_amortized_fee(db_session, client, auth_headers):
    """Стоимость денег по месяцам: проценты из графика + доля разовой комиссии."""
    from backend.schemas.loan import LoanFeeIn, LoanScheduleReplace
    from backend.services import loan_schedule

    project_id = await _create_project(client, auth_headers)
    cp_id = await _create_cp(db_session, project_id, "СимплФинанс")
    svc = LoanService(db_session)
    loan = await svc.create(
        data=LoanCreate(
            counterparty_id=cp_id, direction="INCOMING", principal=Decimal("30000000"),
            currency="RUB", rate=Decimal("0.21"), contract_number="E/2026/0087/02",
            contract_date=date(2026, 3, 18), start_date=date(2026, 3, 18),
            maturity_date=date(2027, 2, 25),
        ),
        project_id=project_id,
    )
    await loan_schedule.replace_schedule(
        db_session, loan_id=loan.id, project_id=project_id,
        data=LoanScheduleReplace(rows=_simple_rows()),
    )
    await loan_schedule.add_fee(
        db_session, loan_id=loan.id, project_id=project_id,
        data=LoanFeeIn(
            fee_kind="ORIGINATION", amount=Decimal("1275000.00"),
            charged_at=date(2026, 3, 25), amortize=True,
        ),
    )

    months = await loan_analytics.accrual_by_month(
        db_session, project_id, date(2026, 6, 30), months=4
    )
    by_month = {m["month"]: m for m in months["items"]}
    # Март: 6 дней из строки 2 (26…31.03), простая ставка дала бы 13 дней
    assert Decimal(by_month["2026-03"]["interest"]) == (Decimal("569589.04") * 6 / 33).quantize(Decimal("0.01"))
    # Комиссия НЕ падает целиком в март: она куплена на весь срок
    assert Decimal(by_month["2026-03"]["fee"]) < Decimal("150000")
    assert Decimal(by_month["2026-04"]["fee"]) > Decimal("0")
    # Полная стоимость месяца = проценты + комиссия
    for m in months["items"]:
        assert Decimal(m["total"]) == Decimal(m["interest"]) + Decimal(m["fee"])
