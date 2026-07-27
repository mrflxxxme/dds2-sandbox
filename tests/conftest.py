"""
Shared fixtures for DDS tests.
"""

import io

import pandas as pd
import pytest
import pytest_asyncio

# Load API test fixtures (client, auth_headers, db_session)
pytest_plugins = ["tests.conftest_api"]


# Real (production-equivalent) projects we never delete in cleanup.
# id=4 — Default seed; id=15 — main user project (вяткин).
# 991/992 — изолированные проекты test_wb_returns: держим, чтобы cleanup не
# удалял их между параллельными воркерами (иначе FK на wb_goods_returns.project_id).
# 1010888 — demo-проект (slug 'demo', WB-зеркала/база знаний): его удаление
# каскадом сносит integration_keys с WB-ключом (инцидент 2026-07-23 — cleanup
# после pytest уничтожил demo-проект и ключ, восстановление из бэкапа невозможно,
# т.к. дамп старше проекта).
_KEEP_PROJECT_IDS = (4, 15, 991, 992, 1010888)


def pytest_sessionstart(session):
    """Reserve the hardcoded test project ids (4/15/991/992) by advancing
    projects_id_seq well past them.

    test_wb_returns hardcodes project_id 991/992 (see its module docstring),
    relying on the sequence already sitting above them — true on the long-lived
    dev DB, but a FRESH CI DB starts the sequence at 1. There it climbs through
    991/992 mid-run and hands those ids to another suite's auto-created project;
    test_wb_returns' teardown then runs `DELETE FROM warehouses WHERE project_id
    IN (991,992)` and trips an FK on that suite's warehouse children — flaky, the
    child table varied run to run (stock_movements, stock_adjustments, …).
    Bumping the sequence makes the "ids are reserved" assumption hold in CI too.

    Idempotent (GREATEST never lowers it) — safe under xdist per-worker re-entry.
    Sync psycopg2 + DATABASE_URL_SYNC: runs before the asyncio loop exists.
    """
    import os

    if os.environ.get("PYTEST_NO_CLEANUP"):
        return
    sync_url = os.environ.get("DATABASE_URL_SYNC")
    if not sync_url:
        return
    try:
        import psycopg2

        conn = psycopg2.connect(sync_url)
        conn.autocommit = True
        cur = conn.cursor()
        cur.execute(
            "SELECT setval('projects_id_seq', GREATEST("
            "(SELECT last_value FROM projects_id_seq),"
            "(SELECT COALESCE(MAX(id), 0) FROM projects),"
            "1000000))"
        )
        cur.close()
        conn.close()
    except Exception as e:
        print(f"\n[pytest_sessionstart] projects_id_seq bump skipped: {type(e).__name__}: {e}")


def pytest_sessionfinish(session, exitstatus):
    """Safety-net cleanup: drop test projects/users left by tests that don't teardown.

    Some tests create projects via POST /api/v1/projects without explicit teardown
    (test_api_isolation, test_loans_api, test_api_*, etc — 20+ files). Without
    this hook the local DB accumulated 74k test projects in 2 months (incident
    2026-05-04). Per-test teardown would be ideal but requires touching every
    test; this is a pragmatic backstop.

    Uses sync psycopg2 + DATABASE_URL_SYNC to avoid asyncio cleanup hassle in
    pytest_sessionfinish (runs after the asyncio event loop is closed).
    """
    import os

    if os.environ.get("PYTEST_NO_CLEANUP"):
        return
    # Под xdist pytest_sessionfinish вызывается на КАЖДОМ воркере при его
    # завершении. Если один воркер финиширует раньше — он удалит все не-KEEP
    # проекты, пока другой воркер ещё гоняет тесты → FK-гонка (напр.
    # wb_goods_returns.project_id violates fkey). Чистим ТОЛЬКО на контроллере
    # (workerinput есть только у xdist-воркеров) — один раз, после всех воркеров.
    if hasattr(session.config, "workerinput"):
        return
    sync_url = os.environ.get("DATABASE_URL_SYNC")
    if not sync_url:
        return

    try:
        import psycopg2

        conn = psycopg2.connect(sync_url)
        conn.autocommit = False
        cur = conn.cursor()
        cur.execute("SET LOCAL session_replication_role = 'replica'")
        cur.execute(
            "SELECT DISTINCT tc.table_name, kcu.column_name "
            "FROM information_schema.referential_constraints rc "
            "JOIN information_schema.table_constraints tc ON rc.constraint_name = tc.constraint_name "
            "JOIN information_schema.key_column_usage kcu ON rc.constraint_name = kcu.constraint_name "
            "JOIN information_schema.constraint_column_usage ccu ON rc.unique_constraint_name = ccu.constraint_name "
            "WHERE ccu.table_name = 'projects' AND ccu.column_name = 'id'"
        )
        for table, col in cur.fetchall():
            cur.execute(
                f'DELETE FROM "{table}" WHERE "{col}" NOT IN %s',  # noqa: S608 — table/col come from information_schema
                (_KEEP_PROJECT_IDS,),
            )
        # «Общие» заявки на оплату (project_id IS NULL) НЕ ловятся фильтром выше
        # (NULL NOT IN (...) → NULL, не TRUE) → копились бы вечно. Чистим явно,
        # child-first (session_replication_role='replica' отключает ON DELETE CASCADE).
        cur.execute(
            "DELETE FROM payment_request_document WHERE payment_request_id IN "
            "(SELECT id FROM payment_request WHERE project_id IS NULL)"
        )
        cur.execute(
            "DELETE FROM payment_request_event WHERE payment_request_id IN "
            "(SELECT id FROM payment_request WHERE project_id IS NULL)"
        )
        cur.execute("DELETE FROM payment_request WHERE project_id IS NULL")
        cur.execute("DELETE FROM projects WHERE id NOT IN %s", (_KEEP_PROJECT_IDS,))
        cur.execute(
            "DELETE FROM users WHERE id NOT IN (SELECT DISTINCT owner_id FROM projects WHERE owner_id IS NOT NULL) "
            "AND id NOT IN (SELECT DISTINCT user_id FROM project_members WHERE user_id IS NOT NULL)"
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        # Don't break the test session on cleanup failure — just log
        print(f"\n[pytest_sessionfinish] cleanup skipped: {type(e).__name__}: {e}")


@pytest.fixture
def vtb_rub_excel() -> bytes:
    """Generate a minimal VTB RUB bank statement Excel file."""
    data = {
        "Дата": ["01.02.2024", "02.02.2024", "ИТОГО"],
        "Номер": ["1", "2", ""],
        "Вид операции": ["Оплата", "Поступление", ""],
        "Контрагент": ["ООО Ромашка", "ИП Иванов", ""],
        "ИНН контрагента": ["7701234567", "770987654321", ""],
        "БИК банка контрагента": ["044525225", "044525226", ""],
        "Счет контрагента": ["40702810000000001234", "40702810000000005678", ""],
        "Дебет RUR": [50000.00, 0, 50000.00],
        "Кредит RUR": [0, 100000.00, 100000.00],
        "Назначение": [
            "Оплата по договору №1 за товар",
            "Оплата по INVOICE INV-2024-001 за услуги",
            "",
        ],
    }
    df = pd.DataFrame(data)
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    return buf.getvalue()


@pytest.fixture
def vtb_rub_excel_bad_dates() -> bytes:
    """VTB RUB statement with some unparseable dates."""
    data = {
        "Дата": ["01.02.2024", "NOT_A_DATE", "03.02.2024"],
        "Номер": ["1", "2", "3"],
        "Вид операции": ["Оплата", "Оплата", "Оплата"],
        "Контрагент": ["ООО A", "ООО B", "ООО C"],
        "ИНН контрагента": ["111", "222", "333"],
        "БИК банка контрагента": ["000", "000", "000"],
        "Счет контрагента": ["40702810000000001111", "40702810000000002222", "40702810000000003333"],
        "Дебет RUR": [1000, 2000, 3000],
        "Кредит RUR": [0, 0, 0],
        "Назначение": ["p1", "p2", "p3"],
    }
    df = pd.DataFrame(data)
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    return buf.getvalue()


@pytest.fixture
def vtb_cny_excel() -> bytes:
    """Generate a minimal VTB CNY bank statement."""
    data = {
        "Дата": ["15.03.2024"],
        "Номер": ["1"],
        "Вид операции": ["Перевод"],
        "Контрагент": ["SUPPLIER CO LTD"],
        "ИНН контрагента": [""],
        "БИК банка контрагента": [""],
        "Счет контрагента": ["CN123456789"],
        "Дебет CNY": [5000.00],
        "Кредит CNY": [0],
        "Дебет RUR": [100000.00],
        "Кредит RUR": [0],
        "Назначение": ["Payment ANNEX #5 according to contract"],
    }
    df = pd.DataFrame(data)
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    return buf.getvalue()


@pytest.fixture
def wb_excel() -> bytes:
    """Generate a minimal WB bank statement."""
    data = {
        "Документ": ["ПП-1", "Наименование", "ПП-2"],
        "Дата операции": ["10.01.2024", "Корреспондент", "11.01.2024"],
        "Корреспондент": ["ООО Вайлдберриз", "Наименование", "ООО Логистик"],
        "ИНН": ["7721546864", "", "1234567890"],
        "КПП": ["", "", ""],
        "Счет": ["40702810000000009999", "", "40702810000000008888"],
        "БИК": ["", "", ""],
        "Вх.остаток": [0, 0, 0],
        "Оборот Дт": [0, 0, 15000],
        "Оборот Кт": [500000, 0, 0],
        "Назначение платежа": [
            "Перечисление по реестру Wildberries",
            "",
            "Комиссия за банковское обслуживание",
        ],
    }
    df = pd.DataFrame(data)
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    return buf.getvalue()


@pytest.fixture
def vtb_multi_excel() -> bytes:
    """Generate a multi-sheet VTB bank statement (3 accounts in one file).

    Mimics real VTB format: each sheet = one account.
    Sheet name = account number.
    Row 1: ВЫПИСКА
    Row 2: Номер счета: <acc>, Валюта: <currency info>
    Row 3-6: metadata
    Row 7: column headers
    Row 8+: data
    """
    import openpyxl

    wb = openpyxl.Workbook()

    # Sheet 1: RUB deposit (will be skipped — usually not needed, but parsed)
    ws1 = wb.active
    ws1.title = "42102810316110029573"
    ws1.append(["ВЫПИСКА"])
    ws1.append(["Номер счета:", "42102810316110029573", "Валюта:", "Валюта 643, Российский рубль"])
    ws1.append(["Начальная дата: ", "01.02.2026", "Конечная дата: ", "06.03.2026"])
    ws1.append(["Входящий остаток RUB:", 0, "Исходящий остаток RUB:", 0])
    ws1.append([None])
    ws1.append([None])
    ws1.append(
        [
            "Дата",
            "Номер",
            "Вид операции",
            "Контрагент",
            "ИНН контрагента",
            "БИК банка контрагента",
            "Счет контрагента",
            "Дебет, RUR",
            "Кредит, RUR",
            "Назначение",
        ]
    )
    ws1.append(
        [
            "01.02.2026",
            "100",
            "01",
            "ООО Тест",
            "7701234567",
            "044525411",
            "40702810000000001234",
            50000,
            0,
            "Размещение депозита",
        ]
    )

    # Sheet 2: RUB main
    ws2 = wb.create_sheet("40702810400810052145")
    ws2.append(["ВЫПИСКА"])
    ws2.append(["Номер счета:", "40702810400810052145", "Валюта:", "Валюта 643, Российский рубль"])
    ws2.append(["Начальная дата: ", "01.02.2026", "Конечная дата: ", "06.03.2026"])
    ws2.append(["Входящий остаток RUB:", 44922.18, "Исходящий остаток RUB:", 310355.96])
    ws2.append([None])
    ws2.append([None])
    ws2.append(
        [
            "Дата",
            "Номер",
            "Вид операции",
            "Контрагент",
            "ИНН контрагента",
            "БИК банка контрагента",
            "Счет контрагента",
            "Дебет, RUR",
            "Кредит, RUR",
            "Назначение",
        ]
    )
    ws2.append(
        [
            "03.02.2026",
            "407",
            "01",
            "ООО Партнер",
            "1800027275",
            "044525450",
            "40702810800000001893",
            0,
            6900000,
            "Перевод собственных средств",
        ]
    )
    ws2.append(
        [
            "04.02.2026",
            "537451",
            "01",
            "ООО Поставщик",
            "7704217370",
            "044525593",
            "40702810901300010687",
            0,
            286032.91,
            "Оплата по договору",
        ]
    )

    # Sheet 3: CNY
    ws3 = wb.create_sheet("40702156916110000346")
    ws3.append(["ВЫПИСКА"])
    ws3.append(["Номер счета:", "40702156916110000346", "Валюта:", "Валюта 156, Китайский юань"])
    ws3.append(["Начальная дата: ", "01.02.2026", "Конечная дата: ", "06.03.2026"])
    ws3.append(["Входящий остаток CNY:", 621.99, "Исходящий остаток CNY:", 58222.79])
    ws3.append(["                                     RUR:", 6760.35, "RUR:", 660747.15])
    ws3.append([None])
    ws3.append(
        [
            "Дата",
            "Номер",
            "Вид операции",
            "Контрагент",
            "ИНН контрагента",
            "БИК банка контрагента",
            "Счет контрагента",
            "Дебет, CNY",
            "Кредит, CNY",
            "Дебет, RUR",
            "Кредит, RUR",
            "Назначение",
        ]
    )
    ws3.append(
        [
            "05.02.2026",
            "18",
            "01",
            "БАНК ВТБ (ПАО)",
            "7702070139",
            "044525411",
            "30301156700810000001",
            106000,
            0,
            1171459,
            0,
            "Перевод CNY",
        ]
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# === Universal project fixtures for AI agents ===
# These reduce test setup boilerplate — use them instead of creating projects from scratch


async def _create_project_with_owner(db_session, project_name: str, slug_prefix: str, user_prefix: str):
    """Create a project + owner via raw SQL. Cleanup handled by pytest_sessionfinish."""
    import uuid

    from sqlalchemy import text

    suffix = uuid.uuid4().hex[:8]
    slug = f"{slug_prefix}-{suffix}"
    result = await db_session.execute(
        text(
            "INSERT INTO users (username, email, password_hash, is_active, created_at) "
            "VALUES (:u, :e, :p, true, NOW()) RETURNING id"
        ),
        {"u": f"{user_prefix}_{suffix}", "e": f"{user_prefix}_{suffix}@test.com", "p": "nohash"},
    )
    user_id = result.scalar()
    result = await db_session.execute(
        text("INSERT INTO projects (name, slug, owner_id, created_at) VALUES (:n, :s, :o, NOW()) RETURNING id"),
        {"n": project_name, "s": slug, "o": user_id},
    )
    project_id = result.scalar()
    await db_session.commit()
    return project_id, user_id, slug


@pytest_asyncio.fixture
async def project(db_session):
    """Ready project for any test. Bulk cleanup happens in pytest_sessionfinish."""
    project_id, user_id, slug = await _create_project_with_owner(
        db_session, "Test Project", "test-project", "proj_user"
    )
    from types import SimpleNamespace

    return SimpleNamespace(id=project_id, name="Test Project", slug=slug, owner_id=user_id)


@pytest_asyncio.fixture
async def other_project(db_session):
    """Another project for multi-tenancy isolation tests. Bulk cleanup in pytest_sessionfinish."""
    project_id, user_id, slug = await _create_project_with_owner(
        db_session, "Other Project", "other-project", "other_user"
    )
    from types import SimpleNamespace

    return SimpleNamespace(id=project_id, name="Other Project", slug=slug, owner_id=user_id)


@pytest.fixture
def wb_multi_xml_xls() -> bytes:
    """Generate a multi-account WB statement in XML SpreadsheetML format.

    Two accounts in one sheet, separated by ИТОГО row.
    Format: XML SpreadsheetML (.xls)
    """
    xml = """<?xml version="1.0" encoding="utf-8"?>
<?mso-application progid="Excel.Sheet"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
 <Worksheet ss:Name="Выписки">
  <Table>
   <Row><Cell><Data ss:Type="String">Выписка по счету 40702810500001001752 RUR (Пассивный)</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String">ООО "ВБ Банк" БИК 044525450</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String">ООО "ТЕСТ"</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell><Cell><Data ss:Type="String">За период с 01.02.2026 по 06.03.2026</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell><Cell><Data ss:Type="String">Дата формирования информации 07.03.2026</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell><Cell><Data ss:Type="String">Дата последней операции 06.03.2026</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell><Cell><Data ss:Type="String">Входящий остаток:</Data></Cell><Cell><Data ss:Type="String">0.00 на 01.02.2026</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell></Row>
   <Row>
    <Cell><Data ss:Type="String">Документ</Data></Cell>
    <Cell><Data ss:Type="String">Дата операции</Data></Cell>
    <Cell><Data ss:Type="String">Корреспондент</Data></Cell>
    <Cell><Data ss:Type="String">Вх.остаток</Data></Cell>
    <Cell><Data ss:Type="String">Оборот Дт</Data></Cell>
    <Cell><Data ss:Type="String">Оборот Кт</Data></Cell>
    <Cell><Data ss:Type="String">Назначение платежа</Data></Cell>
   </Row>
   <Row>
    <Cell><Data ss:Type="String">Наименование</Data></Cell>
    <Cell><Data ss:Type="String">ИНН</Data></Cell>
    <Cell><Data ss:Type="String">КПП</Data></Cell>
    <Cell><Data ss:Type="String">Счет</Data></Cell>
    <Cell><Data ss:Type="String">БИК</Data></Cell>
   </Row>
   <Row>
    <Cell><Data ss:Type="String">Пор. № 100</Data></Cell>
    <Cell><Data ss:Type="String">01.02.2026</Data></Cell>
    <Cell><Data ss:Type="String">РВБ ООО</Data></Cell>
    <Cell><Data ss:Type="String">9714053621</Data></Cell>
    <Cell><Data ss:Type="String">507401001</Data></Cell>
    <Cell><Data ss:Type="String">40702810825620001712</Data></Cell>
    <Cell><Data ss:Type="String">044525411</Data></Cell>
    <Cell><Data ss:Type="Number">0</Data></Cell>
    <Cell><Data ss:Type="String"></Data></Cell>
    <Cell><Data ss:Type="Number">500000</Data></Cell>
    <Cell><Data ss:Type="String">Оплата за товар</Data></Cell>
   </Row>
   <Row>
    <Cell><Data ss:Type="String">Пор. № 101</Data></Cell>
    <Cell><Data ss:Type="String">01.02.2026</Data></Cell>
    <Cell><Data ss:Type="String">ООО "ТЕСТ"</Data></Cell>
    <Cell><Data ss:Type="String">1800027275</Data></Cell>
    <Cell><Data ss:Type="String">370001001</Data></Cell>
    <Cell><Data ss:Type="String">40702810800000001893</Data></Cell>
    <Cell><Data ss:Type="String">044525450</Data></Cell>
    <Cell><Data ss:Type="Number">500000</Data></Cell>
    <Cell><Data ss:Type="Number">495000</Data></Cell>
    <Cell><Data ss:Type="String"></Data></Cell>
    <Cell><Data ss:Type="String">Перевод денежных средств</Data></Cell>
   </Row>
   <Row><Cell><Data ss:Type="String">ИТОГО:</Data></Cell><Cell><Data ss:Type="Number">495000</Data></Cell><Cell><Data ss:Type="Number">500000</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell><Cell><Data ss:Type="String">Исходящий остаток:</Data></Cell><Cell><Data ss:Type="String">5000.00 на 06.03.2026</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell></Row>
   <Row><Cell><Data ss:Type="String">Выписка по счету 40702810800000001893 RUR (Пассивный)</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String">ООО "ВБ Банк" БИК 044525450</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String">ООО "ТЕСТ"</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell><Cell><Data ss:Type="String">За период с 01.02.2026 по 06.03.2026</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell><Cell><Data ss:Type="String">Дата формирования информации 07.03.2026</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell><Cell><Data ss:Type="String">Дата последней операции 06.03.2026</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell><Cell><Data ss:Type="String">Входящий остаток:</Data></Cell><Cell><Data ss:Type="String">100000.00 на 01.02.2026</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell></Row>
   <Row>
    <Cell><Data ss:Type="String">Документ</Data></Cell>
    <Cell><Data ss:Type="String">Дата операции</Data></Cell>
    <Cell><Data ss:Type="String">Корреспондент</Data></Cell>
    <Cell><Data ss:Type="String">Вх.остаток</Data></Cell>
    <Cell><Data ss:Type="String">Оборот Дт</Data></Cell>
    <Cell><Data ss:Type="String">Оборот Кт</Data></Cell>
    <Cell><Data ss:Type="String">Назначение платежа</Data></Cell>
   </Row>
   <Row>
    <Cell><Data ss:Type="String">Наименование</Data></Cell>
    <Cell><Data ss:Type="String">ИНН</Data></Cell>
    <Cell><Data ss:Type="String">КПП</Data></Cell>
    <Cell><Data ss:Type="String">Счет</Data></Cell>
    <Cell><Data ss:Type="String">БИК</Data></Cell>
   </Row>
   <Row>
    <Cell><Data ss:Type="String">Пор. № 200</Data></Cell>
    <Cell><Data ss:Type="String">03.02.2026</Data></Cell>
    <Cell><Data ss:Type="String">ООО "ТЕСТ"</Data></Cell>
    <Cell><Data ss:Type="String">1800027275</Data></Cell>
    <Cell><Data ss:Type="String">370001001</Data></Cell>
    <Cell><Data ss:Type="String">40702810500001001752</Data></Cell>
    <Cell><Data ss:Type="String">044525450</Data></Cell>
    <Cell><Data ss:Type="Number">100000</Data></Cell>
    <Cell><Data ss:Type="String"></Data></Cell>
    <Cell><Data ss:Type="Number">75000</Data></Cell>
    <Cell><Data ss:Type="String">Перевод входящий</Data></Cell>
   </Row>
   <Row><Cell><Data ss:Type="String">ИТОГО:</Data></Cell><Cell><Data ss:Type="Number">0</Data></Cell><Cell><Data ss:Type="Number">75000</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell><Cell><Data ss:Type="String">Исходящий остаток:</Data></Cell><Cell><Data ss:Type="String">175000.00 на 06.03.2026</Data></Cell></Row>
   <Row><Cell><Data ss:Type="String"></Data></Cell></Row>
  </Table>
 </Worksheet>
</Workbook>"""
    return xml.encode("utf-8")
