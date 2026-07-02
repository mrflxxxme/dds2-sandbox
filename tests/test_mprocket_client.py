# ruff: noqa: RUF001, RUF002, RUF003
"""
Tests for the Mprocket («Нитропак») client: XLS stock parser + auth heuristics.

No network: pure parsing/heuristic functions + the write-gate stub. The stock
.xlsx is built in-memory with openpyxl (same lib the parser uses).
"""

import io

import httpx
import openpyxl
import pytest

from backend.integrations.mprocket_client import (
    MprocketApiError,
    _extract_csrf,
    _login_ok,
    _normalize_host,
    _redact,
    _validate_business_id,
    parse_request_rows,
    parse_stock_xlsx,
)


def _stock_xlsx(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        ["Название", "Размер", "Цвет", "ШК WB", "ШК OZON", "Количество на складе", "Количество в отгрузке", "Сумма"]
    )
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── XLS stock parser ─────────────────────────────────────────────────────────


def test_parse_stock_xlsx_maps_columns_by_header():
    content = _stock_xlsx(
        [
            ["Ковер", "200х300", "серый", "2049757911663", "OZN1", "120", "30", "1000"],
            ["Накидка", "", "бежевый", "2043160630135", "", "5", "0", ""],
        ]
    )
    rows = parse_stock_xlsx(content)
    assert len(rows) == 2
    first = rows[0]
    assert first["barcode"] == "2049757911663"
    assert first["qty_stock"] == 120
    assert first["qty_shipping"] == 30
    assert first["name"] == "Ковер · 200х300 · серый"
    # пустые Размер пропускаются в склейке имени
    assert rows[1]["name"] == "Накидка · бежевый"


def test_parse_stock_xlsx_skips_rows_without_barcode():
    content = _stock_xlsx(
        [
            ["Без ШК", "", "", "", "", "10", "0", ""],
            ["С ШК", "", "", "2043160630135", "", "7", "1", ""],
        ]
    )
    rows = parse_stock_xlsx(content)
    assert [r["barcode"] for r in rows] == ["2043160630135"]


def test_parse_stock_xlsx_column_reorder_robust():
    # колонки переставлены — матчинг по заголовку, не позиции
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ШК WB", "Количество на складе", "Название"])
    ws.append(["2049757911663", "42", "Швабра"])
    buf = io.BytesIO()
    wb.save(buf)
    rows = parse_stock_xlsx(buf.getvalue())
    assert rows == [{"barcode": "2049757911663", "name": "Швабра", "qty_stock": 42, "qty_shipping": 0}]


def test_parse_stock_xlsx_missing_barcode_column_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Название", "Количество на складе"])
    ws.append(["Ковер", "10"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(MprocketApiError):
        parse_stock_xlsx(buf.getvalue())


def test_parse_stock_xlsx_garbage_raises():
    with pytest.raises(MprocketApiError):
        parse_stock_xlsx(b"not an xlsx")


# ─── CSRF / host / business_id / redact ──────────────────────────────────────


def test_extract_csrf():
    html = '<form><input type="hidden" name="_token" value="zrh3DtGm0s0PIiMu"></form>'
    assert _extract_csrf(html) == "zrh3DtGm0s0PIiMu"


def test_extract_csrf_absent():
    assert _extract_csrf("<form></form>") is None


def test_normalize_host_default_and_valid():
    assert _normalize_host("") == "https://seller.mprocket.ru"
    assert _normalize_host("seller.mprocket.ru") == "https://seller.mprocket.ru"
    assert _normalize_host("https://seller.mprocket.ru/") == "https://seller.mprocket.ru"


@pytest.mark.parametrize(
    "bad",
    [
        "http://seller.mprocket.ru",  # not https
        "https://evil.com",  # wrong host
        "https://seller.mprocket.ru.evil.com",  # suffix spoof
        "https://seller.mprocket.ru:8080",  # port
        "https://user:pw@seller.mprocket.ru",  # userinfo
        "https://seller.mprocket.ru/path",  # path
    ],
)
def test_normalize_host_rejects(bad: str):
    with pytest.raises(MprocketApiError):
        _normalize_host(bad)


def test_validate_business_id():
    assert _validate_business_id("306") == "306"
    assert _validate_business_id(306) == "306"
    with pytest.raises(MprocketApiError):
        _validate_business_id("abc")
    with pytest.raises(MprocketApiError):
        _validate_business_id("")


def test_login_ok_heuristic():
    assert _login_ok(httpx.Response(302, headers={"location": "https://seller.mprocket.ru/"})) is True
    assert _login_ok(httpx.Response(302, headers={"location": "https://seller.mprocket.ru/businesses"})) is True
    assert _login_ok(httpx.Response(302, headers={"location": "https://seller.mprocket.ru/login"})) is False
    assert _login_ok(httpx.Response(200)) is False  # re-render = ошибка входа


def test_redact_masks_phone():
    out = _redact("вход +79969190097 отклонён")
    assert "+79969190097" not in out
    assert "[phone]" in out


# ─── request-list parser (поставки / приёмки, read-only зеркало) ─────────────

_SHIPMENT_HTML = """
<div class="uk-grid mprs-data-row" data-sortable-id="17006" uk-grid>
    <div class="uk-width-expand@m uk-text-default">
        🟣
        <a href="https://seller.mprocket.ru/businesses/306/ff/shipments/17006/shipment-items">Поставка №17006</a>
        <div>MP Rocket — Отгрузка транспортом клиента</div>
        <div class="uk-text-small">09.07.2026</div>
    </div>
    <div class="uk-width-1-6">
        <div style="color: #9851E6">ШК коробов отправлены</div>
        <div>354 шт</div>
    </div>
    <div class="uk-width-1-6 uk-text-break">Вяткин Ковры 300526 сбор .xlsx</div>
</div>
<ul class="uk-pagination"><li><a href="?page=2">2</a></li></ul>
"""

_RECEIVING_HTML = """
<div class="uk-grid mprs-data-row" data-sortable-id="5039" uk-grid>
    <div class="uk-width-expand uk-text-default">
        <a href="https://seller.mprocket.ru/businesses/306/ff/receivings/5039/receiving-items">№5039: Приёмка на складе</a>
        <div class="uk-text-small">10.01.2026</div>
    </div>
    <div class="uk-width-1-5">Мест: 550 + добавить товары</div>
    <div class="uk-width-1-5">Создана</div>
</div>
"""


def test_parse_request_rows_shipment():
    rows = parse_request_rows(_SHIPMENT_HTML)
    assert len(rows) == 1
    r = rows[0]
    assert r["external_id"] == "17006"
    assert r["number"] == "17006"
    assert "Отгрузка транспортом клиента" in r["type_name"]
    assert r["status_title"] == "ШК коробов отправлены"
    assert r["date_str"] == "2026-07-09"  # dd.mm.yyyy → ISO
    assert r["total_qty"] == 354
    assert "300526" in r["notes"]


def test_parse_request_rows_receiving():
    rows = parse_request_rows(_RECEIVING_HTML)
    assert len(rows) == 1
    r = rows[0]
    assert r["external_id"] == "5039"
    assert r["type_name"] == "Приёмка на складе"  # из текста ссылки, «№5039:» снят
    assert r["status_title"] == "Создана"
    assert r["date_str"] == "2026-01-10"


def test_parse_request_rows_skips_rows_without_id_and_empty():
    assert parse_request_rows("") == []
    assert parse_request_rows('<div class="uk-grid mprs-data-row" uk-grid>нет id</div>') == []


def test_parse_request_rows_multiple():
    rows = parse_request_rows(_SHIPMENT_HTML + _RECEIVING_HTML)
    assert [r["external_id"] for r in rows] == ["17006", "5039"]
