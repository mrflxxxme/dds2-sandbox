# ruff: noqa: RUF001, RUF002, RUF003
"""Тесты генератора описи .xlsx для migfull-портала (чистые, без БД/сети)."""

from io import BytesIO

from openpyxl import load_workbook

from backend.schemas.migfull_portal import MigfullOpisLine
from backend.services.migfull_opis import build_opis_xlsx


def _cells(content: bytes) -> dict[str, object]:
    ws = load_workbook(BytesIO(content), data_only=True).active
    return {c.coordinate: c.value for row in ws.iter_rows() for c in row if c.value is not None}


def test_header_and_table_titles():
    cells = _cells(build_opis_xlsx([], incoming_number="40337302", incoming_date="2026-06-26"))
    assert cells["A1"] == "Перемещение на ответственное хранение"
    assert cells["A2"] == "Номер входящего документа"
    assert cells["B2"] == "40337302"
    assert cells["B3"] == "2026-06-26"
    # заголовки таблицы строго на строке 7
    assert cells["A7"] == "ШК"
    assert cells["B7"] == "Наименование товара"
    assert cells["E7"] == "Кол-во"


def test_box_line_writes_itf14_and_box_count():
    line = MigfullOpisLine(
        barcode="12049985828273", name="ELKA_короб 5 шт", quantity=8, is_box=True, units_per_box=5, pieces=40
    )
    cells = _cells(build_opis_xlsx([line]))
    # ШК числом (как в оригинале migfull), кол-во = число коробов
    assert cells["A8"] == 12049985828273
    assert cells["B8"] == "ELKA_короб 5 шт"
    assert cells["E8"] == 8


def test_loose_line_writes_ean13_and_pieces():
    line = MigfullOpisLine(barcode="2049985828273", name="палас", quantity=4, is_box=False, units_per_box=1, pieces=4)
    cells = _cells(build_opis_xlsx([line]))
    assert cells["A8"] == 2049985828273
    assert cells["E8"] == 4


def test_non_digit_barcode_written_as_text():
    line = MigfullOpisLine(barcode="ABC-12", name="x", quantity=1, pieces=1)
    cells = _cells(build_opis_xlsx([line]))
    assert cells["A8"] == "ABC-12"


def test_multiple_lines_sequential_rows():
    lines = [
        MigfullOpisLine(barcode="111", name="a", quantity=1, pieces=1),
        MigfullOpisLine(barcode="222", name="b", quantity=2, pieces=2),
    ]
    cells = _cells(build_opis_xlsx(lines))
    assert cells["A8"] == 111 and cells["A9"] == 222
    assert cells["E9"] == 2
