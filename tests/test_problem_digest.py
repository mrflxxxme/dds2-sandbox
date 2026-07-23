"""
Tests for backend/services/funnel/problem_digest.py — сводка «Проблемные товары».

Чистая логика: sanitize_settings (дефолты/мусор), enrich_row + pick_sections
(пороги и сортировки всех шести секций), build_tg_text, рендер xlsx
(листы, свёрнутые группы, подсветка остатков), итоги для «Динамики».
"""

import io

from openpyxl import load_workbook

from backend.services.funnel.problem_digest import (
    DEFAULT_SETTINGS,
    _cmp_totals_by_brand,
    build_tg_text,
    enrich_row,
    fmt_days,
    pick_sections,
    rub_short,
    sanitize_settings,
)
from backend.services.funnel.problem_digest_xlsx import render_workbook

CFG = dict(DEFAULT_SETTINGS)


def _row(
    nm_id: int = 100,
    brand: str = "Бренд",
    orders_sum: float = 10000,
    orders_count: int = 10,
    adv_sum: float = 500,
    margin: float = 20,
    drr: float | None = 5,
    profit: float = 2000,
    wb_qty: float = 50,
    own_qty: float = 0,
    wb_cost: float = 5000,
    own_cost: float = 0,
    days_left: float | None = 30,
    wb_days_left: float | None = 30,
) -> dict:
    return {
        "nm_id": nm_id,
        "vendor_code": f"art-{nm_id}",
        "brand": brand,
        "subject": "Ковры",
        "orders_sum_rub": orders_sum,
        "orders_count": orders_count,
        "adv_sum": adv_sum,
        "adv_clicks": 100,
        "margin": margin,
        "drr": drr,
        "profit": profit,
        "avg_price": 1000,
        "cart_to_order_pct": 30,
        "wb_stock_qty": wb_qty,
        "own_stock_qty": own_qty,
        "wb_stock_cost": wb_cost,
        "own_stock_cost": own_cost,
        "stock_days_left": days_left,
        "wb_stock_days_left": wb_days_left,
    }


def _items(*rows: dict) -> list[dict]:
    return [enrich_row(r, None) for r in rows]


# ─── sanitize_settings ───────────────────────────────────────────────────────


def test_sanitize_defaults_on_garbage():
    for raw in (None, [], "x", 42):
        cfg = sanitize_settings(raw)
        assert cfg == DEFAULT_SETTINGS
        assert cfg is not DEFAULT_SETTINGS  # не шарим мутабельный дефолт


def test_sanitize_filters_values():
    cfg = sanitize_settings(
        {
            "enabled": 1,
            "chat_ids": [4910742599, "-100123", "мусор", None],
            "exclude_brands": ["Redmi", "", 5],
            "drr_pct": 12.5,
            "turnover_days": 90,
            "top_n": 50,
            "неизвестный": "ключ",
        }
    )
    assert cfg["enabled"] is True
    assert cfg["chat_ids"] == [4910742599, -100123]
    assert cfg["exclude_brands"] == ["Redmi"]
    assert cfg["drr_pct"] == 12.5
    assert cfg["turnover_days"] == 90
    assert cfg["top_n"] == 50
    assert "неизвестный" not in cfg


def test_sanitize_rejects_out_of_range():
    cfg = sanitize_settings({"turnover_days": -5, "drr_pct": 99999})
    assert cfg["turnover_days"] == DEFAULT_SETTINGS["turnover_days"]
    assert cfg["drr_pct"] == DEFAULT_SETTINGS["drr_pct"]


# ─── pick_sections: пороги ──────────────────────────────────────────────────


def test_drr_above_threshold_and_inf():
    items = _items(
        _row(nm_id=1, drr=10, adv_sum=100),          # ровно порог — не попадает
        _row(nm_id=2, drr=10.1, adv_sum=100),        # выше — попадает
        _row(nm_id=3, drr=None, adv_sum=100, orders_sum=0, orders_count=0),  # ∞ — первым
        _row(nm_id=4, drr=50, adv_sum=0),            # нет расхода — не ДРР-проблема
    )
    sec = pick_sections(items, CFG)["drr"]
    assert [i["r"]["nm_id"] for i in sec] == [3, 2]
    assert sec[0]["drr_inf"] is True


def test_margin_below_threshold_sorted_asc():
    items = _items(
        _row(nm_id=1, margin=5),      # ровно порог — не попадает
        _row(nm_id=2, margin=4.9),
        _row(nm_id=3, margin=-20),    # худший — первым
        _row(nm_id=4, margin=3, orders_sum=0, orders_count=0),  # без заказов — не считаем
    )
    sec = pick_sections(items, CFG)["margin"]
    assert [i["r"]["nm_id"] for i in sec] == [3, 2]


def test_turnover_sorted_by_frozen_money():
    items = _items(
        _row(nm_id=1, days_left=71, wb_cost=1000),
        _row(nm_id=2, days_left=200, wb_cost=99000),  # больше заморожено — первым
        _row(nm_id=3, days_left=70),                   # ровно порог — не попадает
        _row(nm_id=4, days_left=999, wb_qty=0, own_qty=0, wb_cost=0),  # без остатка — мимо
    )
    sec = pick_sections(items, CFG)["turnover"]
    assert [i["r"]["nm_id"] for i in sec] == [2, 1]


def test_no_ads_requires_zero_spend_and_stock():
    items = _items(
        _row(nm_id=1, adv_sum=0, days_left=31),
        _row(nm_id=2, adv_sum=0, days_left=30),   # ровно порог — не попадает
        _row(nm_id=3, adv_sum=1, days_left=100),  # есть расход — мимо
        _row(nm_id=4, adv_sum=0, days_left=100, wb_qty=0, own_qty=0),  # нет остатка — мимо
    )
    sec = pick_sections(items, CFG)["no_ads"]
    assert [i["r"]["nm_id"] for i in sec] == [1]


def test_low_stock_and_out_stock_sections():
    items = _items(
        _row(nm_id=1, wb_qty=5, wb_days_left=6),
        _row(nm_id=2, wb_qty=5, wb_days_left=2),                # меньше дней — первым
        _row(nm_id=3, wb_qty=5, wb_days_left=14),               # ровно порог — не попадает
        _row(nm_id=4, wb_qty=0, orders_sum=50000),              # 🚫 — сортировка по заказам
        _row(nm_id=5, wb_qty=0, orders_sum=90000, own_qty=10),  # 🚫 — первым
    )
    sec = pick_sections(items, CFG)
    assert [i["r"]["nm_id"] for i in sec["low_stock"]] == [2, 1]
    assert [i["r"]["nm_id"] for i in sec["out_stock"]] == [5, 4]


# ─── текст и форматирование ─────────────────────────────────────────────────


def test_rub_short_and_fmt_days():
    assert rub_short(72_956_000) == "73,0 млн ₽"
    assert rub_short(294_000) == "294к ₽"
    assert fmt_days(None) == "—"
    assert fmt_days(999) == "∞"
    assert fmt_days(95.4) == "95"


def _digest(sections_by_brand: dict, gaps: dict | None = None) -> dict:
    return {
        "sections_by_brand": sections_by_brand,
        "gap_by_brand": gaps or {},
        "cmp_rows_by_brand": {},
        "main_rows_by_brand": {},
    }


def test_tg_text_with_sheet_url():
    text = build_tg_text("Тест", sheet_url="https://docs.google.com/spreadsheets/d/abc")
    assert '⚠️ Проблемные товары — в файле и <a href="https://docs.google.com/spreadsheets/d/abc">таблице</a>' in text


def test_tg_text_revenue_only_no_problem_products():
    """Сообщение = общая выручка; проблемные товары в тексте не перечисляются."""
    from backend.services.funnel.problem_digest import revenue_by_brand

    rows = [
        _row(nm_id=1, brand="АРТСПЕЙС", orders_sum=800_000),
        _row(nm_id=2, brand="АРТСПЕЙС", orders_sum=400_000),
        _row(nm_id=3, brand="НУ-НУ", orders_sum=2_000_000),
        _row(nm_id=4, brand="Redmi", orders_sum=999_999),
    ]
    rows[0]["subject"] = "Ковры"
    rows[1]["subject"] = "Чехлы"
    rows[3]["subject"] = "Планшеты"
    revenue = revenue_by_brand(rows, {"Redmi"})
    assert "Redmi" not in revenue

    text = build_tg_text("Сводка · 23.07.2026", revenue=revenue, revenue_label="за 22.07")
    lines = text.splitlines()
    # общий итог по всем товарам (3,2 млн) в заголовке блока
    assert "💰 <b>Выручка за 22.07 — 3,2 млн ₽</b>" in lines
    # бренды по убыванию выручки: НУ-НУ (2,0 млн) раньше АРТСПЕЙС (1,2 млн)
    assert lines.index("<b>НУ-НУ</b> — 2,0 млн ₽") < lines.index("<b>АРТСПЕЙС</b> — 1,2 млн ₽")
    # разбивка по категориям под брендом
    assert "   • Ковры — 800к ₽" in lines
    assert "   • Чехлы — 400к ₽" in lines
    # никаких артикулов/чипсов проблемных товаров в тексте
    assert "🔥" not in text and "🧊" not in text
    assert text.strip().endswith("⚠️ Проблемные товары — в файле")


def test_revenue_by_brand_top_categories_rest():
    from backend.services.funnel.problem_digest import revenue_by_brand

    rows = [_row(nm_id=n, brand="Б", orders_sum=1000 * n) for n in range(1, 9)]
    for n, r in enumerate(rows, 1):
        r["subject"] = f"Кат{n}"
    revenue = revenue_by_brand(rows, set())
    text = build_tg_text("Тест", revenue=revenue, revenue_label="за день")
    # топ-5 категорий + строка «прочие (3)»
    assert "   • Кат8 — 8 000 ₽" in text
    assert "прочие (3)" in text


def test_sanitize_sheet_settings():
    cfg = sanitize_settings(
        {
            "sheet_id": "  1y2_abc  ",
            "sheet_id_weekly": 123,  # не строка — игнор
            "share_emails": ["scramline1336@gmail.com", "мусор", 5, " x@y.ru "],
        }
    )
    assert cfg["sheet_id"] == "1y2_abc"
    assert cfg["sheet_id_weekly"] == ""
    assert cfg["share_emails"] == ["scramline1336@gmail.com", "x@y.ru"]


def test_gsheet_jwt_claims():
    from backend.services.funnel.problem_digest_gsheet import SCOPE, build_jwt_claims

    creds = {"client_email": "bot@p.iam.gserviceaccount.com", "token_uri": "https://oauth2.googleapis.com/token"}
    claims = build_jwt_claims(creds, 1000.9)
    assert claims["iss"] == "bot@p.iam.gserviceaccount.com"
    assert claims["scope"] == SCOPE
    assert claims["aud"] == "https://oauth2.googleapis.com/token"
    assert claims["exp"] - claims["iat"] == 3600


def test_gsheet_disabled_without_key(monkeypatch):
    from backend.services.funnel import problem_digest_gsheet as g

    monkeypatch.setattr(g.settings, "GOOGLE_SA_JSON_PATH", "", raising=False)
    assert g.load_credentials() is None


# ─── xlsx ────────────────────────────────────────────────────────────────────


def _meta(**over) -> dict:
    meta = {
        "cfg": dict(CFG),
        "title": "Проблемные товары · тест",
        "period_label": "неделя 01.06–07.06",
        "cmp_label": "вчера 07.06",
        "cmp_short": "вчера",
        "with_dynamics": False,
    }
    meta.update(over)
    return meta


def test_render_workbook_sheets_and_groups():
    cfg = dict(CFG, top_n=5)
    rows = [_row(nm_id=n, brand="НУ-НУ", days_left=200, wb_cost=1000 + n) for n in range(1, 12)]
    sections = {"НУ-НУ": pick_sections(_items(*rows), cfg)}
    data = render_workbook(_digest(sections), _meta(cfg=cfg))

    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Сводка", "НУ-НУ"]
    ws = wb["НУ-НУ"]
    # топ-5 видимых + 6 в свёрнутой группе
    hidden = [rd for rd in ws.row_dimensions.values() if rd.hidden]
    assert len(hidden) == 6
    assert all(rd.outlineLevel == 1 for rd in hidden)
    values = [c.value for c in ws[1]] + [ws["A2"].value]
    assert any("НУ-НУ" == v for v in values)


def test_render_workbook_wb_stock_highlight():
    items = _items(
        _row(nm_id=1, brand="Б", wb_qty=0, orders_sum=5000),        # «0 · нет»
        _row(nm_id=2, brand="Б", wb_qty=5, wb_days_left=3, drr=99, adv_sum=100),
    )
    data = render_workbook(_digest({"Б": pick_sections(items, CFG)}), _meta())
    ws = load_workbook(io.BytesIO(data))["Б"]
    cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "0 · нет" in cells


def test_render_workbook_dynamics_sheet():
    main = {"Б": {"orders_sum_rub": 200.0, "adv_sum": 20.0, "profit": 50.0, "revenue": 180.0, "drr": 10.0, "margin": 27.8}}
    cmp_totals = {"Б": {"orders_sum_rub": 100.0, "adv_sum": 30.0, "profit": 10.0, "revenue": 90.0, "drr": 30.0, "margin": 11.1}}
    digest = {
        "sections_by_brand": {"Б": pick_sections([], CFG)},
        "gap_by_brand": {},
        "main_rows_by_brand": main,
        "cmp_rows_by_brand": cmp_totals,
    }
    data = render_workbook(digest, _meta(with_dynamics=True, cmp_short="пред. нед"))
    wb = load_workbook(io.BytesIO(data))
    assert "Динамика" in wb.sheetnames


def test_cmp_totals_by_brand_derives_drr_and_margin():
    rows = [
        _row(nm_id=1, brand="А", orders_sum=1000, adv_sum=100, profit=100),
        _row(nm_id=2, brand="А", orders_sum=1000, adv_sum=100, profit=100),
        _row(nm_id=3, brand="Redmi", orders_sum=500, adv_sum=50),
    ]
    for r in rows:
        r["revenue"] = r["orders_sum_rub"] * 0.9
    totals = _cmp_totals_by_brand(rows, {"Redmi"})
    assert set(totals) == {"А"}
    assert totals["А"]["orders_sum_rub"] == 2000
    assert totals["А"]["drr"] == 10.0  # 200/2000
    assert round(totals["А"]["margin"], 1) == 11.1  # 200/1800


def test_budget_gap_section_rendered():
    digest = {
        "sections_by_brand": {"Б": pick_sections([], CFG)},
        "gap_by_brand": {
            "Б": [{"campaign_id": 123, "name": "Тест", "spend_today": 500,
                   "ran_out_at": "2026-06-19T14:30:00+03:00", "needed_potential": 1500}]
        },
        "cmp_rows_by_brand": {},
        "main_rows_by_brand": {},
    }
    data = render_workbook(digest, _meta())
    ws = load_workbook(io.BytesIO(data))["Б"]
    cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("Нехватка бюджета" in v for v in cells)
    assert "14:30" in cells
