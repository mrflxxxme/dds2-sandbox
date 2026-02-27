"""
Себестоимость: номенклатура, правила пошлин, заказы и расчёт.
"""

import streamlit as st
import pandas as pd
from datetime import date
from frontend import api_client as api


DUTY_BASIS_LABELS = {
    "WEIGHT": "Евро/кг",
    "AREA": "Евро/м²",
    "INVOICE": "% от инвойса",
}


def render():
    st.title("📦 Себестоимость")

    tab1, tab2, tab3, tab4 = st.tabs([
        "📋 Заказы", "🔢 Номенклатура", "⚖️ Пошлины / Утиль", "❓ Неопознанные"
    ])

    # ── Вкладка 1: Заказы ─────────────────────────────────────────────────────
    with tab1:
        st.subheader("Заказы (Себестоимость)")

        orders = api.get_cost_orders()

        if orders:
            df_ord = pd.DataFrame(orders)
            df_ord["ship_date"] = pd.to_datetime(df_ord["ship_date"]).dt.strftime("%d.%m.%Y").fillna("—")
            df_ord["transport_label"] = df_ord.get("transport_type", "AUTO").map(
                {"AUTO": "🚛 Авто", "CONTAINER": "🚢 Контейнер"}
            ).fillna("—")
            df_ord["plan_status"] = df_ord.get("has_plan", False).apply(
                lambda x: "✅" if x else "—"
            )

            # Format currency columns
            for col in ["total_rub", "total_cost_rub", "total_delivery_rub", "total_duty_rub", "total_vat_rub", "total_util_rub"]:
                if col in df_ord.columns:
                    df_ord[col + "_fmt"] = df_ord[col].apply(lambda x: f"{x:,.0f} ₽")

            show_cols = {
                "order_no": "Заказ", "invoice_no": "Инвойс",
                "ship_date": "Дата отправки", "transport_label": "Транспорт",
                "items_count": "Позиций", "total_qty": "Кол-во (шт)",
                "total_cost_rub_fmt": "Себестоимость",
                "total_delivery_rub_fmt": "Доставка",
                "total_duty_rub_fmt": "Пошлина",
                "total_vat_rub_fmt": "НДС",
                "total_util_rub_fmt": "Утильсбор",
                "total_rub_fmt": "Итого ₽",
                "unrecognized_count": "❓",
                "plan_status": "План",
            }
            avail_cols = [c for c in show_cols.keys() if c in df_ord.columns]
            st.dataframe(
                df_ord[avail_cols].rename(columns=show_cols),
                hide_index=True, use_container_width=True
            )

        # Выбор заказа для просмотра
        st.markdown("---")
        col_sel, col_del = st.columns([4, 1])
        with col_sel:
            order_nos = [o["order_no"] for o in orders] if orders else []
            selected_order = st.selectbox("Выбрать заказ для просмотра / загрузки", ["— новый —"] + order_nos)

        with col_del:
            if selected_order != "— новый —" and st.button("🗑 Удалить", key="del_order"):
                try:
                    api.delete_cost_order(selected_order)
                    st.success(f"Заказ {selected_order} удалён")
                    st.rerun()
                except Exception as e:
                    st.error(f"Ошибка: {e}")

        if selected_order == "— новый —":
            # Форма создания заказа
            with st.expander("➕ Создать новый заказ", expanded=True):
                c1, c2 = st.columns(2)
                with c1:
                    new_order_no = st.text_input("Номер заказа*", key="new_ord_no")
                    new_invoice = st.text_input("Инвойс", key="new_inv")
                    new_ship_date = st.date_input("Дата отправки", value=date.today(), key="new_date")
                    new_transport = st.selectbox("Тип транспорта", ["AUTO", "CONTAINER"],
                                                 format_func=lambda x: {"AUTO": "🚛 Авто", "CONTAINER": "🚢 Контейнер"}[x],
                                                 key="new_transport")
                with c2:
                    new_delivery_cny = st.number_input("Доставка (CNY)", min_value=0.0, step=100.0, key="new_del_cny")
                    new_delivery_usd = st.number_input("Доставка (USD)", min_value=0.0, step=100.0, key="new_del_usd")
                    new_note = st.text_input("Примечание", key="new_note")

                st.markdown("**Курсы валют:**")
                r1, r2, r3 = st.columns(3)
                with r1:
                    rate_cny = st.number_input("CNY/RUB", value=13.5, step=0.1, format="%.4f", key="new_rcny")
                with r2:
                    rate_eur = st.number_input("EUR/RUB", value=100.0, step=0.5, format="%.4f", key="new_reur")
                with r3:
                    rate_usd = st.number_input("USD/RUB", value=90.0, step=0.5, format="%.4f", key="new_rusd")

                if st.button("💾 Создать заказ", key="btn_create_order"):
                    if not new_order_no:
                        st.error("Введите номер заказа")
                    else:
                        try:
                            api.create_cost_order({
                                "order_no": new_order_no,
                                "invoice_no": new_invoice or None,
                                "ship_date": str(new_ship_date),
                                "transport_type": new_transport,
                                "delivery_cost_cny": new_delivery_cny,
                                "delivery_cost_usd": new_delivery_usd,
                                "rate_cny": rate_cny,
                                "rate_eur": rate_eur,
                                "rate_usd": rate_usd,
                                "note": new_note or None,
                            })
                            st.success(f"✅ Заказ {new_order_no} создан!")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Ошибка: {e}")
        else:
            # Просмотр и редактирование заказа
            order_data = next((o for o in orders if o["order_no"] == selected_order), None)
            if order_data:
                transport_label = {"AUTO": "🚛 Авто", "CONTAINER": "🚢 Контейнер"}.get(
                    order_data.get("transport_type", "AUTO"), "—"
                )

                # Editable fields
                with st.expander("✏️ Редактировать заказ", expanded=False):
                    ec1, ec2 = st.columns(2)
                    with ec1:
                        edit_invoice = st.text_input("Инвойс", value=order_data.get("invoice_no") or "", key="edit_inv")
                        edit_ship = st.date_input("Дата отправки",
                            value=date.fromisoformat(order_data["ship_date"]) if order_data.get("ship_date") else date.today(),
                            key="edit_ship")
                        edit_transport = st.selectbox("Тип транспорта", ["AUTO", "CONTAINER"],
                            index=0 if order_data.get("transport_type", "AUTO") == "AUTO" else 1,
                            format_func=lambda x: {"AUTO": "🚛 Авто", "CONTAINER": "🚢 Контейнер"}[x],
                            key="edit_transport")
                    with ec2:
                        current_arrival = None
                        if order_data.get("actual_arrival_date"):
                            current_arrival = date.fromisoformat(order_data["actual_arrival_date"])
                        use_arrival = st.checkbox("Указать фактическую дату прихода",
                            value=current_arrival is not None, key="edit_use_arrival")
                        if use_arrival:
                            edit_arrival = st.date_input("Факт. дата прихода",
                                value=current_arrival or date.today(), key="edit_arrival")
                        else:
                            edit_arrival = None

                        edit_del_cny = st.number_input("Доставка CNY", value=float(order_data["delivery_cost_cny"]),
                            min_value=0.0, step=100.0, key="edit_del_cny")
                        edit_del_usd = st.number_input("Доставка USD", value=float(order_data["delivery_cost_usd"]),
                            min_value=0.0, step=100.0, key="edit_del_usd")

                    er1, er2, er3 = st.columns(3)
                    with er1:
                        edit_rcny = st.number_input("Курс CNY", value=float(order_data["rate_cny"]),
                            step=0.1, format="%.4f", key="edit_rcny")
                    with er2:
                        edit_reur = st.number_input("Курс EUR", value=float(order_data["rate_eur"]),
                            step=0.5, format="%.4f", key="edit_reur")
                    with er3:
                        edit_rusd = st.number_input("Курс USD", value=float(order_data["rate_usd"]),
                            step=0.5, format="%.4f", key="edit_rusd")

                    if st.button("💾 Сохранить изменения", key="btn_save_edit"):
                        try:
                            api.update_cost_order(selected_order, {
                                "invoice_no": edit_invoice or None,
                                "ship_date": str(edit_ship),
                                "actual_arrival_date": str(edit_arrival) if edit_arrival else None,
                                "transport_type": edit_transport,
                                "delivery_cost_cny": edit_del_cny,
                                "delivery_cost_usd": edit_del_usd,
                                "rate_cny": edit_rcny,
                                "rate_eur": edit_reur,
                                "rate_usd": edit_rusd,
                            })
                            st.success("✅ Заказ обновлён")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Ошибка: {e}")

                # Summary metrics
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("Инвойс", order_data.get("invoice_no") or "—")
                c2.metric("Транспорт", transport_label)
                c3.metric("Доставка CNY", f"{order_data['delivery_cost_cny']:,.0f}")
                c4.metric("Курс CNY", f"{order_data['rate_cny']:.4f}")
                arrival_str = order_data.get("actual_arrival_date") or "план"
                c5.metric("Приход", arrival_str)

                # Generate plan button
                plan_col1, plan_col2 = st.columns([3, 1])
                with plan_col1:
                    has_plan = order_data.get("has_plan", False)
                    if has_plan:
                        st.success("✅ План оплат сгенерирован → Планирование")
                    else:
                        st.info("План оплат не создан")
                with plan_col2:
                    btn_label = "🔄 Пересоздать план" if has_plan else "📋 Сгенерировать план оплат"
                    if st.button(btn_label, key="btn_gen_plan"):
                        try:
                            result = api.generate_cost_plan(selected_order)
                            plan = result.get("plan", {})
                            st.success(
                                f"✅ Создано {result['payments_created']} платежей:\n"
                                f"- Заказ: {plan.get('order_cny', 0):,.0f} CNY → {plan.get('pay_date_order', '')}\n"
                                f"- Доставка: {plan.get('delivery_rub', 0):,.0f} ₽ → {plan.get('pay_date_delivery', '')}\n"
                                f"- Таможня: {plan.get('customs_rub', 0):,.0f} ₽ → {plan.get('pay_date_customs', '')} (приход: {plan.get('arrival_date', '')})"
                            )
                            st.rerun()
                        except Exception as e:
                            st.error(f"Ошибка: {e}")

            # Загрузка файла позиций
            st.markdown("#### 📥 Загрузить файл заказа")
            uploaded = st.file_uploader("Excel файл заказа (дивандек / ковры)", type=["xlsx"], key=f"ul_{selected_order}")
            if uploaded and st.button("📊 Рассчитать себестоимость", key="btn_calc"):
                try:
                    result = api.upload_cost_order_items(selected_order, uploaded.read(), uploaded.name)
                    st.success(f"✅ Загружено позиций: {result['inserted']}, неопознано: {result['unrecognized']}")
                    if result["unrecognized"] > 0:
                        st.warning("⚠️ Есть неопознанные баркоды — проверьте вкладку ❓ Неопознанные")
                    st.rerun()
                except Exception as e:
                    st.error(f"Ошибка: {e}")

            # Таблица позиций
            items = api.get_cost_order_items(selected_order)
            if items:
                df_items = pd.DataFrame(items)

                st.markdown("#### 📊 Расчёт себестоимости по позициям")

                # Filter out totals/invalid rows (barcode=0 or empty)
                df_items = df_items[df_items["barcode"].astype(str).str.strip().ne("0")]
                df_items = df_items[df_items["barcode"].astype(str).str.strip().ne("")]

                # Итоги
                total_qty = df_items["qty"].sum()
                total_cost = (df_items["cost_rub"] * df_items["qty"]).sum()
                total_delivery = (df_items["delivery_rub"] * df_items["qty"]).sum()
                total_duty = (df_items["duty_rub"] * df_items["qty"]).sum()
                total_vat = (df_items["vat_rub"] * df_items["qty"]).sum()
                total_util = (df_items["util_rub"] * df_items["qty"]).sum()
                grand_total = (df_items["total_rub"] * df_items["qty"]).sum()

                def _pct(part):
                    return f"{part / grand_total * 100:.1f}%" if grand_total else "—"

                col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
                col_m1.metric("Кол-во (шт)", f"{total_qty:,}")
                col_m2.metric("Себестоимость", f"{total_cost:,.0f} ₽", delta=_pct(total_cost), delta_color="off")
                col_m3.metric("Доставка", f"{total_delivery:,.0f} ₽", delta=_pct(total_delivery), delta_color="off")
                col_m4.metric("Пошлина", f"{total_duty:,.0f} ₽", delta=_pct(total_duty), delta_color="off")
                col_m5.metric("НДС", f"{total_vat:,.0f} ₽", delta=_pct(total_vat), delta_color="off")

                col_m6, col_m7 = st.columns([1, 4])
                col_m6.metric("Утильсбор", f"{total_util:,.0f} ₽", delta=_pct(total_util), delta_color="off")
                col_m7.metric("🏷 ИТОГО", f"{grand_total:,.0f} ₽")

                # Build WB link from article_wb
                if "article_wb" in df_items.columns:
                    df_items["wb_url"] = df_items["article_wb"].apply(
                        lambda x: f"https://www.wildberries.ru/catalog/{int(x)}/detail.aspx"
                        if pd.notna(x) and str(x).strip() not in ("", "0", "None", "nan")
                        else None
                    )
                else:
                    df_items["wb_url"] = None

                # Таблица по артикулам
                display_cols = {
                    "barcode": "Баркод", "subject": "Категория",
                    "article_seller": "Артикул", "wb_url": "WB",
                    "qty": "Кол-во",
                    "price_cny": "Цена CNY", "cost_rub": "Себест. ₽/шт",
                    "delivery_rub": "Доставка ₽/шт", "duty_rub": "Пошлина ₽/шт",
                    "vat_rub": "НДС ₽/шт", "util_rub": "Утиль ₽/шт",
                    "total_rub": "Итого ₽/шт", "total_cny": "Итого CNY/шт",
                    "unrecognized": "❓",
                }
                df_show = df_items[[c for c in display_cols.keys() if c in df_items.columns]].copy()

                # Format numbers before display
                for col in ["cost_rub", "delivery_rub", "duty_rub", "vat_rub", "util_rub", "total_rub"]:
                    if col in df_show.columns:
                        df_show[col] = df_show[col].apply(lambda x: f"{x:,.0f}")
                for col in ["price_cny", "total_cny"]:
                    if col in df_show.columns:
                        df_show[col] = df_show[col].apply(lambda x: f"{x:,.2f}")

                df_show = df_show.rename(columns=display_cols)

                st.dataframe(
                    df_show,
                    hide_index=True, use_container_width=True, height=400,
                    column_config={
                        "WB": st.column_config.LinkColumn(
                            "WB", display_text="🔗 Открыть",
                        ),
                    },
                )

                # Экспорт в Excel
                import io
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                buf = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = f"Заказ {selected_order}"

                # --- Шапка с итогами ---
                header_font = Font(bold=True, size=12)
                ws["A1"] = f"Себестоимость заказа {selected_order}"
                ws["A1"].font = Font(bold=True, size=14)
                ws.merge_cells("A1:H1")

                summary_labels = ["Кол-во", "Себестоимость", "Доставка", "Пошлина", "НДС", "Утильсбор", "ИТОГО"]
                summary_values = [total_qty, total_cost, total_delivery, total_duty, total_vat, total_util, grand_total]
                summary_pcts = ["", _pct(total_cost), _pct(total_delivery), _pct(total_duty), _pct(total_vat), _pct(total_util), "100%"]

                for col_idx, (lbl, val, pct) in enumerate(zip(summary_labels, summary_values, summary_pcts), 1):
                    cell_lbl = ws.cell(row=3, column=col_idx, value=lbl)
                    cell_lbl.font = Font(bold=True, size=10)
                    cell_lbl.fill = PatternFill("solid", fgColor="D9E1F2")
                    cell_val = ws.cell(row=4, column=col_idx, value=val)
                    cell_val.number_format = '#,##0' if col_idx == 1 else '#,##0 "₽"'
                    cell_val.font = Font(bold=True, size=11)
                    if pct:
                        ws.cell(row=5, column=col_idx, value=pct).font = Font(color="808080", size=9)

                # --- Таблица позиций ---
                table_start = 7
                xl_cols = [
                    ("Баркод", "barcode", None),
                    ("Категория", "subject", None),
                    ("Артикул", "article_seller", None),
                    ("Артикул WB", "article_wb", None),
                    ("Кол-во", "qty", "#,##0"),
                    ("Цена CNY", "price_cny", "#,##0.00"),
                    ("Себест. ₽/шт", "cost_rub", "#,##0"),
                    ("Доставка ₽/шт", "delivery_rub", "#,##0"),
                    ("Пошлина ₽/шт", "duty_rub", "#,##0"),
                    ("НДС ₽/шт", "vat_rub", "#,##0"),
                    ("Утиль ₽/шт", "util_rub", "#,##0"),
                    ("Итого ₽/шт", "total_rub", "#,##0"),
                    ("Итого CNY/шт", "total_cny", "#,##0.00"),
                ]

                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"),
                )
                header_fill = PatternFill("solid", fgColor="4472C4")
                header_font_w = Font(bold=True, color="FFFFFF", size=10)

                for col_idx, (label, _, _) in enumerate(xl_cols, 1):
                    cell = ws.cell(row=table_start, column=col_idx, value=label)
                    cell.font = header_font_w
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border

                for row_idx, (_, row) in enumerate(df_items.iterrows(), table_start + 1):
                    for col_idx, (_, field, fmt) in enumerate(xl_cols, 1):
                        val = row.get(field)
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        if fmt:
                            cell.number_format = fmt
                        cell.border = thin_border
                        # WB hyperlink
                        if field == "article_wb" and pd.notna(val) and str(val).strip() not in ("", "0", "None", "nan"):
                            wb_url = f"https://www.wildberries.ru/catalog/{int(val)}/detail.aspx"
                            cell.hyperlink = wb_url
                            cell.font = Font(color="0563C1", underline="single")

                # Auto-fit column widths
                for col_idx in range(1, len(xl_cols) + 1):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = max(14, len(xl_cols[col_idx - 1][0]) + 4)

                wb.save(buf)
                buf.seek(0)

                st.download_button(
                    "⬇️ Скачать Excel",
                    buf.getvalue(),
                    f"cost_{selected_order}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.info("Загрузите файл заказа для расчёта себестоимости.")

    # ── Вкладка 2: Номенклатура ───────────────────────────────────────────────
    with tab2:
        st.subheader("Справочник номенклатуры (Баркоды)")

        nom = api.get_nomenclature()
        if nom:
            df_nom = pd.DataFrame(nom)
            st.metric("Позиций в справочнике", len(df_nom))
            st.dataframe(
                df_nom[["barcode", "brand", "subject", "article_seller", "article_wb", "volume_l"]].rename(columns={
                    "barcode": "Баркод", "brand": "Бренд", "subject": "Категория",
                    "article_seller": "Артикул продавца", "article_wb": "Артикул WB",
                    "volume_l": "Объём, л",
                }),
                hide_index=True, use_container_width=True, height=400
            )
        else:
            st.info("Справочник пуст. Загрузите файл отчёта WB.")

        st.markdown("---")
        st.markdown("#### 📥 Загрузить номенклатуру из WB")
        uploaded_nom = st.file_uploader("Файл report_YYYY_MM_DD.xlsx от WB", type=["xlsx"], key="ul_nom")
        if uploaded_nom and st.button("📋 Загрузить", key="btn_nom"):
            try:
                result = api.upload_nomenclature(uploaded_nom.read(), uploaded_nom.name)
                st.success(f"✅ Добавлено: {result['inserted']}, обновлено: {result['updated']}")
                st.rerun()
            except Exception as e:
                st.error(f"Ошибка: {e}")

    # ── Вкладка 3: Пошлины и утиль ────────────────────────────────────────────
    with tab3:
        st.subheader("Справочник пошлин и утильсбора")

        rules = api.get_duty_rules()
        if rules:
            df_rules = pd.DataFrame(rules)
            df_rules["basis_label"] = df_rules["basis"].map(DUTY_BASIS_LABELS)
            st.dataframe(
                df_rules[["id", "subject", "basis_label", "rate", "util_collect_rub", "note"]].rename(columns={
                    "id": "ID", "subject": "Категория", "basis_label": "База пошлины",
                    "rate": "Тариф", "util_collect_rub": "Утильсбор ₽/шт", "note": "Примечание",
                }),
                hide_index=True, use_container_width=True
            )

            # Удаление — selectbox вместо ввода ID вручную
            rule_options = {f"{r['subject']} (ID:{r['id']})": r["id"] for r in rules}
            selected_rule = st.selectbox("Выберите правило для удаления", list(rule_options.keys()), key="del_duty_sel")
            if st.button("🗑 Удалить правило", key="btn_del_duty"):
                try:
                    api.delete_duty_rule(rule_options[selected_rule])
                    st.success(f"Удалено: {selected_rule}")
                    st.rerun()
                except Exception as e:
                    st.error(f"Ошибка: {e}")
        else:
            st.info("Правила не добавлены.")

        st.markdown("---")
        st.markdown("#### ➕ Добавить / изменить правило")
        with st.form("duty_form"):
            d1, d2 = st.columns(2)
            with d1:
                d_subject = st.text_input("Категория (Предмет WB)*", placeholder="Ковры")
                d_basis = st.selectbox("База расчёта пошлины", ["INVOICE", "WEIGHT", "AREA"],
                                       format_func=lambda x: DUTY_BASIS_LABELS[x])
                d_rate = st.number_input("Тариф", min_value=0.0, step=0.01, format="%.4f",
                                         help="Для INVOICE — в %, для WEIGHT/AREA — евро за ед.")
            with d2:
                d_util = st.number_input("Утильсбор (₽/шт)", min_value=0.0, step=1.0)
                d_note = st.text_input("Примечание", placeholder="0.61 EUR/кг")

            if st.form_submit_button("💾 Сохранить"):
                if not d_subject:
                    st.error("Введите категорию")
                else:
                    try:
                        api.upsert_duty_rule({
                            "subject": d_subject, "basis": d_basis,
                            "rate": d_rate, "util_collect_rub": d_util, "note": d_note,
                        })
                        st.success(f"✅ Сохранено: {d_subject}")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Ошибка: {e}")

        # Подсказки
        with st.expander("📖 Примеры правил для заполнения"):
            st.markdown("""
| Категория | База | Тариф | Утиль | Пояснение |
|-----------|------|-------|-------|-----------|
| Ковры | Евро/м² | 0.38 | 6.5 | 0.38 EUR за м² |
| Пледы | Евро/кг | 0.61 | 9 | 0.61 EUR за кг |
| Покрывала | % от инвойса | 12 | 9 | 12% от суммы инвойса |
| Чехлы для мебели | Евро/кг | 0.61 | 0 | 0.61 EUR за кг |
""")

    # ── Вкладка 4: Неопознанные ───────────────────────────────────────────────
    with tab4:
        st.subheader("❓ Неопознанные баркоды")
        st.caption("Баркоды из заказов, которых нет в справочнике номенклатуры")

        orders = api.get_cost_orders()
        all_unrecognized = []
        for order in orders:
            if order.get("unrecognized_count", 0) > 0:
                items = api.get_cost_order_items(order["order_no"])
                for item in items:
                    if item.get("unrecognized"):
                        all_unrecognized.append({
                            "order_no": order["order_no"],
                            "barcode": item["barcode"],
                            "qty": item["qty"],
                            "price_cny": item["price_cny"],
                        })

        if all_unrecognized:
            df_unr = pd.DataFrame(all_unrecognized)
            st.metric("Неопознанных позиций", len(df_unr))
            st.dataframe(
                df_unr.rename(columns={
                    "order_no": "Заказ", "barcode": "Баркод",
                    "qty": "Кол-во", "price_cny": "Цена CNY",
                }),
                hide_index=True, use_container_width=True
            )
            st.info("💡 Загрузите актуальный отчёт WB во вкладке 🔢 Номенклатура, затем повторно загрузите файл заказа.")
        else:
            st.success("✅ Все баркоды опознаны!")
