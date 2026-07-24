# ruff: noqa: RUF002, RUF003
"""
Импорт реестра займов из Excel (лист «Расчет процентов»).

Идемпотентно: ключ займа = (project_id, counterparty_id, contract_number).
Повторный импорт обновляет существующие, а не дублирует. «Приход» → займ,
«Возврат» → платёж PRINCIPAL_REPAY + закрытие. Сущность (физ/ИП) и банк —
из второго листа (по имени контрагента), best-effort.
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.cache import invalidate_project_reports
from backend.models.counterparty import Counterparty
from backend.models.loan import Loan, LoanPayment
from backend.schemas.loan import LoanImportResult

logger = logging.getLogger(__name__)

_EXCEL_EPOCH = date(1899, 12, 30)
# Bound decompressed row processing — defends against a zip-bomb xlsx whose
# compressed size passes the upload limit but expands to millions of rows.
_MAX_DATA_ROWS = 100000

_HEADER_ALIASES = {
    "date": {"дата"},
    "type": {"тип"},
    "amount": {"сумма"},
    "counterparty": {"контрагент"},
    "contract": {"номер договора", "договор", "№ договора"},
    "rate": {"ставка"},
    "from": {"от"},
    "to": {"до"},
    "status": {"статус"},
}


def _to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return _EXCEL_EPOCH + timedelta(days=int(value))
        except (ValueError, OverflowError):
            return None
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        try:
            return _EXCEL_EPOCH + timedelta(days=int(float(s)))
        except (ValueError, OverflowError):
            return None
    return None


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _norm_entity(raw: Any) -> str | None:
    if not raw:
        return None
    s = str(raw).strip().lower()
    if s.startswith("ип") or s == "ип":
        return "IP"
    if s.startswith("физ"):
        return "PHYSICAL"
    return None


def _map_headers(header_row: list[Any]) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip().lower()
        for key, aliases in _HEADER_ALIASES.items():
            if name in aliases and key not in out:
                out[key] = idx
    return out


def _find_register_sheet(wb: Any) -> tuple[Any, dict[str, int]] | None:
    """First sheet whose header row maps Тип+Сумма+Контрагент."""
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        cols = _map_headers(list(first))
        if {"type", "amount", "counterparty"} <= set(cols):
            return ws, cols
    return None


def _build_entity_map(wb: Any) -> dict[str, tuple[str | None, str | None]]:
    """name -> (entity_type, bank) from a sheet with «Контрагент» + «куда» columns."""
    result: dict[str, tuple[str | None, str | None]] = {}
    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        header = [str(c).strip().lower() if c is not None else "" for c in first]
        if "контрагент" not in header or "куда" not in header:
            continue
        name_col = header.index("контрагент")
        kuda_cols = [i for i, h in enumerate(header) if h == "куда"]
        entity_col = kuda_cols[0] if kuda_cols else None
        bank_col = kuda_cols[1] if len(kuda_cols) > 1 else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if name_col >= len(row):
                continue
            name = row[name_col]
            if not name:
                continue
            ent = _norm_entity(row[entity_col]) if entity_col is not None and entity_col < len(row) else None
            bank = None
            if bank_col is not None and bank_col < len(row) and row[bank_col]:
                bank = str(row[bank_col]).strip()[:50]
            key = str(name).strip().lower()
            prev = result.get(key, (None, None))
            result[key] = (ent or prev[0], bank or prev[1])
        break
    return result


async def import_loans_from_xlsx(
    db: AsyncSession, *, project_id: int, content: bytes
) -> LoanImportResult:
    """Parse an xlsx loan register and upsert loans/payments/counterparties."""
    res = LoanImportResult()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        raise ValueError(f"Не удалось прочитать Excel: {e}") from e

    found = _find_register_sheet(wb)
    if not found:
        raise ValueError("Не найден лист с колонками Тип/Сумма/Контрагент")
    ws, cols = found
    entity_map = _build_entity_map(wb)

    # Existing counterparties (by lowercased name) and loans (by (cp_id, contract))
    cp_rows = await db.execute(
        select(Counterparty).where(
            Counterparty.project_id == project_id,
            Counterparty.is_deleted == False,  # noqa: E712
        )
    )
    cp_by_name: dict[str, Counterparty] = {cp.name.strip().lower(): cp for cp in cp_rows.scalars().all()}

    loan_rows = await db.execute(
        select(Loan).where(Loan.project_id == project_id, Loan.is_deleted == False)  # noqa: E712
    )
    # Idempotency key = (cp_id, contract, start_date) — same contract number can be
    # reused for distinct tranches (14 cases in the source); keep them separate.
    loans_by_key: dict[tuple[int, str, str], Loan] = {}
    by_cp_contract: dict[tuple[int, str], list[Loan]] = {}
    for loan in loan_rows.scalars().all():
        c = loan.contract_number.strip()
        loans_by_key[(loan.counterparty_id, c, loan.start_date.isoformat())] = loan
        by_cp_contract.setdefault((loan.counterparty_id, c), []).append(loan)

    def col(row, key):
        idx = cols.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    returns: list[tuple[str, str, date, Decimal]] = []  # (cp_name, contract, date, amount)
    lender_names: set[str] = set()

    seen_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        seen_rows += 1
        if seen_rows > _MAX_DATA_ROWS:
            res.warnings.append(f"Превышен лимит строк ({_MAX_DATA_ROWS}) — файл обработан частично")
            break
        cp_name = col(row, "counterparty")
        typ = col(row, "type")
        if not cp_name or not typ:
            continue
        cp_name = str(cp_name).strip()
        typ = str(typ).strip().lower()
        contract = col(row, "contract")
        contract = str(contract).strip() if contract not in (None, "") else ""
        amount = _to_decimal(col(row, "amount")) or Decimal("0")

        # find-or-create counterparty
        key_name = cp_name.lower()
        cp = cp_by_name.get(key_name)
        if cp is None:
            ent, bank = entity_map.get(key_name, (None, None))
            cp = Counterparty(
                project_id=project_id,
                name=cp_name,
                primary_type="OTHER",
                created_by_import=True,
                notes="Заёмщик (импорт реестра займов)",
            )
            db.add(cp)
            await db.flush()
            cp_by_name[key_name] = cp
            res.created_counterparties += 1
        lender_names.add(cp_name)

        if typ.startswith("приход"):
            if not contract:
                res.skipped_rows += 1
                res.warnings.append(f"{cp_name}: приход без номера договора — пропущен")
                continue
            start = _to_date(col(row, "from")) or _to_date(col(row, "date"))
            if start is None:
                res.skipped_rows += 1
                res.warnings.append(f"{cp_name} {contract}: нет даты начала — пропущен")
                continue
            maturity = _to_date(col(row, "to"))
            rate = _to_decimal(col(row, "rate"))
            status_raw = col(row, "status")
            status = "ACTIVE" if (status_raw and "активен" == str(status_raw).strip().lower()) else "CLOSED"
            ent, bank = entity_map.get(key_name, (None, None))

            existing = loans_by_key.get((cp.id, contract, start.isoformat()))
            if existing is not None:
                existing.principal = amount
                existing.rate = rate
                existing.contract_date = start
                existing.maturity_date = maturity
                existing.status = status
                if ent and not existing.entity_type:
                    existing.entity_type = ent
                if bank and not existing.lender_bank:
                    existing.lender_bank = bank
                res.updated_loans += 1
            else:
                loan = Loan(
                    project_id=project_id,
                    counterparty_id=cp.id,
                    direction="INCOMING",
                    principal=amount,
                    currency="RUB",
                    rate=rate,
                    contract_number=contract,
                    contract_date=start,
                    start_date=start,
                    maturity_date=maturity,
                    status=status,
                    entity_type=ent,
                    lender_bank=bank,
                )
                db.add(loan)
                await db.flush()
                loans_by_key[(cp.id, contract, start.isoformat())] = loan
                by_cp_contract.setdefault((cp.id, contract), []).append(loan)
                res.created_loans += 1
        elif typ.startswith("возврат"):
            ret_date = _to_date(col(row, "date")) or _to_date(col(row, "from"))
            if ret_date and contract:
                returns.append((key_name, contract, ret_date, amount))

    # Second pass: returns → PRINCIPAL_REPAY (idempotent) + close loan.
    # Chronological order so an earlier return claims the earlier tranche when a
    # contract number is reused; repaid_ids keeps two returns off the same loan.
    repaid_ids: set[int] = set()
    for key_name, contract, ret_date, amount in sorted(returns, key=lambda r: r[2]):
        cp = cp_by_name.get(key_name)
        if cp is None:
            res.skipped_rows += 1
            continue
        candidates = by_cp_contract.get((cp.id, contract), [])
        target = _pick_repay_target(candidates, ret_date, repaid_ids)
        if target is None:
            res.warnings.append(f"{key_name} {contract}: возврат без займа — пропущен")
            res.skipped_rows += 1
            continue
        dup = (
            await db.execute(
                select(LoanPayment.id).where(
                    LoanPayment.loan_id == target.id,
                    LoanPayment.payment_type == "PRINCIPAL_REPAY",
                    LoanPayment.paid_at == ret_date,
                )
            )
        ).scalar_one_or_none()
        if dup is None:
            db.add(
                LoanPayment(
                    loan_id=target.id,
                    payment_type="PRINCIPAL_REPAY",
                    amount=amount,
                    currency=target.currency,
                    paid_at=ret_date,
                )
            )
            res.created_payments += 1
        target.status = "CLOSED"
        repaid_ids.add(target.id)

    # Third pass: best-effort extension chain linking (06 → 06-01 → 06-02)
    await db.flush()
    _link_extension_chains(by_cp_contract)

    await db.commit()
    await invalidate_project_reports(project_id)
    res.lenders = sorted(lender_names)
    return res


def _pick_repay_target(
    candidates: list[Loan], ret_date: date, repaid_ids: set[int] | None = None
) -> Loan | None:
    """Among loans sharing (cp, contract), choose which one a return closes."""
    if not candidates:
        return None
    done = repaid_ids or set()
    free = [loan for loan in candidates if loan.id not in done]
    pool = free or candidates
    # Продление без смены номера договора: старый транш погашается ровно в день
    # старта преемника, и возврат датирован этим днём. Сначала матчим по дате
    # погашения — иначе «последний старт ≤ возврата» закрывает ПРЕЕМНИКА,
    # а погашенный транш остаётся висеть активным (11 таких пар в реестре).
    exact = [loan for loan in pool if loan.maturity_date == ret_date]
    if exact:
        return min(exact, key=lambda loan: loan.start_date)
    active = [loan for loan in pool if loan.status == "ACTIVE"]
    ranked = active or pool
    # prefer the one started on/before the return, latest such start
    eligible = [loan for loan in ranked if loan.start_date <= ret_date]
    chosen = eligible or ranked
    return max(chosen, key=lambda loan: loan.start_date)


def _link_extension_chains(by_cp_contract: dict[tuple[int, str], list[Loan]]) -> None:
    """Link parent_loan_id for «X-NN» contracts to their predecessor (same cp)."""
    per_cp: dict[int, dict[str, Loan]] = {}
    for (cp_id, contract), loans in by_cp_contract.items():
        # if a contract maps to several tranches, take the earliest as the chain node
        per_cp.setdefault(cp_id, {})[contract] = min(loans, key=lambda loan: loan.start_date)

    for cp_id, contracts in per_cp.items():
        for contract, loan in contracts.items():
            m = re.match(r"^(.+?)\s*-\s*(\d+)\s*$", contract)
            if not m or loan.parent_loan_id is not None:
                continue
            base, num = m.group(1).strip(), int(m.group(2))
            predecessor = f"{base}-{num - 1:02d}" if num > 1 else None
            cands = [predecessor] if predecessor else [base, base.lstrip("0"), f"0{base}"]
            for cand in cands:
                if cand and cand in contracts and contracts[cand] is not loan:
                    loan.parent_loan_id = contracts[cand].id
                    break
