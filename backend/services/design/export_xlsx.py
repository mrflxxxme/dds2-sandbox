# ruff: noqa: RUF001, RUF002, RUF003
"""XLSX-выгрузка аналитики дизайна (волна D v2, CONTRACT-V2 §4).

Два листа: «Сводка» (метрики + три разреза) и «Задачи» (по строке на задачу).

Тихие усечения запрещены: и cap строк, и cap колонок-реквизитов проговариваются
отдельной строкой в «Сводке». Файл собирается openpyxl целиком в память —
как у донора problem_digest_xlsx; StreamingResponse не нужен.
"""

import io
from collections.abc import Sequence
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models.design import DesignAttribute, DesignTask
from backend.schemas.design import (
    DesignStatsByAssigneeOut,
    DesignStatsByAttributeOut,
    DesignStatsFunnelOut,
    DesignStatsOut,
)
from backend.services.design import analytics, stats
from backend.services.design.queries import (
    load_attributes_by_task,
    load_labels_by_task,
    to_list_items,
)

#  Cap листа «Задачи»: при усечении в «Сводке» появляется явная строка.
MAX_TASK_ROWS = 5000
#  Колонок-реквизитов не больше 50 — иначе лист расползается за пределы читаемого.
MAX_ATTRIBUTE_COLUMNS = 50

#  Множественные значения в одной ячейке — через запятую с пробелом; пустое
#  множество это ПУСТАЯ ячейка, а не «—»: прочерк мешал бы фильтрам Excel.
_JOIN = ", "

_HEAD = Font(bold=True)
_TITLE = Font(bold=True, size=13)

STATUS_LABEL = {
    "NEW": "Новые заявки",
    "ASSIGNED": "Назначена",
    "IN_PROGRESS": "В работе",
    "REVIEW": "На проверке",
    "REVISION": "Правки",
    "ON_HOLD": "Отложена",
    "ACCEPTED": "Принято",
    "CANCELLED": "Отменена",
}


def _pct(value: float | None) -> str:
    return "—" if value is None else f"{value * 100:.0f}%"


def _num(value: float | None) -> str:
    return "—" if value is None else f"{value:.1f}"


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


async def build_export(
    db: AsyncSession, project_id: int, date_from: date | None, date_to: date | None
) -> tuple[bytes, str]:
    """Собрать книгу. Возвращает (байты, имя файла)."""
    win_from, win_to = analytics.resolve_stats_window(date_from, date_to)

    metrics = await stats.get_stats(db, project_id, win_from, win_to)
    by_assignee = await analytics.get_by_assignee(db, project_id, win_from, win_to)
    funnel = await analytics.get_funnel(db, project_id, win_from, win_to)
    by_attribute = await analytics.get_by_attribute(db, project_id, win_from, win_to)

    #  Активные поля-реквизиты → колонки листа «Задачи», в порядке sort_order.
    fields_res = await db.execute(
        select(DesignAttribute.id, DesignAttribute.name)
        .where(
            DesignAttribute.project_id == project_id,
            DesignAttribute.is_deleted == False,  # noqa: E712
            DesignAttribute.is_archived == False,  # noqa: E712
        )
        .order_by(DesignAttribute.sort_order, DesignAttribute.id)
        .limit(MAX_ATTRIBUTE_COLUMNS + 1)
    )
    fields = fields_res.all()
    fields_total = len(fields)
    fields_clipped = fields_total > MAX_ATTRIBUTE_COLUMNS
    fields = fields[:MAX_ATTRIBUTE_COLUMNS]

    total_res = await db.execute(
        select(DesignTask)
        .where(
            DesignTask.project_id == project_id,
            DesignTask.is_deleted == False,  # noqa: E712
            *analytics._window(DesignTask.created_at, win_from, win_to),
        )
        .order_by(DesignTask.created_at.desc(), DesignTask.id.desc())
        .limit(MAX_TASK_ROWS + 1)
    )
    tasks = list(total_res.scalars().all())
    tasks_clipped = len(tasks) > MAX_TASK_ROWS
    tasks = tasks[:MAX_TASK_ROWS]

    #  Точное общее число нужно для честной строки усечения.
    total_tasks = len(tasks)
    if tasks_clipped:
        from sqlalchemy import func as sa_func

        count_res = await db.execute(
            select(sa_func.count()).select_from(DesignTask).where(
                DesignTask.project_id == project_id,
                DesignTask.is_deleted == False,  # noqa: E712
                *analytics._window(DesignTask.created_at, win_from, win_to),
            )
        )
        total_tasks = int(count_res.scalar() or 0)

    wb = Workbook()
    summary = wb.active
    summary.title = "Сводка"
    _fill_summary(
        summary,
        win_from,
        win_to,
        metrics,
        by_assignee,
        funnel,
        by_attribute,
        tasks_clipped=tasks_clipped,
        total_tasks=total_tasks,
        fields_clipped=fields_clipped,
        fields_total=fields_total,
    )

    sheet = wb.create_sheet("Задачи")
    await _fill_tasks(db, project_id, sheet, tasks, fields)

    buffer = io.BytesIO()
    wb.save(buffer)
    name = f"design-tasks_{win_from.isoformat()}_{win_to.isoformat()}.xlsx"
    return buffer.getvalue(), name


def _fill_summary(
    ws: Worksheet,
    win_from: date,
    win_to: date,
    metrics: DesignStatsOut,
    by_assignee: DesignStatsByAssigneeOut,
    funnel: DesignStatsFunnelOut,
    by_attribute: DesignStatsByAttributeOut,
    *,
    tasks_clipped: bool,
    total_tasks: int,
    fields_clipped: bool,
    fields_total: int,
) -> None:
    row = 1
    ws.cell(row=row, column=1, value="Дизайн карточек — сводка").font = _TITLE
    row += 1
    ws.cell(row=row, column=1, value=f"Период: {win_from.isoformat()} — {win_to.isoformat()}")
    row += 2

    #  Строки усечения идут ВВЕРХУ: их задача — попасться на глаза, а не потеряться.
    if tasks_clipped:
        ws.cell(
            row=row, column=1,
            value=f"Показано {MAX_TASK_ROWS} из {total_tasks} задач — сузьте период",
        ).font = _HEAD
        row += 1
    if fields_clipped:
        ws.cell(
            row=row, column=1,
            value=f"Показаны первые {MAX_ATTRIBUTE_COLUMNS} полей-реквизитов из {fields_total}",
        ).font = _HEAD
        row += 1
    if tasks_clipped or fields_clipped:
        row += 1

    ws.cell(row=row, column=1, value="Метрики").font = _HEAD
    row += 1
    for label, value in (
        ("В срок", _pct(metrics.on_time_share)),
        ("Версий до приёмки", _num(metrics.avg_versions_to_accept)),
        ("Медиана цикла, дн", _num(metrics.median_cycle_days)),
        ("Без исполнителя >2 дн", str(metrics.unassigned_over_2d)),
        ("Аутсорс", _pct(metrics.outsourced_share)),
        ("С товаром", _pct(metrics.tracked_share)),
    ):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="По исполнителям").font = _HEAD
    row += 1
    for i, head in enumerate(
        ("Исполнитель", "Активных", "Принято", "В срок", "Цикл, дн", "Версий"), start=1
    ):
        ws.cell(row=row, column=i, value=head).font = _HEAD
    row += 1
    for who in by_assignee.rows:
        ws.cell(row=row, column=1, value=who.name)
        ws.cell(row=row, column=2, value=who.active)
        ws.cell(row=row, column=3, value=who.accepted)
        ws.cell(row=row, column=4, value=_pct(who.on_time_share))
        ws.cell(row=row, column=5, value=_num(who.avg_cycle_days))
        ws.cell(row=row, column=6, value=_num(who.avg_versions))
        row += 1
    if by_assignee.truncated:
        ws.cell(row=row, column=1, value="Показаны не все исполнители — сузьте период")
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Воронка").font = _HEAD
    row += 1
    for i, head in enumerate(("Статус", "Задач сейчас", "Среднее время, дн"), start=1):
        ws.cell(row=row, column=i, value=head).font = _HEAD
    row += 1
    for stage in funnel.rows:
        ws.cell(row=row, column=1, value=STATUS_LABEL.get(stage.status, stage.status))
        ws.cell(row=row, column=2, value=stage.count)
        ws.cell(row=row, column=3, value=_num(stage.avg_days_in_status))
        row += 1
    row += 1

    for group in by_attribute.attributes:
        ws.cell(row=row, column=1, value=f"Реквизит: {group.attribute_name}").font = _HEAD
        row += 1
        for value_row in group.rows:
            ws.cell(row=row, column=1, value=value_row.value)
            ws.cell(row=row, column=2, value=value_row.count)
            row += 1
        row += 1

    if by_attribute.labels:
        ws.cell(row=row, column=1, value="Метки").font = _HEAD
        row += 1
        for label_row in by_attribute.labels:
            ws.cell(row=row, column=1, value=label_row.name)
            ws.cell(row=row, column=2, value=label_row.count)
            row += 1

    _autosize(ws, [34, 16, 14, 12, 12, 10])


async def _fill_tasks(
    db: AsyncSession,
    project_id: int,
    ws: Worksheet,
    tasks: list[DesignTask],
    fields: Sequence[Any],
) -> None:
    head = [
        "Номер", "Название", "Тип", "Сложность", "Срочно", "Исполнитель",
        "Автор", "Срок", "Статус", "Версий", "Метки",
        *[name for _fid, name in fields],
        "Создано", "Принято",
    ]
    for i, title in enumerate(head, start=1):
        cell = ws.cell(row=1, column=i, value=title)
        cell.font = _HEAD
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    if not tasks:
        _autosize(ws, [14, 40, 16, 12, 10, 20, 20, 12, 16, 9, 24])
        return

    items = await to_list_items(db, project_id, tasks)
    task_ids = [t.id for t in tasks]
    labels_by_task = await load_labels_by_task(db, project_id, task_ids)
    attrs_by_task = await load_attributes_by_task(db, project_id, task_ids)
    by_id = {t.id: t for t in tasks}

    row = 2
    for item in items:
        task = by_id[item.id]
        values_by_field: dict[int, list[str]] = {}
        for a in attrs_by_task.get(item.id, []):
            values_by_field.setdefault(a.attribute_id, []).append(a.value)

        ws.cell(row=row, column=1, value=item.number)
        ws.cell(row=row, column=2, value=item.title)
        ws.cell(row=row, column=3, value=item.work_type)
        ws.cell(row=row, column=4, value=item.complexity)
        ws.cell(row=row, column=5, value="да" if item.is_urgent else "")
        ws.cell(row=row, column=6, value=item.assignee_name or "")
        ws.cell(row=row, column=7, value=item.author_name or "")
        ws.cell(row=row, column=8, value=item.due_date.isoformat() if item.due_date else "")
        ws.cell(row=row, column=9, value=STATUS_LABEL.get(item.status, item.status))
        ws.cell(row=row, column=10, value=item.versions_count)
        ws.cell(
            row=row, column=11,
            value=_JOIN.join(lb.name for lb in labels_by_task.get(item.id, [])),
        )
        col = 12
        for field_id, _name in fields:
            ws.cell(row=row, column=col, value=_JOIN.join(values_by_field.get(field_id, [])))
            col += 1
        ws.cell(row=row, column=col, value=task.created_at.isoformat(sep=" ", timespec="minutes"))
        col += 1
        ws.cell(
            row=row, column=col,
            value=task.accepted_at.isoformat(sep=" ", timespec="minutes") if task.accepted_at else "",
        )
        row += 1

    _autosize(ws, [14, 40, 16, 12, 10, 20, 20, 12, 16, 9, 24])
