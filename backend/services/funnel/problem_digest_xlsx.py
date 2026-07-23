# ruff: noqa: RUF001, RUF002, RUF003
"""Рендер xlsx для сводки «Проблемные товары».

Дизайн согласован с пользователем (июль 2026): лист «Сводка» + лист на бренд,
на листе бренда — секции по проблемам (у каждой свой приглушённый цвет),
в секции топ-N строк + свёрнутая группа «+» с остальными, единый набор метрик,
колонка «ВБ, дн» с подсветкой «заканчивается» / «нет на ВБ».
"""

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend.services.funnel.problem_digest import (
    PRODUCT_SECTIONS,
    _num,
    fmt_days,
)

# Насыщенный акцент + мягкая подложка (палитра v4, одобрена пользователем)
STYLE: dict[str, dict[str, str]] = {
    "drr": {"accent": "B0443C", "soft": "F6E0DD"},
    "margin": {"accent": "AD731A", "soft": "F8EBD4"},
    "turnover": {"accent": "38639F", "soft": "DFE9F6"},
    "no_ads": {"accent": "6A4FA3", "soft": "E8E1F5"},
    "low_stock": {"accent": "BC6A14", "soft": "FBEBD3"},
    "out_stock": {"accent": "97352F", "soft": "F5DBD9"},
    "budget": {"accent": "2F7D6E", "soft": "DCEEE9"},
}

SECTION_TITLES = {
    "drr": "🔥  Высокий ДРР",
    "margin": "💸  Низкая маржа",
    "turnover": "🐌  Низкая оборачиваемость",
    "no_ads": "💤  Не работает реклама",
    "low_stock": "📦  Заканчиваются остатки на ВБ",
    "out_stock": "🚫  Нет остатков на ВБ",
    "budget": "⏳  Нехватка бюджета",
}

INK = "16181D"
hair = Side(style="hair", color="DFE3E8")
BORDER = Border(left=hair, right=hair, top=hair, bottom=hair)
F_TITLE = Font(name="Arial", bold=True, size=15, color=INK)
F_SUB = Font(name="Arial", size=9, color="6E7480")
F_BASE = Font(name="Arial", size=10, color=INK)
F_NAME = Font(name="Arial", size=10, color="2A2E35")
F_DIM = Font(name="Arial", size=9, color="8A9099", italic=True)
NEG = Font(name="Arial", size=10, bold=True, color="B0443C")
ZEBRA = PatternFill("solid", fgColor="F4F6F8")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

FILL_LOW = PatternFill("solid", fgColor="FBEBD3")
FONT_LOW = Font(name="Arial", size=10, bold=True, color="9A6410")
FILL_OUT = PatternFill("solid", fgColor="F5DBD9")
FONT_OUT = Font(name="Arial", size=10, bold=True, color="97352F")

FMT_RUB = '#,##0" ₽"'
FMT_PCT = '0.0"%"'
FMT_INT = "#,##0"

# (заголовок, ширина); {cmp} — подпись сравнительного периода (вчера / пред. нед)
COLS: list[tuple[str, int]] = [
    ("Артикул", 11), ("Название", 24), ("Маржа %", 9), ("Маржа {cmp}", 10),
    ("ДРР %", 9), ("ДРР {cmp}", 10), ("Расход ₽", 10), ("Заказы ₽", 10),
    ("Заказы шт", 8), ("Прибыль ₽", 10), ("Цена ср ₽", 9), ("Клики", 8),
    ("Корз→заказ %", 11), ("Остаток шт", 9), ("Запас дн", 8), ("ВБ, дн", 8),
    ("Заморожено ₽", 12),
]
NCOLS = len(COLS)
KEY_COL = {"drr": 5, "margin": 3, "turnover": 17, "no_ads": 17, "low_stock": 16, "out_stock": 8}


def _hint(key: str, cfg: dict[str, Any]) -> str:
    return {
        "drr": f"реклама дороже {cfg['drr_pct']:g}% от заказов · сортировка: ДРР по убыванию",
        "margin": f"маржа ниже {cfg['margin_pct']:g}% · сортировка: от меньшей к большей",
        "turnover": f"запаса больше чем на {cfg['turnover_days']} дней · сортировка: по замороженным деньгам",
        "no_ads": f"расход 0₽ при запасе от {cfg['no_ads_stock_days']} дней · сортировка: по замороженным деньгам",
        "low_stock": f"на ВБ меньше чем на {cfg['wb_low_days']} дней · сортировка: по дням до нуля",
        "out_stock": "на ВБ ноль — продажи стоят · сортировка: по заказам за период",
        "budget": "кампании, где бюджет кончается до конца дня",
    }[key]


def _cellf(ws: Worksheet, row: int, col: int, val: Any, font: Font = F_BASE,
           align: Alignment = RIGHT, fmt: str | None = None, fill: PatternFill | None = None) -> None:
    c = ws.cell(row=row, column=col, value=val)
    c.font, c.alignment, c.border = font, align, BORDER
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill


def _section_header(ws: Worksheet, row: int, key: str, count: int, cfg: dict[str, Any]) -> int:
    s = STYLE[key]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws.cell(row=row, column=1, value=f"{SECTION_TITLES[key]}  ·  {count} шт")
    c.font = Font(name="Arial", size=11, bold=True, color=s["accent"])
    c.fill = PatternFill("solid", fgColor=s["soft"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(left=Side(style="thick", color=s["accent"]))
    ws.row_dimensions[row].height = 26
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=NCOLS)
    h = ws.cell(row=row + 1, column=1, value=_hint(key, cfg))
    h.font = Font(name="Arial", size=8.5, color="6E7480", italic=True)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return row + 2


def _col_header(ws: Worksheet, row: int, key: str, cmp_short: str) -> int:
    for c, (title, _w) in enumerate(COLS, 1):
        cell = ws.cell(row=row, column=c, value=title.format(cmp=cmp_short))
        bold = c == KEY_COL[key]
        cell.font = Font(name="Arial", size=9, bold=True, color=STYLE[key]["accent"] if bold else "4A5058")
        cell.border = Border(bottom=Side(style="thin", color="B8BEC6"))
        cell.alignment = CENTER if c > 2 else LEFT
    return row + 1


def _drr_value(i: dict[str, Any]) -> Any:
    if i["drr_inf"]:
        return "∞"
    return i["drr"] if i["drr"] is not None else "—"


def _cmp_drr_value(cmp_row: dict[str, Any]) -> Any:
    if not cmp_row:
        return "—"
    if _num(cmp_row.get("adv_sum")) > 0 and _num(cmp_row.get("orders_sum_rub")) == 0:
        return "∞"
    return _num(cmp_row.get("drr")) if cmp_row.get("drr") is not None else "—"


def _metric_row(ws: Worksheet, row: int, i: dict[str, Any], key: str, zebra: bool, cfg: dict[str, Any]) -> None:
    fill = ZEBRA if zebra else None
    r, cmp_row = i["r"], i["cmp"]
    key_font = Font(name="Arial", size=10, bold=True, color=STYLE[key]["accent"])

    _cellf(ws, row, 1, r["nm_id"], align=LEFT, fmt=FMT_INT, fill=fill)
    _cellf(ws, row, 2, r.get("vendor_code") or "", align=LEFT, fill=fill, font=F_NAME)
    m = _num(r.get("margin"))
    _cellf(ws, row, 3, m, fmt=FMT_PCT, fill=fill,
           font=key_font if key == "margin" else (NEG if m < 0 else F_BASE))
    _cellf(ws, row, 4, _num(cmp_row.get("margin")) if cmp_row else "—", fmt=FMT_PCT, fill=fill)
    _cellf(ws, row, 5, _drr_value(i), fmt=FMT_PCT, fill=fill, font=key_font if key == "drr" else F_BASE)
    _cellf(ws, row, 6, _cmp_drr_value(cmp_row), fmt=FMT_PCT, fill=fill)
    _cellf(ws, row, 7, round(_num(r.get("adv_sum"))), fmt=FMT_RUB, fill=fill)
    _cellf(ws, row, 8, round(_num(r.get("orders_sum_rub"))), fmt=FMT_RUB, fill=fill,
           font=key_font if key == "out_stock" else F_BASE)
    _cellf(ws, row, 9, int(_num(r.get("orders_count"))), fmt=FMT_INT, fill=fill)
    _cellf(ws, row, 10, round(_num(r.get("profit"))), fmt=FMT_RUB, fill=fill,
           font=NEG if _num(r.get("profit")) < 0 else F_BASE)
    _cellf(ws, row, 11, round(_num(r.get("avg_price"))), fmt=FMT_RUB, fill=fill)
    _cellf(ws, row, 12, int(_num(r.get("adv_clicks"))), fmt=FMT_INT, fill=fill)
    _cellf(ws, row, 13, _num(r.get("cart_to_order_pct")), fmt=FMT_PCT, fill=fill)
    _cellf(ws, row, 14, int(i["stock_qty"]), fmt=FMT_INT, fill=fill)
    _cellf(ws, row, 15, fmt_days(i["turnover"]), fill=fill,
           font=key_font if key == "turnover" else F_BASE)
    if i["wb_qty"] == 0:
        _cellf(ws, row, 16, "0 · нет", fill=FILL_OUT, font=FONT_OUT, align=CENTER)
    elif i["wb_days"] is not None and i["wb_days"] < cfg["wb_low_days"]:
        _cellf(ws, row, 16, fmt_days(i["wb_days"]), fill=FILL_LOW, font=FONT_LOW, align=CENTER)
    else:
        _cellf(ws, row, 16, fmt_days(i["wb_days"]), fill=fill)
    _cellf(ws, row, 17, round(i["frozen"]), fmt=FMT_RUB, fill=fill,
           font=key_font if key in ("turnover", "no_ads") else F_BASE)


def _stop_label(g: dict[str, Any]) -> str:
    ran_out = g.get("ran_out_at")
    if ran_out:
        return str(ran_out)[11:16] or "—"
    typical = g.get("typical_stop_hour")
    if typical is not None:
        return f"~{typical:g}ч (прогноз)"
    return "—"


def render_workbook(digest: dict[str, Any], meta: dict[str, Any]) -> bytes:
    """Собрать xlsx: «Сводка» (+«Динамика» у недельной) + лист на каждый бренд."""
    cfg = meta["cfg"]
    top_n = int(cfg["top_n"])
    sections_by_brand: dict[str, dict[str, list[dict[str, Any]]]] = digest["sections_by_brand"]
    gap_by_brand: dict[str, list[dict[str, Any]]] = digest["gap_by_brand"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws.sheet_view.showGridLines = False
    ws["B2"] = meta["title"]
    ws["B2"].font = F_TITLE
    ws["B3"] = (
        f"{meta['period_label']} · сравнение: {meta['cmp_label']} · пороги: "
        f"оборач. >{cfg['turnover_days']} дн · маржа <{cfg['margin_pct']:g}% · "
        f"ДРР >{cfg['drr_pct']:g}% · ВБ <{cfg['wb_low_days']} дн"
    )
    ws["B3"].font = F_SUB
    row = 5
    hdr = ["Бренд", *[SECTION_TITLES[k] for k in PRODUCT_SECTIONS], SECTION_TITLES["budget"], "Заморожено ₽"]
    for c, t in enumerate(hdr, 2):
        cell = ws.cell(row=row, column=c, value=t)
        cell.font = Font(name="Arial", size=9, bold=True, color="4A5058")
        cell.border = Border(bottom=Side(style="medium", color=INK))
        cell.alignment = CENTER if c > 2 else LEFT
        cell.fill = PatternFill("solid", fgColor="EEF1F4")
    row += 1
    for n, (brand, p) in enumerate(sections_by_brand.items()):
        frozen = sum(i["frozen"] for i in p["turnover"])
        vals: list[Any] = [brand, *[len(p[k]) for k in PRODUCT_SECTIONS], len(gap_by_brand.get(brand, [])), frozen]
        for c, v in enumerate(vals, 2):
            fill = ZEBRA if n % 2 else None
            fmt = FMT_RUB if c == len(vals) + 1 else (FMT_INT if c > 2 else None)
            f = Font(name="Arial", size=10.5, bold=True, color=INK) if c == 2 else F_BASE
            if c == 7 and v:  # 📦
                f = FONT_LOW
            if c == 8 and v:  # 🚫
                f = FONT_OUT
            _cellf(ws, row, c, v, align=LEFT if c == 2 else CENTER, fmt=fmt, fill=fill, font=f)
        row += 1
    for idx, w in enumerate([2, 20, 13, 13, 20, 17, 22, 15, 12, 13], 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    if meta.get("with_dynamics"):
        _render_dynamics(wb, digest, meta)

    for brand, p in sections_by_brand.items():
        wsb = wb.create_sheet(brand[:31])
        wsb.sheet_view.showGridLines = False
        wsb.sheet_properties.outlinePr.summaryBelow = True
        wsb["A1"] = brand
        wsb["A1"].font = F_TITLE
        wsb["A2"] = (
            f"{meta['period_label']} · сравнение: {meta['cmp_label']} · "
            f"топ-{top_n} в каждой секции, остальное — раскрыть «+» слева"
        )
        wsb["A2"].font = F_SUB
        row = 4
        for key in PRODUCT_SECTIONS:
            sec = p[key]
            if not sec:
                continue
            row = _section_header(wsb, row, key, len(sec), cfg)
            row = _col_header(wsb, row, key, meta["cmp_short"])
            bar_from = row
            for n, i in enumerate(sec[:top_n]):
                _metric_row(wsb, row, i, key, bool(n % 2), cfg)
                row += 1
            rest = sec[top_n:]
            if rest:
                grp_start = row
                for n, i in enumerate(rest):
                    _metric_row(wsb, row, i, key, bool(n % 2), cfg)
                    row += 1
                for rr in range(grp_start, row):
                    wsb.row_dimensions[rr].outlineLevel = 1
                    wsb.row_dimensions[rr].hidden = True
                extra = ""
                if key in ("turnover", "no_ads"):
                    frozen_rest = round(sum(i["frozen"] for i in rest))
                    extra = f" · заморожено {frozen_rest:,} ₽".replace(",", " ")
                _cellf(wsb, row, 1, f"⊕ скрыто ещё {len(rest)} товаров — раскрыть «+» слева{extra}",
                       font=F_DIM, align=LEFT)
                row += 1
            if key == "drr":
                wsb.conditional_formatting.add(
                    f"E{bar_from}:E{row - 1}",
                    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100,
                                color="CE8B85", showValue=True),
                )
            row += 1

        gaps = gap_by_brand.get(brand, [])
        if gaps:
            row = _section_header(wsb, row, "budget", len(gaps), cfg)
            for c, t in enumerate(["Кампания", "Название", "Стоп", "Расход сегодня", "Долить до потенциала"], 1):
                cell = wsb.cell(row=row, column=c, value=t)
                cell.font = Font(name="Arial", size=9, bold=True, color="4A5058")
                cell.border = Border(bottom=Side(style="thin", color="B8BEC6"))
            row += 1
            for n, g in enumerate(gaps):
                fill = ZEBRA if n % 2 else None
                _cellf(wsb, row, 1, g.get("campaign_id"), align=LEFT, fill=fill)
                _cellf(wsb, row, 2, g.get("name") or "", align=LEFT, fill=fill, font=F_NAME)
                _cellf(wsb, row, 3, _stop_label(g), align=CENTER, fill=fill)
                _cellf(wsb, row, 4, round(_num(g.get("spend_today"))), fmt=FMT_RUB, fill=fill)
                needed = g.get("needed_potential")
                _cellf(wsb, row, 5, round(_num(needed)) if needed is not None else "—", fmt=FMT_RUB, fill=fill)
                row += 1

        for idx, (_t, w) in enumerate(COLS, 1):
            wsb.column_dimensions[get_column_letter(idx)].width = w
        wsb.freeze_panes = "C4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_dynamics(wb: Workbook, digest: dict[str, Any], meta: dict[str, Any]) -> None:
    """Лист «Динамика» (недельная сводка): метрики брендов неделя против недели."""
    ws = wb.create_sheet("Динамика")
    ws.sheet_view.showGridLines = False
    ws["B2"] = "Динамика по брендам"
    ws["B2"].font = F_TITLE
    ws["B3"] = f"{meta['period_label']} против «{meta['cmp_label']}» · Δ = стало − было"
    ws["B3"].font = F_SUB
    main = digest["main_rows_by_brand"]
    cmp_totals = digest["cmp_rows_by_brand"]
    metrics = [
        ("Заказы ₽", "orders_sum_rub", FMT_RUB, False),
        ("Расход ₽", "adv_sum", FMT_RUB, True),
        ("Прибыль ₽", "profit", FMT_RUB, False),
        ("ДРР %", "drr", FMT_PCT, True),
        ("Маржа %", "margin", FMT_PCT, False),
    ]
    row = 5
    hdr = ["Бренд"]
    for title, *_rest in metrics:
        hdr += [title, "было", "Δ"]
    for c, t in enumerate(hdr, 2):
        cell = ws.cell(row=row, column=c, value=t)
        cell.font = Font(name="Arial", size=9, bold=True, color="4A5058")
        cell.border = Border(bottom=Side(style="medium", color=INK))
        cell.alignment = CENTER if c > 2 else LEFT
        cell.fill = PatternFill("solid", fgColor="EEF1F4")
    row += 1
    for n, brand in enumerate(sorted(main)):
        fill = ZEBRA if n % 2 else None
        _cellf(ws, row, 2, brand, align=LEFT, fill=fill, font=Font(name="Arial", size=10.5, bold=True, color=INK))
        col = 3
        for _title, key, fmt, invert in metrics:
            cur = round(main[brand].get(key, 0), 2)
            was = round(cmp_totals.get(brand, {}).get(key, 0), 2)
            delta = round(cur - was, 2)
            worse = delta > 0 if invert else delta < 0
            _cellf(ws, row, col, cur, fmt=fmt, fill=fill)
            _cellf(ws, row, col + 1, was, fmt=fmt, fill=fill, font=F_DIM)
            _cellf(ws, row, col + 2, delta, fmt=fmt, fill=fill,
                   font=NEG if worse else Font(name="Arial", size=10, bold=True, color="2F7D6E"))
            col += 3
        row += 1
    ws.column_dimensions["B"].width = 20
    for idx in range(3, 3 + len(metrics) * 3):
        ws.column_dimensions[get_column_letter(idx)].width = 11
