# ruff: noqa: RUF002, RUF003
"""
Генератор описи .xlsx для заявки на отгрузку в migfull-портал.

Формат снят с реальной описи migfull (download существующей заявки):
- A1: «Перемещение на ответственное хранение» (тип документа)
- A2..A5: метки шапки (Номер/Дата входящего документа, Наименование/ИНН поклажедателя)
- строка 7: заголовки таблицы — ШК · Наименование товара · Размер · Цвет · Кол-во · Цена, руб. · Примечание
- строки 8+: данные. Для коробов ШК = ITF14 + Кол-во = число коробов; для россыпи
  ШК = EAN13 + Кол-во = штук. Размер/Цвет/Цена/Примечание — пусто (как в оригинале).

migfull парсит таблицу на своей стороне → Планируемое/Контейнеры. Пишем ШК числом
(как в оригинале): ITF14/EAN13 без ведущих нулей укладываются в int точно.

Чистая функция (без БД/сети) — юнит-тестируемо.
"""

from io import BytesIO

from openpyxl import Workbook

from backend.schemas.migfull_portal import MigfullOpisLine

OPIS_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_TABLE_HEADERS = ["ШК", "Наименование товара", "Размер", "Цвет", "Кол-во", "Цена, руб.", "Примечание (не обяз.)"]
_TABLE_HEADER_ROW = 7


def _barcode_cell(barcode: str) -> object:
    """ШК числом (как в оригинале), если строго цифры; иначе текстом."""
    b = (barcode or "").strip()
    return int(b) if b.isdigit() else b


def build_opis_xlsx(
    lines: list[MigfullOpisLine],
    *,
    incoming_number: str | None = None,
    incoming_date: str | None = None,
    consignor_name: str | None = None,
    consignor_inn: str | None = None,
) -> bytes:
    """Собрать .xlsx-опись из строк. Возвращает байты файла."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    # Шапка (метка в A, значение в B — как в шаблоне поклажедателя)
    ws["A1"] = "Перемещение на ответственное хранение"
    header_meta = [
        ("Номер входящего документа", incoming_number),
        ("Дата входящего документа", incoming_date),
        ("Наименование поклажедателя", consignor_name),
        ("ИНН поклажедатель", consignor_inn),
    ]
    for i, (label, value) in enumerate(header_meta, start=2):
        ws.cell(row=i, column=1, value=label)
        if value:
            ws.cell(row=i, column=2, value=value)

    # Заголовки таблицы
    for col, title in enumerate(_TABLE_HEADERS, start=1):
        ws.cell(row=_TABLE_HEADER_ROW, column=col, value=title)

    # Данные
    row = _TABLE_HEADER_ROW + 1
    for line in lines:
        ws.cell(row=row, column=1, value=_barcode_cell(line.barcode))
        ws.cell(row=row, column=2, value=line.name or "")
        ws.cell(row=row, column=3, value=line.size or "")
        ws.cell(row=row, column=4, value=line.color or "")
        ws.cell(row=row, column=5, value=int(line.quantity))
        # F (Цена) и G (Примечание) — пусто
        row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
