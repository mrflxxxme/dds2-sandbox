"""
ETL parsers: raw bank statements → normalized DataFrame.
Each parser returns a DataFrame with standardized columns:
  date, bank, account, currency, counterparty, inn, counterparty_account,
  purpose, income, expense
"""

import logging
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

import pandas as pd

logger = logging.getLogger(__name__)


# ─── Column mapping & normalization helpers ──────────────────────────────────

def _to_decimal(val) -> Decimal:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return Decimal("0")
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0")


def _clean_str(val) -> Optional[str]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    return str(val).strip() or None


def _parse_date(val) -> datetime:
    if isinstance(val, datetime):
        return val.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(val, str):
        for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    try:
        return pd.to_datetime(val).to_pydatetime().replace(hour=0, minute=0, second=0, microsecond=0)
    except Exception:
        raise ValueError(f"Cannot parse date: {val!r}")


NORM_COLS = [
    "date", "bank", "account", "currency",
    "counterparty", "inn", "counterparty_account",
    "purpose", "income", "expense",
]


def _find_columns(df: pd.DataFrame, mapping: dict[str, list[str]]) -> dict[str, str]:
    """
    Find actual column names by partial case-insensitive match.

    Args:
        df: DataFrame with column headers
        mapping: {logical_name: [search_patterns...]}

    Returns:
        {logical_name: actual_column_name}

    Raises:
        KeyError if a required column is not found.
    """
    result = {}
    cols_lower = {str(c).strip().lower(): c for c in df.columns}

    for logical, patterns in mapping.items():
        found = None
        for pattern in patterns:
            pattern_l = pattern.lower()
            for cl, original in cols_lower.items():
                if pattern_l in cl:
                    found = original
                    break
            if found:
                break
        if found is None:
            raise KeyError(
                f"Column '{logical}' not found. "
                f"Searched for {patterns} in columns: {list(df.columns)}"
            )
        result[logical] = found

    return result


# ─── VTB RUB parser (MAIN / TRANSIT) ─────────────────────────────────────────

_VTB_RUB_COLUMNS = {
    "date": ["Дата"],
    "counterparty": ["Контрагент"],
    "inn": ["ИНН"],
    "cp_account": ["Счет контрагента", "Счёт контрагента"],
    "debit": ["Дебет"],
    "credit": ["Кредит"],
    "purpose": ["Назначение"],
}


def parse_vtb_rub(data: bytes, source_type: str, account_no: str) -> tuple[pd.DataFrame, int]:
    """
    Columns: Дата, Номер, Вид операции, Контрагент, ИНН контрагента,
             БИК банка контрагента, Счет контрагента, Дебет RUR, Кредит RUR, Назначение
    Debit  = expense (money out), Credit = income (money in).

    Returns: (DataFrame, skipped_rows_count)
    """
    df = pd.read_excel(BytesIO(data), header=0)
    col_map = _find_columns(df, _VTB_RUB_COLUMNS)

    # drop rows where date is NaN
    df = df.dropna(subset=[col_map["date"]])

    bank = "VTB"
    currency = "RUB"
    skipped = 0

    rows = []
    for idx, row in df.iterrows():
        try:
            dt = _parse_date(row[col_map["date"]])
        except Exception as e:
            logger.warning("VTB_RUB row %s skipped: date parse error: %s", idx, e)
            skipped += 1
            continue

        debit = _to_decimal(row[col_map["debit"]])
        credit = _to_decimal(row[col_map["credit"]])

        rows.append({
            "date": dt,
            "bank": bank,
            "account": account_no,
            "currency": currency,
            "counterparty": _clean_str(row[col_map["counterparty"]]),
            "inn": _clean_str(row[col_map["inn"]]),
            "counterparty_account": _clean_str(row[col_map["cp_account"]]),
            "purpose": _clean_str(row[col_map["purpose"]]),
            "income": credit,
            "expense": debit,
        })

    if skipped:
        logger.info("VTB_RUB parsing: %d rows parsed, %d skipped", len(rows), skipped)

    result = pd.DataFrame(rows, columns=NORM_COLS) if rows else pd.DataFrame(columns=NORM_COLS)
    return result, skipped


# ─── VTB CNY parser ───────────────────────────────────────────────────────────

_VTB_CNY_COLUMNS = {
    "date": ["Дата"],
    "counterparty": ["Контрагент"],
    "inn": ["ИНН"],
    "cp_account": ["Счет контрагента", "Счёт контрагента"],
    "debit": ["Дебет CNY", "Дебет юан", "Дебет"],
    "credit": ["Кредит CNY", "Кредит юан", "Кредит"],
    "purpose": ["Назначение"],
}


def parse_vtb_cny(data: bytes, account_no: str) -> tuple[pd.DataFrame, int]:
    """
    CNY columns: Дебет CNY, Кредит CNY.
    Returns: (DataFrame, skipped_rows_count)
    """
    df = pd.read_excel(BytesIO(data), header=0)
    col_map = _find_columns(df, _VTB_CNY_COLUMNS)

    df = df.dropna(subset=[col_map["date"]])
    skipped = 0

    rows = []
    for idx, row in df.iterrows():
        try:
            dt = _parse_date(row[col_map["date"]])
        except Exception as e:
            logger.warning("VTB_CNY row %s skipped: date parse error: %s", idx, e)
            skipped += 1
            continue

        debit_cny = _to_decimal(row[col_map["debit"]])
        credit_cny = _to_decimal(row[col_map["credit"]])

        rows.append({
            "date": dt,
            "bank": "VTB",
            "account": account_no,
            "currency": "CNY",
            "counterparty": _clean_str(row[col_map["counterparty"]]),
            "inn": _clean_str(row[col_map["inn"]]),
            "counterparty_account": _clean_str(row[col_map["cp_account"]]),
            "purpose": _clean_str(row[col_map["purpose"]]),
            "income": credit_cny,
            "expense": debit_cny,
        })

    if skipped:
        logger.info("VTB_CNY parsing: %d rows parsed, %d skipped", len(rows), skipped)

    result = pd.DataFrame(rows, columns=NORM_COLS) if rows else pd.DataFrame(columns=NORM_COLS)
    return result, skipped


# ─── WB (MAIN / PAYOUT) parser ───────────────────────────────────────────────

_WB_COLUMNS = {
    "date": ["Дата операции", "Дата"],
    "counterparty": ["Корреспондент", "Наименование"],
    "inn": ["ИНН"],
    "cp_account": ["Счет"],
    "debit": ["Оборот Дт", "Дебет"],
    "credit": ["Оборот Кт", "Кредит"],
    "purpose": ["Назначение платежа", "Назначение"],
}


def _parse_wb(data: bytes, account_no: str, bank_name: str = "WB") -> tuple[pd.DataFrame, int]:
    """
    WB format has header row 0 as column names, row 1 as sub-headers.
    Returns: (DataFrame, skipped_rows_count)
    """
    df = pd.read_excel(BytesIO(data), header=0, skiprows=0)
    # row 0 is the real header, row 1 has sub-headers → skip row 1
    if len(df) > 0 and str(df.iloc[0, 2]).strip() in ("Наименование", "Корреспондент"):
        df = df.iloc[1:].reset_index(drop=True)

    col_map = _find_columns(df, _WB_COLUMNS)
    skipped = 0

    rows = []
    for idx, row in df.iterrows():
        date_val = row[col_map["date"]]
        if date_val is None or (isinstance(date_val, float) and pd.isna(date_val)):
            continue
        try:
            dt = _parse_date(date_val)
        except Exception as e:
            logger.warning("WB row %s skipped: date parse error: %s", idx, e)
            skipped += 1
            continue

        debit = _to_decimal(row[col_map["debit"]])
        credit = _to_decimal(row[col_map["credit"]])

        rows.append({
            "date": dt,
            "bank": bank_name,
            "account": account_no,
            "currency": "RUB",
            "counterparty": _clean_str(row[col_map["counterparty"]]),
            "inn": _clean_str(row[col_map["inn"]]),
            "counterparty_account": _clean_str(row[col_map["cp_account"]]),
            "purpose": _clean_str(row[col_map["purpose"]]),
            "income": credit,
            "expense": debit,
        })

    if skipped:
        logger.info("WB parsing: %d rows parsed, %d skipped", len(rows), skipped)

    result = pd.DataFrame(rows, columns=NORM_COLS) if rows else pd.DataFrame(columns=NORM_COLS)
    return result, skipped


def parse_wb_main(data: bytes, account_no: str) -> tuple[pd.DataFrame, int]:
    return _parse_wb(data, account_no, "WB")


def parse_wb_payout(data: bytes, account_no: str) -> tuple[pd.DataFrame, int]:
    return _parse_wb(data, account_no, "WB")


# ─── WB Payout Cabinet parser ─────────────────────────────────────────────────

def parse_wb_payout_cabinet(data: bytes) -> list[dict]:
    """
    Parse WB seller cabinet payout Excel.
    Supports two formats:
    1. With headers: ID заявки на оплату, Сумма, Валюта, Дата создания, Статус оплаты, Комментарий банка
    2. Without headers: columns A-F positionally mapped (request_id, amount, currency, date, status, comment)
    Returns list of dicts (not DataFrame — this is not a bank statement).
    """
    df = pd.read_excel(BytesIO(data), header=0)

    # Flexible column matching
    col_map: dict[str, str] = {}
    for c in df.columns:
        cl = str(c).strip().lower()
        if "id" in cl and "заявк" in cl:
            col_map["request_id"] = c
        elif cl.startswith("сумм"):
            col_map["amount"] = c
        elif "валют" in cl:
            col_map["currency"] = c
        elif "дата" in cl:
            col_map["created_at"] = c
        elif "статус" in cl:
            col_map["status"] = c
        elif "коммент" in cl:
            col_map["comment"] = c

    # If header matching failed (no headers in the file), re-read without header
    # and use positional mapping: A=request_id, B=amount, C=currency, D=date, E=status, F=comment
    if "request_id" not in col_map or "amount" not in col_map:
        logger.info("WB_CABINET: no header detected, using positional column mapping")
        df = pd.read_excel(BytesIO(data), header=None)
        cols = list(df.columns)
        if len(cols) >= 5:
            col_map = {
                "request_id": cols[0],
                "amount": cols[1],
                "currency": cols[2] if len(cols) > 2 else None,
                "created_at": cols[3] if len(cols) > 3 else None,
                "status": cols[4] if len(cols) > 4 else None,
                "comment": cols[5] if len(cols) > 5 else None,
            }

    results = []
    skipped = 0
    for idx, row in df.iterrows():
        request_id = _clean_str(row.get(col_map.get("request_id")))
        if not request_id:
            continue

        # Parse amount: "800 373.93" → Decimal("800373.93")
        raw_amount = str(row.get(col_map.get("amount"), "0"))
        raw_amount = raw_amount.replace("\xa0", "").replace(" ", "").replace(",", ".")
        amount = _to_decimal(raw_amount)
        if amount <= 0:
            continue

        # Parse date
        raw_date = row.get(col_map.get("created_at"))
        try:
            created_at = _parse_date(raw_date)
        except Exception as e:
            logger.warning("WB_CABINET row %s skipped: date parse error: %s", idx, e)
            skipped += 1
            continue

        raw_status = _clean_str(row.get(col_map.get("status")))

        # Derive status enum
        status = "PENDING"
        if raw_status:
            lower_s = raw_status.lower()
            if "успешно проведена" in lower_s or "оплата" == lower_s.strip():
                status = "RECEIVED"
            elif "поручение" in lower_s:
                status = "TRANSIT"
            elif "обрабатывается" in lower_s:
                status = "PROCESSING"

        results.append({
            "request_id": request_id,
            "amount_rub": amount,
            "currency": "RUB",
            "created_at": created_at,
            "wb_status_raw": raw_status,
            "status": status,
            "bank_comment": _clean_str(row.get(col_map.get("comment"))),
        })

    if skipped:
        logger.info("WB_CABINET parsing: %d rows parsed, %d skipped", len(results), skipped)

    return results


# ─── VTB Multi-account parser ─────────────────────────────────────────────────

def parse_vtb_multi(data: bytes) -> tuple[pd.DataFrame, int]:
    """
    Parse a VTB multi-account statement (one .xlsx file, multiple sheets).

    Each sheet = one bank account. Sheet name = account number.
    Row 2: "Номер счета:" | <account_number> | "Валюта:" | <currency_info>
    Row 7: column headers (Дата, Номер, ..., Дебет/Кредит, Назначение)
    Row 8+: transaction data.

    Auto-detects currency from column headers:
    - "Дебет, CNY" → CNY
    - "Дебет, RUR" or "Дебет RUR" → RUB

    Returns: (combined_dataframe, total_skipped_rows)
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    all_frames: list[pd.DataFrame] = []
    total_skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # Read metadata rows (R1–R7)
        header_rows: list[tuple] = []
        for i, row in enumerate(rows_iter):
            header_rows.append(row)
            if i >= 6:  # R7 (0-indexed=6) is the column header row
                break

        if len(header_rows) < 7:
            logger.warning("VTB_MULTI sheet '%s' has < 7 rows, skipping", sheet_name)
            continue

        # Extract account number from R2 cell B (index 1)
        account_no = str(header_rows[1][1]).strip() if header_rows[1][1] else sheet_name

        # Column headers are in R7 (index 6)
        col_headers = [str(c).strip() if c else "" for c in header_rows[6]]

        # Detect currency from column header names
        is_cny = any("cny" in h.lower() for h in col_headers if h)
        currency = "CNY" if is_cny else "RUB"

        # Determine debit/credit column indices
        debit_col = None
        credit_col = None
        for ci, h in enumerate(col_headers):
            hl = h.lower()
            if "дебет" in hl:
                if is_cny and "cny" in hl:
                    debit_col = ci
                elif not is_cny and debit_col is None:
                    debit_col = ci
            if "кредит" in hl:
                if is_cny and "cny" in hl:
                    credit_col = ci
                elif not is_cny and credit_col is None:
                    credit_col = ci

        # Fallback: first Дебет/Кредит columns
        if debit_col is None or credit_col is None:
            for ci, h in enumerate(col_headers):
                hl = h.lower()
                if "дебет" in hl and debit_col is None:
                    debit_col = ci
                if "кредит" in hl and credit_col is None:
                    credit_col = ci

        if debit_col is None or credit_col is None:
            logger.warning("VTB_MULTI sheet '%s': cannot find Дебет/Кредит columns, skipping", sheet_name)
            continue

        # Find other standard columns by name
        col_idx = {}
        for ci, h in enumerate(col_headers):
            hl = h.lower()
            if hl == "дата":
                col_idx["date"] = ci
            elif "контрагент" in hl and "инн" not in hl and "бик" not in hl and "счет" not in hl and "счёт" not in hl:
                col_idx["counterparty"] = ci
            elif "инн" in hl:
                col_idx["inn"] = ci
            elif ("счет контрагента" in hl or "счёт контрагента" in hl):
                col_idx["cp_account"] = ci
            elif "назначение" in hl:
                col_idx["purpose"] = ci

        if "date" not in col_idx:
            logger.warning("VTB_MULTI sheet '%s': no 'Дата' column, skipping", sheet_name)
            continue

        # Parse data rows (R8+)
        skipped = 0
        rows_data = []
        for row in rows_iter:
            date_val = row[col_idx["date"]] if col_idx["date"] < len(row) else None
            if date_val is None or (isinstance(date_val, float) and pd.isna(date_val)):
                continue

            try:
                dt = _parse_date(date_val)
            except Exception as e:
                logger.warning("VTB_MULTI sheet '%s' skipped row: date error: %s", sheet_name, e)
                skipped += 1
                continue

            debit = _to_decimal(row[debit_col] if debit_col < len(row) else None)
            credit = _to_decimal(row[credit_col] if credit_col < len(row) else None)

            rows_data.append({
                "date": dt,
                "bank": "VTB",
                "account": account_no,
                "currency": currency,
                "counterparty": _clean_str(row[col_idx.get("counterparty", 999)] if col_idx.get("counterparty", 999) < len(row) else None),
                "inn": _clean_str(row[col_idx.get("inn", 999)] if col_idx.get("inn", 999) < len(row) else None),
                "counterparty_account": _clean_str(row[col_idx.get("cp_account", 999)] if col_idx.get("cp_account", 999) < len(row) else None),
                "purpose": _clean_str(row[col_idx.get("purpose", 999)] if col_idx.get("purpose", 999) < len(row) else None),
                "income": credit,
                "expense": debit,
            })

        total_skipped += skipped

        if rows_data:
            sheet_df = pd.DataFrame(rows_data, columns=NORM_COLS)
            all_frames.append(sheet_df)
            logger.info("VTB_MULTI sheet '%s': %d rows, %d skipped (%s %s)",
                        sheet_name, len(rows_data), skipped, account_no, currency)

    wb.close()

    if all_frames:
        result = pd.concat(all_frames, ignore_index=True)
    else:
        result = pd.DataFrame(columns=NORM_COLS)

    logger.info("VTB_MULTI total: %d rows across %d sheets, %d skipped",
                len(result), len(all_frames), total_skipped)
    return result, total_skipped


# ─── WB Multi-account (XML SpreadsheetML) parser ─────────────────────────────

def parse_wb_multi(data: bytes) -> tuple[pd.DataFrame, int]:
    """
    Parse WB multi-account statement in XML SpreadsheetML (.xls) format.

    WB exports often save as XML SpreadsheetML. Multiple accounts appear in
    one sheet, separated by "Выписка по счету <account_no>" header rows.

    Each account section:
      - Header: "Выписка по счету XXXXXXXXXX RUR ..."
      - Column headers: "Документ", "Дата операции", "Корреспондент", ...
      - Sub-headers: "Наименование", "ИНН", "КПП", "Счет", "БИК"
      - Data rows: interleaved 2-row format per transaction
      - Separator: "ИТОГО:" row

    Returns: (combined_dataframe, total_skipped_rows)
    """
    import xml.etree.ElementTree as ET
    import re

    tree = ET.parse(BytesIO(data))
    root = tree.getroot()
    ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}

    all_frames: list[pd.DataFrame] = []
    total_skipped = 0

    for ws in root.findall('.//ss:Worksheet', ns):
        table = ws.find('ss:Table', ns)
        if table is None:
            continue

        xml_rows = table.findall('ss:Row', ns)

        # Convert XML rows to list of cell values
        def _row_cells(row) -> list[str]:
            cells = row.findall('ss:Cell', ns)
            return [
                (c.find('ss:Data', ns).text or "").strip()
                if c.find('ss:Data', ns) is not None else ""
                for c in cells
            ]

        # Find all "Выписка по счету" header positions
        account_sections: list[tuple[str, int]] = []
        for i, row in enumerate(xml_rows):
            cells = _row_cells(row)
            if cells and "Выписка по счету" in cells[0]:
                # Extract account number: "Выписка по счету 40702810500001001752 RUR ..."
                match = re.search(r'Выписка по счету\s+(\d+)', cells[0])
                if match:
                    account_sections.append((match.group(1), i))

        if not account_sections:
            logger.warning("WB_MULTI: no 'Выписка по счету' found")
            continue

        # Process each account section
        for sec_idx, (account_no, start_row) in enumerate(account_sections):
            # End of section: next account start or end of sheet
            end_row = (
                account_sections[sec_idx + 1][1]
                if sec_idx + 1 < len(account_sections)
                else len(xml_rows)
            )

            # Find column header row within section (contains "Документ" or "Дата операции")
            col_header_idx = None
            for i in range(start_row, min(start_row + 15, end_row)):
                cells = _row_cells(xml_rows[i])
                if any("Дата операции" in c for c in cells):
                    col_header_idx = i
                    break

            if col_header_idx is None:
                logger.warning("WB_MULTI account %s: no column headers found", account_no)
                continue

            # Build column index map from header row
            headers = _row_cells(xml_rows[col_header_idx])
            col_idx: dict[str, int] = {}
            for ci, h in enumerate(headers):
                hl = h.lower().strip()
                if "дата" in hl:
                    col_idx["date"] = ci
                elif hl == "корреспондент" or hl == "наименование":
                    if "counterparty" not in col_idx:
                        col_idx["counterparty"] = ci
                elif "оборот дт" in hl or "дебет" in hl:
                    col_idx["debit"] = ci
                elif "оборот кт" in hl or "кредит" in hl:
                    col_idx["credit"] = ci
                elif "назначение" in hl:
                    col_idx["purpose"] = ci

            if "date" not in col_idx:
                logger.warning("WB_MULTI account %s: no date column", account_no)
                continue

            # Sub-header row: "Наименование", "ИНН", "КПП", "Счет", "БИК"
            sub_header_idx = col_header_idx + 1
            sub_headers = _row_cells(xml_rows[sub_header_idx]) if sub_header_idx < end_row else []
            sub_col_idx: dict[str, int] = {}
            for ci, h in enumerate(sub_headers):
                hl = h.lower().strip()
                if "инн" == hl:
                    sub_col_idx["inn"] = ci
                elif "счет" in hl or "счёт" in hl:
                    sub_col_idx["cp_account"] = ci

            # Parse data rows (after sub-header)
            data_start = sub_header_idx + 1
            skipped = 0
            rows_data = []

            i = data_start
            while i < end_row:
                cells = _row_cells(xml_rows[i])
                if not cells:
                    i += 1
                    continue

                # Check for ИТОГО or end markers
                if cells[0].startswith("ИТОГО"):
                    break

                # WB format: main data row has >= 7 cells
                # and has a date-like string in the date column
                date_ci = col_idx.get("date", 1)
                date_val = cells[date_ci] if date_ci < len(cells) else ""
                if not date_val or date_val in ("", "Дата операции"):
                    i += 1
                    continue

                try:
                    dt = _parse_date(date_val)
                except Exception:
                    i += 1
                    skipped += 1
                    continue

                # Extract from main row
                debit_ci = col_idx.get("debit", 4)
                credit_ci = col_idx.get("credit", 5)
                purpose_ci = col_idx.get("purpose", 6)
                cp_ci = col_idx.get("counterparty", 2)

                debit = _to_decimal(cells[debit_ci] if debit_ci < len(cells) else None)
                credit = _to_decimal(cells[credit_ci] if credit_ci < len(cells) else None)
                counterparty = _clean_str(cells[cp_ci] if cp_ci < len(cells) else None)
                purpose = _clean_str(cells[purpose_ci] if purpose_ci < len(cells) else None)

                # WB interleaved format: the counterparty name in main data row
                # is at index 2, but INN is in the same row at index 3
                # In the actual file, data rows have 11 cells:
                # [doc, date, cp, inn, kpp, account, bik, balance, debit, credit, purpose]
                inn = None
                cp_account = None
                if len(cells) > 5:
                    # Real data row: [doc, date, cp, inn, kpp, account, bik, balance, debit, credit, purpose]
                    inn = _clean_str(cells[3] if 3 < len(cells) else None)
                    cp_account = _clean_str(cells[5] if 5 < len(cells) else None)
                    # In this wide format, debit/credit are at positions 8,9 and purpose at 10
                    if len(cells) >= 11:
                        debit = _to_decimal(cells[8] if cells[8] else None)
                        credit = _to_decimal(cells[9] if cells[9] else None)
                        purpose = _clean_str(cells[10] if 10 < len(cells) else None)
                        counterparty = _clean_str(cells[2] if 2 < len(cells) else None)

                rows_data.append({
                    "date": dt,
                    "bank": "WB",
                    "account": account_no,
                    "currency": "RUB",
                    "counterparty": counterparty,
                    "inn": inn,
                    "counterparty_account": cp_account,
                    "purpose": purpose,
                    "income": credit,
                    "expense": debit,
                })

                i += 1

            total_skipped += skipped

            if rows_data:
                sheet_df = pd.DataFrame(rows_data, columns=NORM_COLS)
                all_frames.append(sheet_df)
                logger.info("WB_MULTI account '%s': %d rows, %d skipped",
                            account_no, len(rows_data), skipped)

    if all_frames:
        result = pd.concat(all_frames, ignore_index=True)
    else:
        result = pd.DataFrame(columns=NORM_COLS)

    logger.info("WB_MULTI total: %d rows across %d accounts, %d skipped",
                len(result), len(all_frames), total_skipped)
    return result, total_skipped


# ─── Registry ─────────────────────────────────────────────────────────────────

SOURCE_PARSERS = {
    "VTB_RUB_MAIN": lambda data, acc: parse_vtb_rub(data, "VTB_RUB_MAIN", acc),
    "VTB_RUB_TRANSIT": lambda data, acc: parse_vtb_rub(data, "VTB_RUB_TRANSIT", acc),
    "VTB_CNY": lambda data, acc: parse_vtb_cny(data, acc),
    "VTB_MULTI": lambda data, acc: parse_vtb_multi(data),
    "WB_MAIN": parse_wb_main,
    "WB_PAYOUT": parse_wb_payout,
    "WB_MULTI": lambda data, acc: parse_wb_multi(data),
}


def parse_statement(source_type: str, data: bytes, account_no: str) -> tuple[pd.DataFrame, int]:
    """
    Parse a bank statement file.
    Returns: (normalized_dataframe, skipped_rows_count)
    """
    parser = SOURCE_PARSERS.get(source_type)
    if not parser:
        raise ValueError(f"Unknown source_type: {source_type}")
    return parser(data, account_no)

