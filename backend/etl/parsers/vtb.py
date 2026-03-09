"""
VTB bank statement parsers: RUB, CNY, Multi-account.
"""

import logging
from io import BytesIO

import pandas as pd

from backend.etl.parsers.helpers import (
    NORM_COLS, _clean_str, _find_columns, _parse_date, _to_decimal,
)

logger = logging.getLogger(__name__)


# ─── VTB RUB parser (MAIN / TRANSIT) ─────────────────────────────────────────

_VTB_RUB_COLUMNS = {
    "date": ["Дата"],
    "counterparty": ["Контрагент"],
    "inn": ["ИНН"],
    "cp_account": ["Счет контрагента", "Счёт контрагента"],
    "debit": ["Дебет"],
    "credit": ["Кредит"],
    "purpose": ["Назначение"],
}


def parse_vtb_rub(data: bytes, source_type: str, account_no: str) -> tuple[pd.DataFrame, int]:
    """
    Columns: Дата, Номер, Вид операции, Контрагент, ИНН контрагента,
             БИК банка контрагента, Счет контрагента, Дебет RUR, Кредит RUR, Назначение
    Debit  = expense (money out), Credit = income (money in).

    Returns: (DataFrame, skipped_rows_count)
    """
    df = pd.read_excel(BytesIO(data), header=0)
    col_map = _find_columns(df, _VTB_RUB_COLUMNS)

    # drop rows where date is NaN
    df = df.dropna(subset=[col_map["date"]])

    bank = "VTB"
    currency = "RUB"
    skipped = 0

    rows = []
    for idx, row in df.iterrows():
        try:
            dt = _parse_date(row[col_map["date"]])
        except Exception as e:
            logger.warning("VTB_RUB row %s skipped: date parse error: %s", idx, e)
            skipped += 1
            continue

        debit = _to_decimal(row[col_map["debit"]])
        credit = _to_decimal(row[col_map["credit"]])

        rows.append({
            "date": dt,
            "bank": bank,
            "account": account_no,
            "currency": currency,
            "counterparty": _clean_str(row[col_map["counterparty"]]),
            "inn": _clean_str(row[col_map["inn"]]),
            "counterparty_account": _clean_str(row[col_map["cp_account"]]),
            "purpose": _clean_str(row[col_map["purpose"]]),
            "income": credit,
            "expense": debit,
        })

    if skipped:
        logger.info("VTB_RUB parsing: %d rows parsed, %d skipped", len(rows), skipped)

    result = pd.DataFrame(rows, columns=NORM_COLS) if rows else pd.DataFrame(columns=NORM_COLS)
    return result, skipped


# ─── VTB CNY parser ───────────────────────────────────────────────────────────

_VTB_CNY_COLUMNS = {
    "date": ["Дата"],
    "counterparty": ["Контрагент"],
    "inn": ["ИНН"],
    "cp_account": ["Счет контрагента", "Счёт контрагента"],
    "debit": ["Дебет CNY", "Дебет юан", "Дебет"],
    "credit": ["Кредит CNY", "Кредит юан", "Кредит"],
    "purpose": ["Назначение"],
}


def parse_vtb_cny(data: bytes, account_no: str) -> tuple[pd.DataFrame, int]:
    """
    CNY columns: Дебет CNY, Кредит CNY.
    Returns: (DataFrame, skipped_rows_count)
    """
    df = pd.read_excel(BytesIO(data), header=0)
    col_map = _find_columns(df, _VTB_CNY_COLUMNS)

    df = df.dropna(subset=[col_map["date"]])
    skipped = 0

    rows = []
    for idx, row in df.iterrows():
        try:
            dt = _parse_date(row[col_map["date"]])
        except Exception as e:
            logger.warning("VTB_CNY row %s skipped: date parse error: %s", idx, e)
            skipped += 1
            continue

        debit_cny = _to_decimal(row[col_map["debit"]])
        credit_cny = _to_decimal(row[col_map["credit"]])

        rows.append({
            "date": dt,
            "bank": "VTB",
            "account": account_no,
            "currency": "CNY",
            "counterparty": _clean_str(row[col_map["counterparty"]]),
            "inn": _clean_str(row[col_map["inn"]]),
            "counterparty_account": _clean_str(row[col_map["cp_account"]]),
            "purpose": _clean_str(row[col_map["purpose"]]),
            "income": credit_cny,
            "expense": debit_cny,
        })

    if skipped:
        logger.info("VTB_CNY parsing: %d rows parsed, %d skipped", len(rows), skipped)

    result = pd.DataFrame(rows, columns=NORM_COLS) if rows else pd.DataFrame(columns=NORM_COLS)
    return result, skipped


# ─── VTB Multi-account parser ─────────────────────────────────────────────────

def parse_vtb_multi(data: bytes) -> tuple[pd.DataFrame, int]:
    """
    Parse a VTB multi-account statement (one .xlsx file, multiple sheets).

    Each sheet = one bank account. Sheet name = account number.
    Row 2: "Номер счета:" | <account_number> | "Валюта:" | <currency_info>
    Row 7: column headers (Дата, Номер, ..., Дебет/Кредит, Назначение)
    Row 8+: transaction data.

    Auto-detects currency from column headers:
    - "Дебет, CNY" → CNY
    - "Дебет, RUR" or "Дебет RUR" → RUB

    Returns: (combined_dataframe, total_skipped_rows)
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    all_frames: list[pd.DataFrame] = []
    total_skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # Read metadata rows (R1–R7)
        header_rows: list[tuple] = []
        for i, row in enumerate(rows_iter):
            header_rows.append(row)
            if i >= 6:  # R7 (0-indexed=6) is the column header row
                break

        if len(header_rows) < 7:
            logger.warning("VTB_MULTI sheet '%s' has < 7 rows, skipping", sheet_name)
            continue

        # Extract account number from R2 cell B (index 1)
        account_no = str(header_rows[1][1]).strip() if header_rows[1][1] else sheet_name

        # Column headers are in R7 (index 6)
        col_headers = [str(c).strip() if c else "" for c in header_rows[6]]

        # Detect currency from column header names
        is_cny = any("cny" in h.lower() for h in col_headers if h)
        currency = "CNY" if is_cny else "RUB"

        # Determine debit/credit column indices
        debit_col = None
        credit_col = None
        for ci, h in enumerate(col_headers):
            hl = h.lower()
            if "дебет" in hl:
                if is_cny and "cny" in hl:
                    debit_col = ci
                elif not is_cny and debit_col is None:
                    debit_col = ci
            if "кредит" in hl:
                if is_cny and "cny" in hl:
                    credit_col = ci
                elif not is_cny and credit_col is None:
                    credit_col = ci

        # Fallback: first Дебет/Кредит columns
        if debit_col is None or credit_col is None:
            for ci, h in enumerate(col_headers):
                hl = h.lower()
                if "дебет" in hl and debit_col is None:
                    debit_col = ci
                if "кредит" in hl and credit_col is None:
                    credit_col = ci

        if debit_col is None or credit_col is None:
            logger.warning("VTB_MULTI sheet '%s': cannot find Дебет/Кредит columns, skipping", sheet_name)
            continue

        # Find other standard columns by name
        col_idx = {}
        for ci, h in enumerate(col_headers):
            hl = h.lower()
            if hl == "дата":
                col_idx["date"] = ci
            elif "контрагент" in hl and "инн" not in hl and "бик" not in hl and "счет" not in hl and "счёт" not in hl:
                col_idx["counterparty"] = ci
            elif "инн" in hl:
                col_idx["inn"] = ci
            elif ("счет контрагента" in hl or "счёт контрагента" in hl):
                col_idx["cp_account"] = ci
            elif "назначение" in hl:
                col_idx["purpose"] = ci

        if "date" not in col_idx:
            logger.warning("VTB_MULTI sheet '%s': no 'Дата' column, skipping", sheet_name)
            continue

        # Parse data rows (R8+)
        skipped = 0
        rows_data = []
        for row in rows_iter:
            date_val = row[col_idx["date"]] if col_idx["date"] < len(row) else None
            if date_val is None or (isinstance(date_val, float) and pd.isna(date_val)):
                continue

            try:
                dt = _parse_date(date_val)
            except Exception as e:
                logger.warning("VTB_MULTI sheet '%s' skipped row: date error: %s", sheet_name, e)
                skipped += 1
                continue

            debit = _to_decimal(row[debit_col] if debit_col < len(row) else None)
            credit = _to_decimal(row[credit_col] if credit_col < len(row) else None)

            rows_data.append({
                "date": dt,
                "bank": "VTB",
                "account": account_no,
                "currency": currency,
                "counterparty": _clean_str(row[col_idx.get("counterparty", 999)] if col_idx.get("counterparty", 999) < len(row) else None),
                "inn": _clean_str(row[col_idx.get("inn", 999)] if col_idx.get("inn", 999) < len(row) else None),
                "counterparty_account": _clean_str(row[col_idx.get("cp_account", 999)] if col_idx.get("cp_account", 999) < len(row) else None),
                "purpose": _clean_str(row[col_idx.get("purpose", 999)] if col_idx.get("purpose", 999) < len(row) else None),
                "income": credit,
                "expense": debit,
            })

        total_skipped += skipped

        if rows_data:
            sheet_df = pd.DataFrame(rows_data, columns=NORM_COLS)
            all_frames.append(sheet_df)
            logger.info("VTB_MULTI sheet '%s': %d rows, %d skipped (%s %s)",
                        sheet_name, len(rows_data), skipped, account_no, currency)

    wb.close()

    if all_frames:
        result = pd.concat(all_frames, ignore_index=True)
    else:
        result = pd.DataFrame(columns=NORM_COLS)

    logger.info("VTB_MULTI total: %d rows across %d sheets, %d skipped",
                len(result), len(all_frames), total_skipped)
    return result, total_skipped
