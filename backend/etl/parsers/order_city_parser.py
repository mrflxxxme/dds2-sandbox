"""Parser for WB order feed Excel (Лента заказов) — extracts srid → city + okrug + order_date mapping."""

import logging
from datetime import date, datetime
from io import BytesIO

import openpyxl

logger = logging.getLogger("dds.parsers.order_city")


def _detect_header_row(ws) -> tuple[int, int, int, int, int]:
    """Scan rows 1-10 looking for header columns.

    Returns:
        (data_start_row, srid_col, city_col, okrug_col, date_col) — all 1-based.
        0 means column not found.
    """
    for row_num in range(1, min(11, ws.max_row + 1)):
        srid_col = None
        city_col = None
        okrug_col = None
        date_col = None

        for col_num in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=row_num, column=col_num).value
            if cell_val is None:
                continue
            text = str(cell_val).strip().lower()

            if "id заказа" in text:
                srid_col = col_num
            elif "город прибытия" in text or text == "город":
                city_col = col_num
            elif "округ прибытия" in text or "регион прибытия" in text:
                okrug_col = col_num
            elif "дата оформления" in text or "дата заказа" in text:
                date_col = col_num

        if srid_col is not None and city_col is not None:
            if okrug_col is None:
                logger.warning(f"Header row {row_num}: found srid+city but no okrug column, okrug will be None")
            logger.info(
                f"Detected header row={row_num}, srid_col={srid_col}, "
                f"city_col={city_col}, okrug_col={okrug_col}, date_col={date_col}"
            )
            return (row_num + 1, srid_col, city_col, okrug_col or 0, date_col or 0)

    # Fallback: legacy fixed positions
    logger.warning("Header row not detected, using fallback column positions")
    return (2, 21, 16, 15, 8)


def _parse_date(val) -> date | None:
    """Try to extract a date from cell value."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    text = str(val).strip()
    if not text:
        return None
    # Try common formats
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%dT%H:%M:%S", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(text[: len(fmt) + 2], fmt).date()
        except (ValueError, IndexError):  # noqa: S112
            continue
    return None


def parse_order_city_excel(data: bytes) -> list[dict]:
    """Parse WB order feed Excel, extract srid→city+okrug+order_date mapping.

    WB "Лента заказов" Excel structure:
      Auto-detects header row by scanning rows 1-10 for known column names.
      Fallback: Row 1 = merged title (skip), data from row 2,
      Column O (15) = округ, Column P (16) = город, Column U (21) = srid, Column H (8) = дата.

    Returns:
        list of {srid: str, city: str, okrug: str | None, order_date: date | None}
    """
    wb = openpyxl.load_workbook(BytesIO(data), read_only=False, data_only=True)

    # Find sheet 'Заказы'
    sheet_name = None
    for name in wb.sheetnames:
        if "заказ" in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        logger.warning("Sheet 'Заказы' not found in Excel")
        return []

    ws = wb[sheet_name]

    data_start, srid_col, city_col, okrug_col, date_col = _detect_header_row(ws)

    results = []
    row_count = 0

    for row_num in range(data_start, ws.max_row + 1):
        row_count += 1
        srid_val = ws.cell(row=row_num, column=srid_col).value
        city_val = ws.cell(row=row_num, column=city_col).value
        okrug_val = ws.cell(row=row_num, column=okrug_col).value if okrug_col else None
        date_val = ws.cell(row=row_num, column=date_col).value if date_col else None

        if not srid_val or not city_val:
            continue

        srid_str = str(srid_val).strip()
        city_str = str(city_val).strip()
        okrug_str = str(okrug_val).strip() if okrug_val else None

        if not srid_str or not city_str or city_str == "-":
            continue

        results.append(
            {
                "srid": srid_str,
                "city": city_str,
                "okrug": okrug_str,
                "order_date": _parse_date(date_val),
            }
        )

    logger.info(f"Parsed order_city Excel: {len(results)} mappings from {row_count} rows")
    wb.close()
    return results
