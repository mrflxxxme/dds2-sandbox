"""
AI Chat router — web interface for the AI agent system.
Provides conversation CRUD, message streaming (SSE), and file upload.
"""

from __future__ import annotations

import asyncio
import base64
import io
import json
import logging

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from backend.auth import get_current_user
from backend.config import settings
from backend.database import AsyncSessionLocal, get_db
from backend.models.auth import Project, User
from backend.project_context import get_current_project
from backend.schemas.ai_chat import (
    AiConversationCreate,
    AiConversationList,
    AiConversationSchema,
    AiConversationUpdate,
    AiFileUploadResponse,
    AiMessageCreate,
    AiMessageSchema,
)
from backend.services import ai_chat_service
from backend.services.ai.orchestrator import ask as orchestrator_ask
from backend.utils.rate_limit import rate_limit_write

logger = logging.getLogger("dds.routers.ai_chat")

router = APIRouter(prefix="/ai", tags=["AI Chat"])

# ── Conversations CRUD ───────────────────────────────────────────────


@router.post("/conversations", response_model=AiConversationSchema)
async def create_conversation(
    body: AiConversationCreate,
    project: Project = Depends(get_current_project),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rl: None = Depends(rate_limit_write),
):
    """Create a new AI chat conversation."""
    return await ai_chat_service.create_conversation(
        db,
        project_id=project.id,
        user_id=user.id,
        brand=body.brand,
        title=body.title,
    )


@router.get("/conversations", response_model=list[AiConversationList])
async def list_conversations(
    limit: int = 50,
    offset: int = 0,
    project: Project = Depends(get_current_project),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List conversations for sidebar (newest first)."""
    return await ai_chat_service.list_conversations(
        db,
        project_id=project.id,
        user_id=user.id,
        limit=min(limit, 100),
        offset=offset,
    )


@router.get("/conversations/{conversation_id}", response_model=AiConversationSchema)
async def get_conversation(
    conversation_id: int,
    project: Project = Depends(get_current_project),
    _user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get a conversation by ID."""
    conv = await ai_chat_service.get_conversation(db, project.id, conversation_id)
    if not conv:
        raise HTTPException(status_code=404, detail="Чат не найден")
    return conv


@router.patch("/conversations/{conversation_id}", response_model=AiConversationSchema)
async def update_conversation(
    conversation_id: int,
    body: AiConversationUpdate,
    project: Project = Depends(get_current_project),
    _user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rl: None = Depends(rate_limit_write),
):
    """Rename or change brand of a conversation."""
    conv = await ai_chat_service.update_conversation(
        db, project.id, conversation_id, title=body.title, brand=body.brand
    )
    if not conv:
        raise HTTPException(status_code=404, detail="Чат не найден")
    return conv


@router.delete("/conversations/{conversation_id}")
async def delete_conversation(
    conversation_id: int,
    project: Project = Depends(get_current_project),
    _user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rl: None = Depends(rate_limit_write),
):
    """Soft-delete a conversation."""
    ok = await ai_chat_service.delete_conversation(db, project.id, conversation_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Чат не найден")
    return {"message": "Чат удалён"}


# ── Messages ─────────────────────────────────────────────────────────


@router.get(
    "/conversations/{conversation_id}/messages",
    response_model=list[AiMessageSchema],
)
async def list_messages(
    conversation_id: int,
    limit: int = 100,
    offset: int = 0,
    project: Project = Depends(get_current_project),
    _user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List messages in a conversation."""
    # Verify conversation exists and belongs to project
    conv = await ai_chat_service.get_conversation(db, project.id, conversation_id)
    if not conv:
        raise HTTPException(status_code=404, detail="Чат не найден")
    return await ai_chat_service.list_messages(db, conversation_id, limit=min(limit, 500), offset=offset)


@router.post("/conversations/{conversation_id}/messages")
async def send_message(
    conversation_id: int,
    body: AiMessageCreate,
    project: Project = Depends(get_current_project),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    _rl: None = Depends(rate_limit_write),
):
    """Send a message and get AI response via SSE stream.

    SSE events:
    - data: {"type": "token", "content": "word "}
    - data: {"type": "done", "message": {...}}
    """
    # Verify conversation exists and belongs to project
    conv = await ai_chat_service.get_conversation(db, project.id, conversation_id)
    if not conv:
        raise HTTPException(status_code=404, detail="Чат не найден")

    # Save user message
    await ai_chat_service.save_message(
        db,
        conversation_id=conversation_id,
        role="user",
        content=body.content,
        files=body.files,
    )

    # Auto-title from first message
    await ai_chat_service.auto_title(db, conversation_id, body.content)

    # Determine brand: message override > conversation brand
    brand = body.brand or conv.brand

    # Capture values needed in stream (db session from Depends will be closed)
    project_id = project.id
    tax_rate = float(project.tax_rate or 6)
    question = body.content

    async def event_stream():
        """Generate SSE events: stream AI response word by word, then send final message.

        Uses its own DB session because FastAPI closes the Depends(get_db) session
        before the StreamingResponse generator runs.
        """
        async with AsyncSessionLocal() as stream_db:
            try:
                # Get full AI response with dedicated session
                answer = await orchestrator_ask(
                    db=stream_db,
                    project_id=project_id,
                    brand=brand,
                    chat_id=conversation_id,
                    question=question,
                    tax_rate=tax_rate,
                )

                # Fake streaming: yield word by word
                words = answer.split(" ")
                for i, word in enumerate(words):
                    token = word if i == len(words) - 1 else word + " "
                    event = json.dumps({"type": "token", "content": token}, ensure_ascii=False)
                    yield f"data: {event}\n\n"
                    await asyncio.sleep(0.02)

                # Save assistant message
                assistant_msg = await ai_chat_service.save_message(
                    stream_db,
                    conversation_id=conversation_id,
                    role="assistant",
                    content=answer,
                )

                # Send final "done" event with full message
                done_data = {
                    "type": "done",
                    "message": {
                        "id": assistant_msg.id,
                        "conversation_id": assistant_msg.conversation_id,
                        "role": assistant_msg.role,
                        "content": assistant_msg.content,
                        "tokens_used": assistant_msg.tokens_used,
                        "tools_used": assistant_msg.tools_used,
                        "created_at": assistant_msg.created_at.isoformat(),
                    },
                }
                yield f"data: {json.dumps(done_data, ensure_ascii=False)}\n\n"

            except Exception:
                logger.exception("AI chat error for conversation=%s", conversation_id)
                error_event = json.dumps(
                    {"type": "error", "content": "Ошибка при обработке запроса. Попробуйте позже."},
                    ensure_ascii=False,
                )
                yield f"data: {error_event}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# ── File Upload ──────────────────────────────────────────────────────

_ALLOWED_EXCEL = {".xlsx", ".xls"}
_ALLOWED_IMAGES = {".png", ".jpg", ".jpeg", ".webp"}


@router.post("/upload", response_model=AiFileUploadResponse)
async def upload_file(
    file: UploadFile,
    _project: Project = Depends(get_current_project),
    _user: User = Depends(get_current_user),
    _rl: None = Depends(rate_limit_write),
):
    """Upload a file for AI chat context.

    - Excel (.xlsx, .xls): converted to text representation
    - Images (.png, .jpg, .jpeg, .webp): converted to base64
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Имя файла не указано")

    # Determine file extension
    ext = "." + file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in _ALLOWED_EXCEL | _ALLOWED_IMAGES:
        raise HTTPException(
            status_code=400,
            detail=f"Неподдерживаемый формат. Допустимые: {', '.join(sorted(_ALLOWED_EXCEL | _ALLOWED_IMAGES))}",
        )

    # Read and check size
    data = await file.read()
    max_bytes = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024
    if len(data) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"Файл слишком большой. Максимум: {settings.MAX_UPLOAD_SIZE_MB} МБ",
        )

    # Process based on type
    if ext in _ALLOWED_EXCEL:
        content = _excel_to_text(data, ext)
        file_type = "excel"
    else:
        content = base64.b64encode(data).decode("ascii")
        file_type = "image"

    return AiFileUploadResponse(
        name=file.filename,
        type=file_type,
        size=len(data),
        content=content,
    )


def _excel_to_text(data: bytes, ext: str) -> str:
    """Convert Excel file bytes to text representation using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"=== Лист: {sheet_name} ===")
        for row_count, row in enumerate(ws.iter_rows(values_only=True)):
            if row_count >= 500:
                parts.append(f"... (обрезано, всего строк: {ws.max_row})")
                break
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("\t".join(cells))

    wb.close()
    return "\n".join(parts)
